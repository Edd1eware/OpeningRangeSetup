import argparse
import csv
import ctypes
import json
import os
import re
import sys
import time
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook

import replay_sync_runner_common_after_sync as replay_sync
import telegram_run_summary_after_sync as telegram


PROJECT_DIR = Path(__file__).resolve().parent
CONTEXT_FOLDER = Path(r"C:\Users\k_99_\Desktop\Codex_cotexto")
VISUAL_TESTS_FOLDER = replay_sync.RESULTS_FOLDER / "visual_tests"
CONFLICT_OUTPUT_FOLDER = VISUAL_TESTS_FOLDER / "test_fechas_conflictivas_runs"
LADDER_OUTPUT_ROOT = VISUAL_TESTS_FOLDER / "sync_ladder_runs"
STATE_PATH = PROJECT_DIR / "sync_ladder_state.json"
ARCHIVE_PREFIX = "SYNC_PASS"
# Meta previa a la busqueda del edge: 500 sesiones recolectadas (con o sin trade).
# Recien al llegar a 500 empieza la optimizacion de WR 87.2% / PF 9.77.
SESSION_TARGET = 500
PRIMARY_OBJECTIVE = (
    "Objetivo principal: acercarnos a WR 87.2% y PF 9.77, "
    "sin matar la operativa o afectandola lo menos posible."
)

CONFLICTING_DATES = [
    "2025-07-16",
    "2025-03-14",
    "2025-03-26",
    "2025-04-17",
    "2025-05-12",
    "2025-05-21",
    "2025-06-23",
]


class WindowsSleepBlocker:
    ES_CONTINUOUS = 0x80000000
    ES_SYSTEM_REQUIRED = 0x00000001
    ES_DISPLAY_REQUIRED = 0x00000002

    def __init__(self, keep_display_on=True):
        self.keep_display_on = keep_display_on
        self.enabled = False

    def __enter__(self):
        if os.name != "nt":
            return self

        flags = self.ES_CONTINUOUS | self.ES_SYSTEM_REQUIRED
        if self.keep_display_on:
            flags |= self.ES_DISPLAY_REQUIRED

        result = ctypes.windll.kernel32.SetThreadExecutionState(flags)
        self.enabled = bool(result)
        if self.enabled:
            print("Anti-suspension activo mientras este Python siga corriendo.")
        else:
            print("WARNING: Windows no confirmo el bloqueo de suspension.")
        return self

    def refresh(self):
        if self.enabled:
            self.__enter__()

    def __exit__(self, exc_type, exc, tb):
        if os.name == "nt" and self.enabled:
            ctypes.windll.kernel32.SetThreadExecutionState(self.ES_CONTINUOUS)
            print("Anti-suspension liberado.")


def nth_weekday(year, month, weekday, n):
    current = date(year, month, 1)
    days_until = (weekday - current.weekday()) % 7
    return current + timedelta(days=days_until + 7 * (n - 1))


def last_weekday(year, month, weekday):
    if month == 12:
        current = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        current = date(year, month + 1, 1) - timedelta(days=1)
    return current - timedelta(days=(current.weekday() - weekday) % 7)


def observed_market_holiday(day):
    if day.weekday() == 5:
        return day - timedelta(days=1)
    if day.weekday() == 6:
        return day + timedelta(days=1)
    return day


def easter_date(year):
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def market_closed_dates_for_year(year):
    holidays = {
        observed_market_holiday(date(year, 1, 1)),
        nth_weekday(year, 1, 0, 3),
        nth_weekday(year, 2, 0, 3),
        easter_date(year) - timedelta(days=2),
        last_weekday(year, 5, 0),
        observed_market_holiday(date(year, 6, 19)),
        observed_market_holiday(date(year, 7, 4)),
        nth_weekday(year, 9, 0, 1),
        nth_weekday(year, 11, 3, 4),
        observed_market_holiday(date(year, 12, 25)),
    }
    if year == 2025:
        holidays.add(date(2025, 1, 9))
    return holidays


def market_closed_dates(start_date, end_date):
    closed = set()
    for year in range(start_date.year - 1, end_date.year + 2):
        closed.update(market_closed_dates_for_year(year))
    return closed


def ny_dst_trading_window(year):
    # ATAS Replay está configurado en UTC-4. Para evitar timeouts largos en
    # invierno, la escalera valida solo sesiones NY dentro de DST.
    dst_starts = nth_weekday(year, 3, 6, 2)
    dst_ends = nth_weekday(year, 11, 6, 1)
    first_trading_day = dst_starts + timedelta(days=1)
    last_trading_day = dst_ends - timedelta(days=2)
    return first_trading_day, last_trading_day


def is_ny_dst_trading_day(value):
    first_trading_day, last_trading_day = ny_dst_trading_window(value.year)
    return first_trading_day <= value <= last_trading_day


def add_months(value, months):
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    last_day = (next_month - timedelta(days=1)).day
    return date(year, month, min(value.day, last_day))


def trading_date_isos(start_date, end_date, dst_only=True):
    if start_date > end_date:
        return []

    closed = market_closed_dates(start_date, end_date)
    dates = []
    current = start_date
    while current <= end_date:
        if (
            current.weekday() < 5
            and current not in closed
            and (not dst_only or is_ny_dst_trading_day(current))
        ):
            dates.append(current.isoformat())
        current += timedelta(days=1)
    return dates


def sanitize_key(value):
    value = value.lower()
    value = re.sub(r"[^a-z0-9_]+", "_", value)
    return re.sub(r"_+", "_", value).strip("_")


def build_stage(index, key, label, start_date, end_date, date_isos, output_folder, report_prefix, run_id):
    return {
        "index": index,
        "key": key,
        "label": label,
        "run_id": run_id,
        "start_date": start_date.isoformat() if start_date else "",
        "end_date": end_date.isoformat() if end_date else "",
        "dates": date_isos,
        "output_folder": output_folder,
        "report_prefix": report_prefix,
    }


def build_ladder_stages(until_date, run_id, include_one_month=False):
    today_ny = datetime.now(ZoneInfo("America/New_York")).date()
    last_replay_date = today_ny - timedelta(days=1)
    stages = []
    ladder_run_root = LADDER_OUTPUT_ROOT / run_id

    stages.append(
        build_stage(
            1,
            "01_conflictivas",
            "Fechas conflictivas canonicas",
            None,
            None,
            CONFLICTING_DATES,
            CONFLICT_OUTPUT_FOLDER / run_id / "01_conflictivas",
            "test_fechas_conflictivas",
            run_id,
        )
    )

    next_index = 2
    if include_one_month:
        start_date = add_months(last_replay_date, -1) + timedelta(days=1)
        key = f"{next_index:02d}_current_1m_{start_date:%Y%m%d}_{last_replay_date:%Y%m%d}"
        stages.append(
            build_stage(
                next_index,
                key,
                "Smoke actual 1 mes",
                start_date,
                last_replay_date,
                trading_date_isos(start_date, last_replay_date),
                ladder_run_root / key,
                key,
                run_id,
            )
        )
        next_index += 1

    # Bloques de ~6 meses hacia atras, SIEMPRE dentro de una sola temporada DST.
    # Nunca cruzan el cambio de horario (invierno EST): ATAS Replay esta en UTC-4,
    # asi que fuera de DST las sesiones NY de las 09:30 quedarian corridas 1 hora.
    cursor_end = last_replay_date
    while cursor_end >= until_date:
        season_first, season_last = ny_dst_trading_window(cursor_end.year)

        if cursor_end > season_last:
            # Despues del fin de DST: bajar al ultimo dia DST de la temporada.
            cursor_end = season_last
            continue

        if cursor_end < season_first:
            # Invierno: saltar a la ultima sesion DST de la temporada anterior.
            cursor_end = ny_dst_trading_window(cursor_end.year - 1)[1]
            continue

        season_floor = max(season_first, until_date)
        block_start = max(season_floor, add_months(cursor_end, -6) + timedelta(days=1))
        key = f"{next_index:02d}_dst_6m_{block_start:%Y%m%d}_{cursor_end:%Y%m%d}"
        stages.append(
            build_stage(
                next_index,
                key,
                "Bloque DST 6 meses",
                block_start,
                cursor_end,
                trading_date_isos(block_start, cursor_end),
                ladder_run_root / key,
                key,
                run_id,
            )
        )
        next_index += 1
        cursor_end = block_start - timedelta(days=1)

    return stages


# Plan completo del ladder (las 9 etapas), seteado en main() antes de filtrar.
# Es informativo: el ladder real lo re-deriva build_ladder_stages cada corrida.
LADDER_PLAN_STAGES = []


def stage_saved_count(stage):
    """Cuenta CSVs X1 ya guardados de una etapa (para el plan en el .json)."""
    x1_folder = Path(stage["output_folder"]) / "X1_R1"
    if not x1_folder.exists():
        return 0
    return len(list(x1_folder.glob("score_trade_result_*_NY.csv")))


def build_ladder_plan(stages):
    plan = []
    for stage in stages:
        date_count = len(stage["dates"])
        saved = stage_saved_count(stage)
        plan.append(
            {
                "index": stage["index"],
                "key": stage["key"],
                "label": stage["label"],
                "start_date": stage["start_date"],
                "end_date": stage["end_date"],
                "date_count": date_count,
                "saved": saved,
                "complete": date_count > 0 and saved >= date_count,
                "output_folder": str(stage["output_folder"]),
            }
        )
    return plan


def write_state(status, stage=None, **extra):
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "status": status,
        "state_file": str(STATE_PATH),
        "primary_objective": PRIMARY_OBJECTIVE,
        "context_folder": str(CONTEXT_FOLDER),
        "session_target": SESSION_TARGET,
    }
    if LADDER_PLAN_STAGES:
        payload["ladder_plan"] = LADDER_PLAN_STAGES
        payload["stage_count_total"] = len(LADDER_PLAN_STAGES)
    if stage:
        payload["stage"] = {
            "key": stage["key"],
            "label": stage["label"],
            "run_id": stage["run_id"],
            "start_date": stage["start_date"],
            "end_date": stage["end_date"],
            "date_count": len(stage["dates"]),
            "output_folder": str(stage["output_folder"]),
            "report_prefix": stage["report_prefix"],
        }
    payload.update(extra)
    STATE_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def classify_stage_outcome(stage, failures):
    """Separa el resultado de una etapa en dos categorías:

    - real_divergences: desync VERDADERO de operativa (X1 != X10 en campos que
      afectan WR/PF). Esto SÍ detiene el ladder: es lo que importa para confiar.
    - gap_dates: huecos re-corregibles (timeout, replay no arrancó, foco perdido,
      MISSING_CSV). NO detienen el ladder; se rellenan con un re-run sin --force.
    """
    gap_dates = set()
    for date_iso, _run_name, _reason in failures:
        if date_iso:
            gap_dates.add(date_iso)

    real_divergences = []
    report_path = stage["output_folder"] / f"{stage['report_prefix']}_comparacion.csv"
    if report_path.exists():
        with open(report_path, "r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if row.get("Estado") != "FAIL":
                    continue
                diff = (row.get("Campos_Diferentes") or "").strip()
                if diff in ("", "MISSING_CSV"):
                    gap_dates.add(row.get("Fecha"))
                else:
                    real_divergences.append((row.get("Fecha"), diff))

    return real_divergences, sorted(d for d in gap_dates if d)


def clean_divergent_date(stage, run_plan, date_iso):
    """Borra los artefactos de una fecha (CSV X1/X10, timeline, snapshot, signal)
    para forzar un re-pareo limpio X1->X10 en el reintento automatico de desync.
    """
    for run_name, _speed, _timeout in run_plan:
        replay_sync.destination_result_path(
            stage["output_folder"], date_iso, run_name
        ).unlink(missing_ok=True)
        replay_sync.destination_timeline_path(
            stage["output_folder"], date_iso, run_name
        ).unlink(missing_ok=True)
    replay_sync.sync_result_path(date_iso).unlink(missing_ok=True)
    replay_sync.sync_signal_path(date_iso).unlink(missing_ok=True)


def read_csv_rows(path):
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def append_rows_sheet(workbook, sheet_name, rows):
    ws = workbook.create_sheet(sheet_name[:31])
    if not rows:
        ws.append(["sin_datos"])
        return

    headers = []
    for row in rows:
        for key in row.keys():
            if key not in headers:
                headers.append(key)

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    for column_cells in ws.columns:
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 2, 10), 42)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def collect_run_rows(stage, run_name):
    rows = []
    for date_iso in stage["dates"]:
        path = replay_sync.destination_result_path(stage["output_folder"], date_iso, run_name)
        row = replay_sync.read_csv_row(path)
        if row:
            enriched = {"Stage": stage["key"], "Run": run_name, "Fecha": date_iso}
            enriched.update(row)
            rows.append(enriched)
    return rows


def create_summary_workbook(stage, run_plan, passed, failures, archive_path, timestamp):
    output_folder = stage["output_folder"]
    report_prefix = stage["report_prefix"]
    comparison_path = output_folder / f"{report_prefix}_comparacion.csv"
    summary_text_path = output_folder / f"{report_prefix}_resumen.txt"
    workbook_path = output_folder / (
        f"{stage['key']}_{stage['run_id']}_{timestamp}_summary.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    rows = [
        ("Stage", stage["key"]),
        ("Run ID", stage["run_id"]),
        ("Objetivo principal", PRIMARY_OBJECTIVE),
        ("Ruta contexto Codex", str(CONTEXT_FOLDER)),
        ("Nombre", stage["label"]),
        ("Inicio", stage["start_date"]),
        ("Fin", stage["end_date"]),
        ("Fechas", len(stage["dates"])),
        ("Estado", "PASS" if passed and not failures else "FAIL"),
        ("Exporter esperado", replay_sync.EXPECTED_EXPORTER_VERSION),
        ("Timeline esperado", replay_sync.EXPECTED_TIMELINE_VERSION),
        ("Output folder", str(output_folder)),
        ("Archive zip", str(archive_path)),
    ]
    for row in rows:
        ws.append(row)
    ws.append([])
    ws.append(["Run", "Replay", "Timeout segundos"])
    for run_name, speed_label, timeout_seconds in run_plan:
        ws.append([run_name, speed_label, timeout_seconds])
    ws.append([])
    ws.append(["Failure date", "Run", "Reason"])
    for failure in failures:
        ws.append(list(failure))

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 18

    append_rows_sheet(wb, "Comparacion", read_csv_rows(comparison_path))
    for run_name, _, _ in run_plan:
        append_rows_sheet(wb, run_name, collect_run_rows(stage, run_name))

    if summary_text_path.exists():
        summary_ws = wb.create_sheet("Resumen TXT")
        for line in summary_text_path.read_text(encoding="utf-8").splitlines():
            summary_ws.append([line])
        summary_ws.column_dimensions["A"].width = 120

    output_folder.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(workbook_path)
    except PermissionError:
        fallback = output_folder / (
            f"{stage['key']}_{stage['run_id']}_summary_"
            f"{datetime.now():%Y%m%d_%H%M%S}_fallback.xlsx"
        )
        wb.save(fallback)
        workbook_path = fallback
    return workbook_path


def add_path_to_zip(zip_file, path, arc_root):
    if path.is_file():
        zip_file.write(path, arc_root / path.name)
        return

    for child in path.rglob("*"):
        if child.is_file():
            zip_file.write(child, arc_root / child.relative_to(path))


def archive_success(stage, workbook_path, timestamp):
    archive_path = PROJECT_DIR / (
        f"{ARCHIVE_PREFIX}_{stage['run_id']}_{stage['key']}_{timestamp}.zip"
    )
    output_folder = stage["output_folder"]
    arc_root = Path(stage["key"])

    with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        if workbook_path.exists():
            zip_file.write(workbook_path, arc_root / workbook_path.name)
        if output_folder.exists():
            add_path_to_zip(zip_file, output_folder, arc_root / "generated")

    print(f"Zip PASS creado: {archive_path}")
    return archive_path


def filter_stages(stages, start_at=None, stop_after=None, max_stages=None):
    selected = stages
    if start_at:
        start_index = None
        for stage in selected:
            if stage["key"] == start_at or str(stage["index"]) == start_at:
                start_index = stage["index"]
                break
        if start_index is not None:
            selected = [stage for stage in selected if stage["index"] >= start_index]
    if stop_after:
        stop_index = None
        for stage in selected:
            if stage["key"] == stop_after or str(stage["index"]) == stop_after:
                stop_index = stage["index"]
                break
        if stop_index is not None:
            selected = [stage for stage in selected if stage["index"] <= stop_index]
    if max_stages is not None:
        selected = selected[:max_stages]
    return selected


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Escalera automatica de sincronicidad X1/X10. Corre fechas "
            "conflictivas, 6 meses, 1 ano y bloques historicos de 6 meses."
        )
    )
    parser.add_argument("--prepare-only", action="store_true", help="Muestra el plan sin iniciar Replay.")
    parser.add_argument("--compare-only", action="store_true", help="Solo recompone reportes usando archivos guardados.")
    parser.add_argument("--force", action="store_true", help="Repite corridas aunque ya existan CSV guardados.")
    parser.add_argument("--step", action="store_true", help="Pide ENTER antes de cada fecha.")
    parser.add_argument("--full-x10", action="store_true", help="Usa tres repeticiones X10 por etapa.")
    parser.add_argument("--include-one-month", action="store_true", help="Agrega una etapa smoke de 1 mes antes de 6 meses.")
    parser.add_argument("--until-date", default="2022-06-15", help="Fecha historica minima inclusive, formato YYYY-MM-DD.")
    parser.add_argument("--start-at", help="Stage key o indice desde donde continuar.")
    parser.add_argument("--stop-after", help="Stage key o indice donde detenerse.")
    parser.add_argument("--max-stages", type=int, help="Limita cantidad de etapas para pruebas.")
    parser.add_argument("--allow-sleep", action="store_true", help="No bloquea suspension de Windows.")
    parser.add_argument("--run-id", help="Identificador unico de la corrida. Por defecto usa timestamp.")
    parser.add_argument(
        "--desync-retries",
        type=int,
        default=3,
        help="Reintentos automaticos por desync real (limpia y re-corre la fecha). "
        "Si persiste tras estos, se detiene como bug genuino y avisa por Telegram. "
        "0 = sin reintento.",
    )
    return parser.parse_args()


def print_plan(stages, run_plan):
    print("\nESCALERA DE SINCRONICIDAD X1/X10")
    print(f"Version esperada: {replay_sync.EXPECTED_EXPORTER_VERSION}")
    print(PRIMARY_OBJECTIVE)
    print(f"Ruta contexto Codex: {CONTEXT_FOLDER}")
    print(f"Estado: {STATE_PATH}")
    print(f"Zips PASS: {PROJECT_DIR}")
    if stages:
        print(f"Run ID: {stages[0]['run_id']}")
    print("\nRun plan:")
    for run_name, speed_label, _ in run_plan:
        print(f"  - {run_name}: Replay {speed_label}")
    print("\nEtapas:")
    for stage in stages:
        period = (
            f"{stage['start_date']} -> {stage['end_date']}"
            if stage["start_date"]
            else "fechas fijas"
        )
        print(
            f"  {stage['index']:02d}. {stage['key']} | "
            f"{period} | {len(stage['dates'])} fechas | {stage['output_folder']}"
        )


def main():
    args = parse_args()
    until_date = datetime.strptime(args.until_date, "%Y-%m-%d").date()
    run_id = sanitize_key(args.run_id or datetime.now().strftime("ladder_%Y%m%d_%H%M%S"))
    run_plan = replay_sync.build_run_plan(quick=not args.full_x10)
    stages = build_ladder_stages(until_date, run_id, include_one_month=args.include_one_month)
    total_stage_count = len(stages)
    # Registra el plan completo (todas las etapas) en el .json, no solo la actual.
    global LADDER_PLAN_STAGES
    LADDER_PLAN_STAGES = build_ladder_plan(stages)
    stages = filter_stages(
        stages,
        start_at=args.start_at,
        stop_after=args.stop_after,
        max_stages=args.max_stages,
    )

    print_plan(stages, run_plan)
    if args.prepare_only:
        write_state("prepared", stages[0] if stages else None, stage_count=len(stages))
        print("\nPREPARE-ONLY correcto. No se inicio Replay.")
        return 0

    # Limpieza masiva de mensajes Telegram de corridas anteriores (una vez por
    # corrida real; en compare-only no se toca Telegram).
    if not args.compare_only:
        telegram.clear_telegram_before_run(replay_sync.RESULTS_FOLDER)

    session_roots = [
        LADDER_OUTPUT_ROOT / run_id,
        CONFLICT_OUTPUT_FOLDER / run_id,
    ]

    sleep_context = (
        WindowsSleepBlocker(keep_display_on=True)
        if not args.allow_sleep
        else None
    )

    # Huecos re-corregibles acumulados (timeout/foco/missing) que NO detienen el
    # ladder; se reportan al final para un re-run sin --force que los rellene.
    all_gaps = []

    context = sleep_context if sleep_context is not None else nullcontext()
    with context as blocker:
        for stage in stages:
            if not stage["dates"]:
                print(f"\nEtapa sin fechas operables; saltando: {stage['key']}")
                continue

            if blocker:
                blocker.refresh()

            print(f"\n=== ETAPA {stage['index']:02d}: {stage['key']} ===")
            write_state("running", stage)
            stage["output_folder"].mkdir(parents=True, exist_ok=True)

            stage_period = (
                f"{stage['start_date']} -> {stage['end_date']}"
                if stage["start_date"]
                else "fechas fijas"
            )
            progress_meta = {
                "stage_index": stage["index"],
                "stage_total": total_stage_count,
                "stage_label": stage["key"],
                "stage_period": stage_period,
                "global_target": SESSION_TARGET,
                "session_roots": session_roots,
                "run_label": "Ladder X1/X10",
            }

            # Reintento automatico ante desync REAL: un desync puede ser un glitch
            # transitorio (re-correr limpio lo arregla) o un bug genuino (persiste).
            # Se limpia la(s) fecha(s) divergente(s) y se re-corre hasta
            # args.desync_retries veces; si persiste, recien ahi se detiene.
            desync_attempt = 0
            while True:
                passed, failures = replay_sync.run_replay_period(
                    stage["dates"],
                    output_folder=stage["output_folder"],
                    run_plan=run_plan,
                    report_prefix=stage["report_prefix"],
                    force=args.force,
                    step=args.step,
                    compare_only=args.compare_only,
                    replay_to_time=replay_sync.DEFAULT_REPLAY_TO_TIME,
                    progress_meta=progress_meta,
                )

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                archive_path = PROJECT_DIR / (
                    f"{ARCHIVE_PREFIX}_{stage['run_id']}_{stage['key']}_{timestamp}.zip"
                )
                workbook_path = create_summary_workbook(
                    stage,
                    run_plan,
                    passed,
                    failures,
                    archive_path,
                    timestamp,
                )

                real_divergences, gap_dates = classify_stage_outcome(stage, failures)

                if not real_divergences:
                    break

                diverging = [date_iso for date_iso, _ in real_divergences]

                # Persiste tras los reintentos -> bug GENUINO, detener + avisar.
                if desync_attempt >= args.desync_retries:
                    write_state(
                        "failed",
                        stage,
                        latest_summary=str(workbook_path),
                        real_divergences=[
                            {"date": date_iso, "fields": fields}
                            for date_iso, fields in real_divergences
                        ],
                        gaps=gap_dates,
                        desync_retries_used=desync_attempt,
                    )
                    telegram.send_text(
                        replay_sync.RESULTS_FOLDER,
                        "\n".join(
                            [
                                "EW Opening Range | DESYNC PERSISTENTE - Fase 1",
                                f"Fase: ETAPA {stage['index']:02d}/{total_stage_count:02d} - {stage['key']}",
                                f"Divergencia REAL tras {desync_attempt} reintentos en: "
                                + ", ".join(diverging),
                                "Campos: "
                                + " | ".join(
                                    f"{date_iso}: {fields}"
                                    for date_iso, fields in real_divergences
                                ),
                                "Ladder DETENIDO. Requiere arreglo manual del .cs.",
                            ]
                        ),
                    )
                    print(
                        "\nFALLO LA SINCRONICIDAD: divergencia REAL PERSISTENTE "
                        f"tras {desync_attempt} reintento(s) -> bug genuino."
                    )
                    for date_iso, fields in real_divergences:
                        print(f"  {date_iso}: {fields}")
                    print(f"Estado para Codex: {STATE_PATH}")
                    print(f"Resumen Excel: {workbook_path}")
                    print(f"Carpeta de reportes: {stage['output_folder']}")
                    return 1

                # Reintento: limpiar artefactos de las fechas divergentes para
                # forzar un re-pareo X1/X10 desde cero (solo esas se re-corren).
                desync_attempt += 1
                print(
                    f"\nDesync real en {', '.join(diverging)} -> reintento "
                    f"{desync_attempt}/{args.desync_retries}: limpio y re-corro esas fechas."
                )
                for date_iso in diverging:
                    clean_divergent_date(stage, run_plan, date_iso)
                write_state("desync_retry", stage, diverging=diverging, attempt=desync_attempt)

            # Solo huecos re-corregibles (timeout/foco/missing) -> NO detiene;
            # se aisla la(s) fecha(s), se loguea y el ladder sigue a la siguiente.
            if gap_dates:
                all_gaps.extend((stage["key"], date_iso) for date_iso in gap_dates)
                write_state(
                    "stage_pass_with_gaps",
                    stage,
                    latest_summary=str(workbook_path),
                    gaps=gap_dates,
                )
                print(
                    f"\nEtapa {stage['key']} OK con {len(gap_dates)} hueco(s) "
                    f"re-corregible(s): {', '.join(gap_dates)}"
                )
                print(
                    "Sin divergencia real de operativa. El ladder continua; "
                    "los huecos se rellenan luego con un re-run sin --force."
                )
            else:
                # Limpio total -> archiva el zip de PASS y avanza.
                archive_path = archive_success(stage, workbook_path, timestamp)
                write_state(
                    "stage_pass",
                    stage,
                    latest_summary=str(workbook_path),
                    latest_archive=str(archive_path),
                )

            # META: la etapa cerro con X1+X10+comparacion, asi que todo lo contado
            # esta sincronizado. Si cruzamos 500, detenemos el Replay y avisamos.
            global_done = replay_sync.count_distinct_sessions(session_roots)
            if global_done is not None and global_done >= SESSION_TARGET:
                telegram.send_progress_update(
                    replay_sync.RESULTS_FOLDER,
                    stage_index=stage["index"],
                    stage_total=total_stage_count,
                    stage_label=stage["key"],
                    stage_period=stage_period,
                    done=global_done,
                    total=SESSION_TARGET,
                    passed=0,
                    failed=0,
                    skipped=0,
                    eta_seconds=None,
                    avg_seconds=None,
                    remaining=0,
                    global_done=global_done,
                    global_target=SESSION_TARGET,
                    run_label="Ladder X1/X10",
                    goal=True,
                )
                write_state(
                    "goal_reached",
                    stage,
                    latest_summary=str(workbook_path),
                    sessions=global_done,
                    gaps=[
                        {"stage": stage_key, "date": date_iso}
                        for stage_key, date_iso in all_gaps
                    ],
                )
                print(
                    f"\nMETA ALCANZADA: {global_done}/{SESSION_TARGET} sesiones "
                    "sincronizadas X1==X10. Replay detenido. Fase 1 completa."
                )
                if all_gaps:
                    print(
                        f"Huecos re-corregibles pendientes ({len(all_gaps)}): "
                        + ", ".join(f"{s} {d}" for s, d in all_gaps)
                    )
                print("Siguiente: Fase 2 (definir estrategia).")
                return 0

    write_state(
        "complete",
        stages[-1] if stages else None,
        stage_count=len(stages),
        gaps=[{"stage": stage_key, "date": date_iso} for stage_key, date_iso in all_gaps],
    )
    if all_gaps:
        print(
            f"\nESCALERA COMPLETA con {len(all_gaps)} hueco(s) re-corregible(s) "
            "(timeout/foco/missing, sin desync real):"
        )
        for stage_key, date_iso in all_gaps:
            print(f"  {stage_key}  {date_iso}")
        print(
            "\nRellenarlos: re-run SIN --force (salta lo bueno, re-corre solo los "
            "huecos). Si una fecha vuelve a dar timeout, es hueco de datos real."
        )
    else:
        print("\nESCALERA COMPLETA: todas las etapas pasaron limpio.")
    return 0


class nullcontext:
    def __enter__(self):
        return None

    def __exit__(self, exc_type, exc, tb):
        return False


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        replay_sync.click_stop()
        write_state("cancelled")
        print("\nEscalera cancelada por el usuario.")
        raise SystemExit(130)
