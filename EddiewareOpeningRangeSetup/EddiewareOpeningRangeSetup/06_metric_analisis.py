# -*- coding: utf-8 -*-
from pathlib import Path
import re
import statistics

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIG
# ============================================================

FOOTPRINT_DIR = Path(
    r"C:\Users\k_99_\Desktop\codding\data_footprint_generator\footprints_generados"
)
OUTPUT_XLSX = Path(__file__).with_name("metric results.xlsx")

OPENING_CANDLE_TIME = "09:30"
TICK_SIZE = 0.25
MIN_CONTINUATION_TICKS = 60

BODY_FILLS = {
    "00C6EFCE",  # bullish body
    "00FFC7CE",  # bearish body
    "C6EFCE",
    "FFC7CE",
}
BODY_BORDER_COLORS = {
    "0000B050",  # bullish body border
    "00C00000",  # bearish body border
    "00B050",
    "C00000",
}
BULLISH_BODY_COLORS = {"00C6EFCE", "C6EFCE", "0000B050", "00B050"}
BEARISH_BODY_COLORS = {"00FFC7CE", "FFC7CE", "00C00000", "C00000"}
VOLUME_RE = re.compile(r"^\s*\d+\s*x\s*\d+\s*$", re.IGNORECASE)


# ============================================================
# HELPERS
# ============================================================

def detect_date_from_filename(path):
    match = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    return match.group(1) if match else path.stem


def round_to_tick(value):
    return round(float(value) / TICK_SIZE, 2)


def safe_mean(values):
    values = [v for v in values if v is not None]
    return statistics.mean(values) if values else None


def rgb(color):
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    return None


def is_volume_cell(value):
    return isinstance(value, str) and bool(VOLUME_RE.match(value))


def is_body_cell(cell):
    fill_rgb = rgb(cell.fill.fgColor)
    if fill_rgb in BODY_FILLS:
        return True

    left = cell.border.left
    right = cell.border.right
    left_rgb = rgb(left.color)
    right_rgb = rgb(right.color)

    return (
        left.style == "medium"
        and right.style == "medium"
        and (left_rgb in BODY_BORDER_COLORS or right_rgb in BODY_BORDER_COLORS)
    )


def body_direction(cell):
    fill_rgb = rgb(cell.fill.fgColor)
    if fill_rgb in BULLISH_BODY_COLORS:
        return "BUY"
    if fill_rgb in BEARISH_BODY_COLORS:
        return "SELL"

    colors = [
        rgb(cell.border.left.color),
        rgb(cell.border.right.color),
        rgb(cell.border.top.color),
        rgb(cell.border.bottom.color),
    ]
    if any(color in BULLISH_BODY_COLORS for color in colors):
        return "BUY"
    if any(color in BEARISH_BODY_COLORS for color in colors):
        return "SELL"
    return None


def find_minute_column(ws, minute):
    for cell in ws[2]:
        if str(cell.value).strip() == minute:
            return cell.column
    return None


def get_minute_columns(ws):
    minute_columns = []
    for cell in ws[2]:
        value = str(cell.value or "").strip()
        if re.match(r"^\d{2}:\d{2}$", value):
            minute_columns.append((value, cell.column))
    return minute_columns


def find_label_row(ws, label):
    target = str(label).strip().lower()
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if str(value or "").strip().lower() == target:
            return row_idx
    return None


def read_metric_by_minute(ws, minute_columns, label):
    label_row = find_label_row(ws, label)
    if label_row is None:
        return {}

    return {
        minute: ws.cell(row=label_row, column=minute_col).value
        for minute, minute_col in minute_columns
    }


def iter_price_rows(ws):
    for row_idx in range(3, ws.max_row + 1):
        price = ws.cell(row=row_idx, column=1).value
        if price in (None, "", "VWAP", "Métrica", "Delta", "Volumen"):
            continue

        try:
            yield row_idx, float(price)
        except (TypeError, ValueError):
            continue


def read_candle(ws, minute, minute_col):
    traded_prices = []
    body_prices = []
    directions = []

    for row_idx, price in iter_price_rows(ws):
        cell = ws.cell(row=row_idx, column=minute_col)

        if is_volume_cell(cell.value):
            traded_prices.append(price)

        if is_body_cell(cell):
            body_prices.append(price)
            direction = body_direction(cell)
            if direction:
                directions.append(direction)

    if not traded_prices:
        return None

    high = max(traded_prices)
    low = min(traded_prices)
    range_points = high - low
    range_ticks = round_to_tick(range_points)

    body_ticks = None
    body_pct = None
    body_high = None
    body_low = None
    direction = None
    close = None

    if body_prices and range_points > 0:
        body_high = max(body_prices)
        body_low = min(body_prices)
        body_ticks = round_to_tick(body_high - body_low)
        body_pct = (body_ticks / range_ticks) * 100 if range_ticks else 0

        if directions:
            direction = max(set(directions), key=directions.count)
            close = body_high if direction == "BUY" else body_low

    return {
        "minute": minute,
        "low": low,
        "high": high,
        "range_points": range_points,
        "range_ticks": range_ticks,
        "body_low": body_low,
        "body_high": body_high,
        "body_ticks": body_ticks,
        "body_pct": body_pct,
        "direction": direction,
        "close": close,
    }


def find_breakout_continuation(candles, opening_candle):
    or_high = opening_candle["high"]
    or_low = opening_candle["low"]
    minutes = sorted(candles)

    for idx, minute in enumerate(minutes):
        if minute <= OPENING_CANDLE_TIME:
            continue
        if idx + 1 >= len(minutes):
            break

        candle = candles[minute]
        next_minute = minutes[idx + 1]
        next_candle = candles[next_minute]
        direction = candle.get("direction")
        close = candle.get("close")

        if direction == "BUY" and close is not None and close > or_high:
            breakout_ticks = round_to_tick(close - or_high)
            continuation_ticks = round_to_tick(next_candle["high"] - close)
            if continuation_ticks >= MIN_CONTINUATION_TICKS:
                return {
                    "breakout_time": minute,
                    "breakout_side": "BUY",
                    "breakout_body_price": close,
                    "breakout_body_ticks": breakout_ticks,
                    "next_candle_time": next_minute,
                    "next_continuation_ticks": continuation_ticks,
                    "continuation_60t": True,
                }

        if direction == "SELL" and close is not None and close < or_low:
            breakout_ticks = round_to_tick(or_low - close)
            continuation_ticks = round_to_tick(close - next_candle["low"])
            if continuation_ticks >= MIN_CONTINUATION_TICKS:
                return {
                    "breakout_time": minute,
                    "breakout_side": "SELL",
                    "breakout_body_price": close,
                    "breakout_body_ticks": breakout_ticks,
                    "next_candle_time": next_minute,
                    "next_continuation_ticks": continuation_ticks,
                    "continuation_60t": True,
                }

    return {
        "breakout_time": None,
        "breakout_side": None,
        "breakout_body_price": None,
        "breakout_body_ticks": None,
        "next_candle_time": None,
        "next_continuation_ticks": None,
        "continuation_60t": False,
    }


def add_breakout_metrics(breakout, volume_by_minute, delta_by_minute, vwap_by_minute):
    breakout_time = breakout.get("breakout_time")
    if not breakout_time:
        breakout.update(
            {
                "breakout_volume": None,
                "breakout_delta": None,
                "breakout_vwap": None,
            }
        )
        return breakout

    breakout.update(
        {
            "breakout_volume": volume_by_minute.get(breakout_time),
            "breakout_delta": delta_by_minute.get(breakout_time),
            "breakout_vwap": vwap_by_minute.get(breakout_time),
        }
    )
    return breakout


def add_trade_simulation(row):
    side = row.get("breakout_side")
    entry_price = row.get("breakout_body_price")

    if side == "BUY" and entry_price is not None:
        row["entry_price"] = entry_price
        row["SL_price"] = row.get("or_low")
        row["TP_price"] = entry_price + (MIN_CONTINUATION_TICKS * TICK_SIZE)
    elif side == "SELL" and entry_price is not None:
        row["entry_price"] = entry_price
        row["SL_price"] = row.get("or_high")
        row["TP_price"] = entry_price - (MIN_CONTINUATION_TICKS * TICK_SIZE)
    else:
        row["entry_price"] = None
        row["SL_price"] = None
        row["TP_price"] = None

    return row


def passes_continuation_filter(row):
    return (
        row.get("status") == "OK"
        and row.get("breakout_time") is not None
        and row.get("breakout_side") in {"BUY", "SELL"}
        and row.get("breakout_body_ticks") is not None
        and row.get("breakout_volume") is not None
        and row.get("breakout_delta") is not None
        and row.get("breakout_vwap") is not None
        and row.get("next_continuation_ticks") is not None
        and row.get("next_continuation_ticks") >= MIN_CONTINUATION_TICKS
    )


def analyze_opening_candle(path):
    wb = load_workbook(path, read_only=False, data_only=True)
    if "Footprint" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'Footprint'.")

    ws = wb["Footprint"]
    minute_columns = get_minute_columns(ws)
    candles = {}
    for minute, minute_col in minute_columns:
        candle = read_candle(ws, minute, minute_col)
        if candle:
            candles[minute] = candle

    opening_candle = candles.get(OPENING_CANDLE_TIME)
    if opening_candle is None:
        raise ValueError(f"No hay datos de volumen en {OPENING_CANDLE_TIME}.")

    breakout = find_breakout_continuation(candles, opening_candle)
    volume_by_minute = read_metric_by_minute(ws, minute_columns, "Volumen")
    delta_by_minute = read_metric_by_minute(ws, minute_columns, "Delta")
    vwap_by_minute = read_metric_by_minute(ws, minute_columns, "VWAP")
    breakout = add_breakout_metrics(
        breakout,
        volume_by_minute,
        delta_by_minute,
        vwap_by_minute,
    )

    row = {
        "date": detect_date_from_filename(path),
        "file": path.name,
        "minute": OPENING_CANDLE_TIME,
        "or_low": opening_candle["low"],
        "or_high": opening_candle["high"],
        "range_points": opening_candle["range_points"],
        "range_ticks": opening_candle["range_ticks"],
        "body_low": opening_candle["body_low"],
        "body_high": opening_candle["body_high"],
        "body_ticks": opening_candle["body_ticks"],
        "body_pct": opening_candle["body_pct"],
        **breakout,
        "status": "OK",
        "error": "",
    }
    return add_trade_simulation(row)


def format_number(value, digits=2):
    if value is None:
        return ""
    return f"{value:.{digits}f}"


def write_excel(rows, summary):
    headers = [
        "date",
        "file",
        "minute",
        "or_low",
        "or_high",
        "range_points",
        "range_ticks",
        "body_low",
        "body_high",
        "body_ticks",
        "body_pct",
        "breakout_time",
        "breakout_side",
        "breakout_body_price",
        "breakout_body_ticks",
        "breakout_volume",
        "breakout_delta",
        "breakout_vwap",
        "entry_price",
        "SL_price",
        "TP_price",
        "next_candle_time",
        "next_continuation_ticks",
        "continuation_60t",
        "filter_taken",
        "status",
        "error",
    ]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_detail = wb.create_sheet("Detail")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    error_fill = PatternFill("solid", fgColor="FFC7CE")

    ws_summary["A1"] = "Metric Results - Vela 09:30"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A1"].fill = header_fill
    ws_summary.merge_cells("A1:B1")

    summary_rows = [
        ("Archivos analizados OK", summary["ok_files"]),
        ("Archivos totales", summary["total_files"]),
        ("Promedio rango vela 09:30 (ticks)", summary["avg_range_ticks"]),
        ("Promedio body vela 09:30 (ticks)", summary["avg_body_ticks"]),
        ("Promedio % body vela 09:30", summary["avg_body_pct"]),
        ("Continuaciones simples >= 60 ticks", summary["continuation_count"]),
        ("% continuaciones simples >= 60 ticks", summary["continuation_pct"]),
        ("Promedio breakout con cuerpo (ticks)", summary["avg_breakout_body_ticks"]),
        ("Trades tomados con filtro", summary["filter_taken_count"]),
        ("% trades tomados con filtro", summary["filter_taken_pct"]),
        ("% continuaciones capturadas por filtro", summary["filter_capture_pct"]),
        ("Filtro usado", summary["filter_rule"]),
        ("Ruta analizada", str(FOOTPRINT_DIR)),
    ]

    for row_idx, (label, value) in enumerate(summary_rows, start=3):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary["C12"] = "equivale a trades"
    ws_summary["C12"].font = Font(bold=True)
    ws_summary["D12"] = summary["filter_taken_count"]
    ws_summary["D12"].font = Font(bold=True)
    ws_summary["D12"].fill = ok_fill

    percent_labels = {
        "Promedio % body vela 09:30",
        "% continuaciones simples >= 60 ticks",
        "% trades tomados con filtro",
        "% continuaciones capturadas por filtro",
    }

    for row_idx in range(3, 3 + len(summary_rows)):
        label = ws_summary.cell(row=row_idx, column=1).value
        value_cell = ws_summary.cell(row=row_idx, column=2)
        if label in percent_labels and isinstance(value_cell.value, (int, float)):
            value_cell.value = value_cell.value / 100
            value_cell.number_format = "0.00%"
        elif isinstance(value_cell.value, float):
            value_cell.number_format = "0.00"

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
            if header == "body_pct" and isinstance(value, (int, float)):
                cell.value = value / 100
                cell.number_format = "0.00%"
            elif isinstance(value, float):
                cell.number_format = "0.00"

        if row.get("filter_taken"):
            for col_idx in range(1, len(headers) + 1):
                ws_detail.cell(row=row_idx, column=col_idx).fill = ok_fill

        status_cell = ws_detail.cell(row=row_idx, column=headers.index("status") + 1)
        if row.get("filter_taken"):
            status_cell.fill = ok_fill
        else:
            status_cell.fill = ok_fill if row.get("status") == "OK" else error_fill

    for ws in (ws_summary, ws_detail):
        for column in ws.columns:
            max_len = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                max_len = max(max_len, len(str(cell.value or "")))
            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 70)
        ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)


def main():
    files = sorted(FOOTPRINT_DIR.glob("*.xlsx"))
    if not files:
        raise SystemExit(f"No se encontraron archivos .xlsx en: {FOOTPRINT_DIR}")

    rows = []
    for path in files:
        try:
            rows.append(analyze_opening_candle(path))
        except Exception as exc:
            rows.append(
                {
                    "date": detect_date_from_filename(path),
                    "file": path.name,
                    "minute": OPENING_CANDLE_TIME,
                    "status": "ERROR",
                    "error": str(exc),
                }
            )

    for row in rows:
        row["filter_taken"] = passes_continuation_filter(row)

    ok_rows = [row for row in rows if row.get("status") == "OK"]
    avg_range_ticks = safe_mean([row.get("range_ticks") for row in ok_rows])
    avg_body_pct = safe_mean([row.get("body_pct") for row in ok_rows])
    avg_body_ticks = safe_mean([row.get("body_ticks") for row in ok_rows])
    continuation_rows = [row for row in ok_rows if row.get("continuation_60t")]
    continuation_pct = (len(continuation_rows) / len(ok_rows) * 100) if ok_rows else None
    avg_breakout_body_ticks = safe_mean(
        [row.get("breakout_body_ticks") for row in continuation_rows]
    )
    filter_taken_rows = [row for row in ok_rows if row.get("filter_taken")]
    filter_taken_pct = (len(filter_taken_rows) / len(ok_rows) * 100) if ok_rows else None
    filter_capture_pct = (
        len([row for row in filter_taken_rows if row.get("continuation_60t")])
        / len(continuation_rows)
        * 100
    ) if continuation_rows else None
    filter_rule = (
        "status=OK; breakout_time no vacio; breakout_side BUY/SELL; "
        "breakout body/volume/delta/vwap presentes; next_continuation_ticks >= 60"
    )
    summary = {
        "ok_files": len(ok_rows),
        "total_files": len(rows),
        "avg_range_ticks": avg_range_ticks,
        "avg_body_ticks": avg_body_ticks,
        "avg_body_pct": avg_body_pct,
        "continuation_count": len(continuation_rows),
        "continuation_pct": continuation_pct,
        "avg_breakout_body_ticks": avg_breakout_body_ticks,
        "filter_taken_count": len(filter_taken_rows),
        "filter_taken_pct": filter_taken_pct,
        "filter_capture_pct": filter_capture_pct,
        "filter_rule": filter_rule,
    }

    write_excel(rows, summary)

    print("============================================================")
    print("METRICAS VELA DE APERTURA 09:30")
    print("============================================================")
    print(f"Archivos analizados OK: {len(ok_rows)} / {len(rows)}")
    print(f"Promedio rango vela 09:30: {format_number(avg_range_ticks)} ticks")
    print(f"Promedio body vela 09:30: {format_number(avg_body_ticks)} ticks")
    print(f"Promedio % body vela 09:30: {format_number(avg_body_pct)}%")
    print(
        f"Continuaciones simples >= {MIN_CONTINUATION_TICKS} ticks: "
        f"{len(continuation_rows)} / {len(ok_rows)}"
    )
    print(f"Promedio breakout con cuerpo: {format_number(avg_breakout_body_ticks)} ticks")
    print(
        f"Trades tomados con filtro: {len(filter_taken_rows)} / {len(ok_rows)} "
        f"({format_number(filter_taken_pct)}%)"
    )
    print(f"Excel guardado en: {OUTPUT_XLSX}")

    errors = [row for row in rows if row.get("status") != "OK"]
    if errors:
        print("")
        print("Archivos con error:")
        for row in errors[:20]:
            print(f"- {row['file']}: {row['error']}")
        if len(errors) > 20:
            print(f"- ... {len(errors) - 20} errores mas")


if __name__ == "__main__":
    main()
