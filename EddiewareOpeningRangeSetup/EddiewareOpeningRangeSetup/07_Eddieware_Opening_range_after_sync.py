# -*- coding: utf-8 -*-
from pathlib import Path
import re

from openpyxl import load_workbook


# ============================================================
# PURE OPENING RANGE SETUP CONFIG
# ============================================================

FOOTPRINT_DIR = Path(
    r"C:\Users\k_99_\Desktop\codding\data_footprint_generator\footprints_generados"
)

OPENING_CANDLE_TIME = "09:30"
TICK_SIZE = 0.25

MIN_CONTINUATION_TICKS = 60
MIN_TRADE_TICKS = 60
MAX_TRADE_TICKS = 120

BODY_BORDER_COLORS = {"00B050", "C00000"}
BULLISH_BODY_COLORS = {"C6EFCE", "00B050"}
BEARISH_BODY_COLORS = {"FFC7CE", "C00000"}
VOLUME_RE = re.compile(r"^\s*\d+\s*x\s*\d+\s*$", re.IGNORECASE)


# ============================================================
# GENERIC HELPERS
# ============================================================

def detect_date_from_filename(path):
    match = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    return match.group(1) if match else path.stem


def round_to_tick(value):
    return round(float(value) / TICK_SIZE, 2)


def ticks_to_points(ticks):
    return float(ticks) * TICK_SIZE


def clamp_ticks(ticks, min_ticks=MIN_TRADE_TICKS, max_ticks=MAX_TRADE_TICKS):
    return max(min_ticks, min(float(ticks), max_ticks))


def rgb(color):
    if color is None:
        return None
    if color.type != "rgb":
        return None

    value = color.rgb
    if value and len(value) == 8:
        return value[-6:]
    return value


def is_volume_cell(value):
    return isinstance(value, str) and bool(VOLUME_RE.match(value))


# ============================================================
# FOOTPRINT READING
# ============================================================

def is_body_cell(cell):
    left = cell.border.left
    right = cell.border.right
    left_rgb = rgb(left.color)
    right_rgb = rgb(right.color)

    return (
        (left.style == "medium" and left_rgb in BODY_BORDER_COLORS)
        or (right.style == "medium" and right_rgb in BODY_BORDER_COLORS)
    )


def body_direction(cell):
    colors = [
        rgb(cell.border.left.color),
        rgb(cell.border.right.color),
    ]
    if any(color in BULLISH_BODY_COLORS for color in colors):
        return "BUY"
    if any(color in BEARISH_BODY_COLORS for color in colors):
        return "SELL"
    return None


def get_minute_columns(ws):
    minute_columns = []
    for cell in ws[2]:
        value = str(cell.value or "").strip()
        if re.match(r"^\d{2}:\d{2}$", value):
            minute_columns.append((value, cell.column))
    return minute_columns


def find_label_row(ws, label):
    target = str(label).strip().lower()
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if str(value or "").strip().lower() == target:
            return row_idx
    return None


def read_metric_by_minute(ws, minute_columns, label):
    label_row = find_label_row(ws, label)
    if label_row is None:
        return {}

    return {
        minute: ws.cell(row=label_row, column=minute_col).value
        for minute, minute_col in minute_columns
    }


def iter_price_rows(ws):
    for row_idx in range(3, ws.max_row + 1):
        price = ws.cell(row=row_idx, column=1).value
        if price in (None, "", "VWAP", "Métrica", "MÃ©trica", "Delta", "Volumen"):
            continue

        try:
            yield row_idx, float(price)
        except (TypeError, ValueError):
            continue


def read_candle(ws, minute, minute_col):
    traded_prices = []
    body_prices = []
    directions = []

    for row_idx, price in iter_price_rows(ws):
        cell = ws.cell(row=row_idx, column=minute_col)

        if is_volume_cell(cell.value):
            traded_prices.append(price)

        if is_body_cell(cell):
            body_prices.append(price)
            direction = body_direction(cell)
            if direction:
                directions.append(direction)

    if not traded_prices:
        return None

    high = max(traded_prices)
    low = min(traded_prices)
    range_points = high - low
    range_ticks = round_to_tick(range_points)

    body_high = None
    body_low = None
    body_ticks = None
    body_pct = None
    direction = None
    close = None

    if body_prices and range_points > 0:
        body_high = max(body_prices)
        body_low = min(body_prices)
        body_ticks = round_to_tick(body_high - body_low)
        body_pct = (body_ticks / range_ticks) * 100 if range_ticks else 0

        if directions:
            direction = max(set(directions), key=directions.count)
            close = body_high if direction == "BUY" else body_low

    return {
        "minute": minute,
        "low": low,
        "high": high,
        "range_points": range_points,
        "range_ticks": range_ticks,
        "body_low": body_low,
        "body_high": body_high,
        "body_ticks": body_ticks,
        "body_pct": body_pct,
        "direction": direction,
        "close": close,
    }


# ============================================================
# PURE SETUP LOGIC
# ============================================================

def find_breakout_continuation(candles, opening_candle):
    or_high = opening_candle["high"]
    or_low = opening_candle["low"]
    minutes = sorted(candles)

    for idx, minute in enumerate(minutes):
        if minute <= OPENING_CANDLE_TIME:
            continue
        if idx + 1 >= len(minutes):
            break

        candle = candles[minute]
        next_minute = minutes[idx + 1]
        next_candle = candles[next_minute]
        direction = candle.get("direction")
        close = candle.get("close")

        if direction == "BUY" and close is not None and close > or_high:
            breakout_ticks = round_to_tick(close - or_high)
            continuation_ticks = round_to_tick(next_candle["high"] - close)
            if continuation_ticks >= MIN_CONTINUATION_TICKS:
                return {
                    "breakout_time": minute,
                    "breakout_side": "BUY",
                    "breakout_body_price": close,
                    "breakout_body_ticks": breakout_ticks,
                    "next_candle_time": next_minute,
                    "next_continuation_ticks": continuation_ticks,
                    "continuation_60t": True,
                }

        if direction == "SELL" and close is not None and close < or_low:
            breakout_ticks = round_to_tick(or_low - close)
            continuation_ticks = round_to_tick(close - next_candle["low"])
            if continuation_ticks >= MIN_CONTINUATION_TICKS:
                return {
                    "breakout_time": minute,
                    "breakout_side": "SELL",
                    "breakout_body_price": close,
                    "breakout_body_ticks": breakout_ticks,
                    "next_candle_time": next_minute,
                    "next_continuation_ticks": continuation_ticks,
                    "continuation_60t": True,
                }

    return {
        "breakout_time": None,
        "breakout_side": None,
        "breakout_body_price": None,
        "breakout_body_ticks": None,
        "next_candle_time": None,
        "next_continuation_ticks": None,
        "continuation_60t": False,
    }


def add_breakout_metrics(breakout, volume_by_minute, delta_by_minute, vwap_by_minute):
    breakout_time = breakout.get("breakout_time")
    if not breakout_time:
        breakout.update(
            {
                "breakout_volume": None,
                "breakout_delta": None,
                "breakout_vwap": None,
            }
        )
        return breakout

    breakout.update(
        {
            "breakout_volume": volume_by_minute.get(breakout_time),
            "breakout_delta": delta_by_minute.get(breakout_time),
            "breakout_vwap": vwap_by_minute.get(breakout_time),
        }
    )
    return breakout


def add_trade_simulation(row):
    side = row.get("breakout_side")
    entry_price = row.get("breakout_body_price")

    if side == "BUY" and entry_price is not None:
        trade_ticks = clamp_ticks(round_to_tick(entry_price - row.get("or_low")))
        trade_points = ticks_to_points(trade_ticks)
        row["entry_time"] = row.get("breakout_time")
        row["entry_price"] = entry_price
        row["SL_price"] = entry_price - trade_points
        row["TP_price"] = entry_price + trade_points
    elif side == "SELL" and entry_price is not None:
        trade_ticks = clamp_ticks(round_to_tick(row.get("or_high") - entry_price))
        trade_points = ticks_to_points(trade_ticks)
        row["entry_time"] = row.get("breakout_time")
        row["entry_price"] = entry_price
        row["SL_price"] = entry_price + trade_points
        row["TP_price"] = entry_price - trade_points
    else:
        row["entry_time"] = None
        row["entry_price"] = None
        row["SL_price"] = None
        row["TP_price"] = None

    return row


def passes_continuation_filter(row):
    return (
        row.get("status") == "OK"
        and row.get("breakout_time") is not None
        and row.get("breakout_side") in {"BUY", "SELL"}
        and row.get("breakout_body_ticks") is not None
        and row.get("breakout_volume") is not None
        and row.get("breakout_delta") is not None
        and row.get("breakout_vwap") is not None
        and row.get("next_continuation_ticks") is not None
        and row.get("next_continuation_ticks") >= MIN_CONTINUATION_TICKS
    )


def analyze_footprint_file(path):
    wb = load_workbook(path, read_only=False, data_only=True)
    if "Footprint" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'Footprint'.")

    ws = wb["Footprint"]
    minute_columns = get_minute_columns(ws)
    candles = {}

    for minute, minute_col in minute_columns:
        candle = read_candle(ws, minute, minute_col)
        if candle:
            candles[minute] = candle

    opening_candle = candles.get(OPENING_CANDLE_TIME)
    if opening_candle is None:
        raise ValueError(f"No hay datos de volumen en {OPENING_CANDLE_TIME}.")

    breakout = find_breakout_continuation(candles, opening_candle)
    breakout = add_breakout_metrics(
        breakout,
        read_metric_by_minute(ws, minute_columns, "Volumen"),
        read_metric_by_minute(ws, minute_columns, "Delta"),
        read_metric_by_minute(ws, minute_columns, "VWAP"),
    )

    row = {
        "date": detect_date_from_filename(path),
        "file": path.name,
        "minute": OPENING_CANDLE_TIME,
        "or_low": opening_candle["low"],
        "or_high": opening_candle["high"],
        "range_points": opening_candle["range_points"],
        "range_ticks": opening_candle["range_ticks"],
        "body_low": opening_candle["body_low"],
        "body_high": opening_candle["body_high"],
        "body_ticks": opening_candle["body_ticks"],
        "body_pct": opening_candle["body_pct"],
        **breakout,
        "status": "OK",
        "error": "",
    }
    row = add_trade_simulation(row)
    row["filter_taken"] = passes_continuation_filter(row)
    return row


def analyze_directory(footprint_dir=FOOTPRINT_DIR):
    rows = []
    files = sorted(
        path for path in Path(footprint_dir).glob("*.xlsx")
        if not path.name.startswith("~$")
    )

    for path in files:
        try:
            rows.append(analyze_footprint_file(path))
        except Exception as exc:
            rows.append(
                {
                    "date": detect_date_from_filename(path),
                    "file": path.name,
                    "minute": OPENING_CANDLE_TIME,
                    "status": "ERROR",
                    "error": str(exc),
                    "filter_taken": False,
                }
            )

    return rows


def main():
    rows = analyze_directory()
    ok_rows = [row for row in rows if row.get("status") == "OK"]
    taken_rows = [row for row in ok_rows if row.get("filter_taken")]

    print("============================================================")
    print("EDDIEWARE OPENING RANGE FILTER")
    print("============================================================")
    print(f"Archivos OK: {len(ok_rows)} / {len(rows)}")
    print(f"Trades tomados por filtro: {len(taken_rows)} / {len(ok_rows)}")
    if ok_rows:
        print(f"Porcentaje tomado: {(len(taken_rows) / len(ok_rows) * 100):.2f}%")


if __name__ == "__main__":
    main()
