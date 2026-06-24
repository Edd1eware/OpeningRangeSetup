# -*- coding: utf-8 -*-
from pathlib import Path
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# ============================================================
# CONFIG
# ============================================================

RAW_DIR = Path(r"C:\Users\k_99_\Desktop\codding\data_footprint_generator")
OUTPUT_DIR = RAW_DIR / "footprints_generados"

START_TIME = "09:30"
END_TIME = "10:30"
# Set to "YYYY-MM-DD" to process one date, or None to process every file.
PROCESS_ONLY_DATE = None

# ATAS style imbalance
IMBALANCE_RATIO = 3.0
IMBALANCE_MIN_VOLUME = 100
VOLUME_FILTER_MODES = ["atas_compare"]
DIAGONAL_COMPARE_MIN_VOLUME = 5

# V26: enfocarnos en niveles reales, no zonas artificiales
ALLOW_GAP_LEVELS = 0
MIN_STACK_RAW_IMBALANCES = 1
MAX_STACK_SPAN_LEVELS = 9999

# V26: compresión visual tipo ATAS
# En tu observación: ATAS sube/baja por bloques aprox de 1.25 puntos.
# NQ tick = 0.25, entonces 5 ticks = 1.25 puntos.
TICK_SIZE = 0.25
PRICE_GROUP_TICKS = 5

# VWAP ATAS style
DRAW_VWAP = True
VWAP_SESSION_BEGIN = "09:30"
VWAP_SESSION_END = "16:00"
VWAP_BULLISH_COLOR = "2196F3"
VWAP_BEARISH_COLOR = "B22222"
VWAP_NEUTRAL_COLOR = "404040"

OUTPUT_DIR.mkdir(exist_ok=True)


# ============================================================
# HELPERS
# ============================================================

def normalize_columns(df):
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(" ", "_", regex=False)
        .str.replace("-", "_", regex=False)
        .str.replace(".", "_", regex=False)
    )
    return df


def find_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def detect_date_from_filename(path):
    m = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    return m.group(1) if m else path.stem


def read_csv_safe(path):
    for enc in ["utf-8-sig", "utf-8", "latin1"]:
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception:
            pass
    return pd.read_csv(path)


def compress_price(price):
    """
    Agrupa precios en bloques tipo ATAS.

    PRICE_GROUP_TICKS = 5
    TICK_SIZE = 0.25
    group_size = 1.25 puntos
    """
    group_size = PRICE_GROUP_TICKS * TICK_SIZE
    return round(round(float(price) / group_size) * group_size, 2)


# ============================================================
# READ RAW
# ============================================================

def read_raw_file(path):
    if path.suffix.lower() == ".csv":
        df = read_csv_safe(path)
    elif path.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(path)
    else:
        raise ValueError(f"Formato no soportado: {path.name}")

    df = normalize_columns(df)
    print(f"  Columnas detectadas: {list(df.columns)}")

    date_col = find_col(df, ["date", "fecha", "date_ny", "date_utc"])
    time_col = find_col(df, ["time", "hora", "time_ny", "time_utc", "datetime", "timestamp", "ts_event"])
    price_col = find_col(df, ["price", "precio", "trade_price", "level", "nivel_de_precio"])
    bid_col = find_col(df, ["bid", "bid_volume", "bidvol", "bid_size"])
    ask_col = find_col(df, ["ask", "ask_volume", "askvol", "ask_size"])

    open_col = find_col(df, ["candle_open", "open"])
    high_col = find_col(df, ["candle_high", "high"])
    low_col = find_col(df, ["candle_low", "low"])
    close_col = find_col(df, ["candle_close", "close"])
    volume_col = find_col(df, ["candle_volume", "volume", "vol", "total_volume", "level_volume"])

    if time_col is None or price_col is None or bid_col is None or ask_col is None:
        raise ValueError("Faltan columnas necesarias: time/date, price, bid, ask.")

    if date_col:
        df["_dt"] = pd.to_datetime(
            df[date_col].astype(str) + " " + df[time_col].astype(str),
            errors="coerce"
        )
    else:
        df["_dt"] = pd.to_datetime(df[time_col], errors="coerce")

    df["_price"] = pd.to_numeric(df[price_col], errors="coerce")
    df["_bid"] = pd.to_numeric(df[bid_col], errors="coerce").fillna(0)
    df["_ask"] = pd.to_numeric(df[ask_col], errors="coerce").fillna(0)

    df = df.dropna(subset=["_dt", "_price"])
    df["_minute"] = df["_dt"].dt.strftime("%H:%M")

    df = df[(df["_minute"] >= START_TIME) & (df["_minute"] <= END_TIME)].copy()

    if open_col and high_col and low_col and close_col:
        df["_open"] = pd.to_numeric(df[open_col], errors="coerce")
        df["_high"] = pd.to_numeric(df[high_col], errors="coerce")
        df["_low"] = pd.to_numeric(df[low_col], errors="coerce")
        df["_close"] = pd.to_numeric(df[close_col], errors="coerce")
    else:
        df["_open"] = None
        df["_high"] = None
        df["_low"] = None
        df["_close"] = None

    if volume_col:
        df["_volume"] = pd.to_numeric(df[volume_col], errors="coerce").fillna(0)
    else:
        df["_volume"] = df["_bid"] + df["_ask"]

    return df


# ============================================================
# IMBALANCES
# ============================================================

def get_level(matrix, price, minute):
    item = matrix.get((price, minute), {"bid": 0, "ask": 0})
    return int(item["bid"]), int(item["ask"])


def group_indices(indices, allow_gap=0):
    if not indices:
        return []

    indices = sorted(indices)
    groups = []
    current = [indices[0]]

    for idx in indices[1:]:
        if idx - current[-1] <= allow_gap + 1:
            current.append(idx)
        else:
            groups.append(current)
            current = [idx]

    groups.append(current)
    return groups


def passes_volume_filter(side, bid, ask, mode):
    if mode == "winner":
        winner_volume = bid if side == "sell" else ask
        return winner_volume >= IMBALANCE_MIN_VOLUME

    if mode == "level":
        return (bid + ask) >= IMBALANCE_MIN_VOLUME

    if mode == "atas_compare":
        # ATAS marks levels like 30 x 57 even though level total is 87,
        # so Volume filter is not a per-level minimum in this export.
        return True

    raise ValueError(f"Modo de filtro no soportado: {mode}")


def detect_diagonal_imbalances(prices, minutes, matrix, volume_filter_mode="winner"):
    """
    V32 - ATAS clean mode.

    Orientación diagonal:

    SELL:
        bid_actual >= ask_superior * ratio

    BUY:
        ask_actual >= bid_inferior * ratio

    Volume filter modes:
    - winner: el lado ganador debe ser >= IMBALANCE_MIN_VOLUME.
    - level: el volumen total del nivel debe ser >= IMBALANCE_MIN_VOLUME.
    """

    imbalances = {}

    for minute in minutes:
        raw_buy = []
        raw_sell = []
        raw_info = {}

        for i, price in enumerate(prices):
            bid, ask = get_level(matrix, price, minute)

            if bid == 0 and ask == 0:
                continue

            upper_bid = upper_ask = 0
            lower_bid = lower_ask = 0

            if i > 0:
                upper_price = prices[i - 1]
                upper_bid, upper_ask = get_level(matrix, upper_price, minute)

            if i < len(prices) - 1:
                lower_price = prices[i + 1]
                lower_bid, lower_ask = get_level(matrix, lower_price, minute)

            # SELL diagonal: Bid actual vs Ask superior
            sell_ratio = bid / upper_ask if upper_ask > 0 else 0

            if (
                upper_ask > 0
                and upper_ask >= DIAGONAL_COMPARE_MIN_VOLUME
                and bid >= upper_ask * IMBALANCE_RATIO
                and sell_ratio >= IMBALANCE_RATIO
                and passes_volume_filter("sell", bid, ask, volume_filter_mode)
            ):
                raw_sell.append(i)
                raw_info[("sell", i)] = {
                    "side": "sell",
                    "bid": bid,
                    "ask": ask,
                    "upper_bid": upper_bid,
                    "upper_ask": upper_ask,
                    "lower_bid": lower_bid,
                    "lower_ask": lower_ask,
                    "comparison": f"SELL {bid} >= {upper_ask} x {IMBALANCE_RATIO} | ratio={sell_ratio:.2f}",
                }

            # BUY diagonal: Ask actual vs Bid inferior
            buy_ratio = ask / lower_bid if lower_bid > 0 else 0

            if (
                lower_bid > 0
                and lower_bid >= DIAGONAL_COMPARE_MIN_VOLUME
                and ask >= lower_bid * IMBALANCE_RATIO
                and buy_ratio >= IMBALANCE_RATIO
                and passes_volume_filter("buy", bid, ask, volume_filter_mode)
            ):
                raw_buy.append(i)
                raw_info[("buy", i)] = {
                    "side": "buy",
                    "bid": bid,
                    "ask": ask,
                    "upper_bid": upper_bid,
                    "upper_ask": upper_ask,
                    "lower_bid": lower_bid,
                    "lower_ask": lower_ask,
                    "comparison": f"BUY {ask} >= {lower_bid} x {IMBALANCE_RATIO} | ratio={buy_ratio:.2f}",
                }

        for side, raw_indices in [("buy", raw_buy), ("sell", raw_sell)]:
            groups = group_indices(raw_indices, ALLOW_GAP_LEVELS)

            for group in groups:
                if len(group) < MIN_STACK_RAW_IMBALANCES:
                    continue

                start = min(group)
                end = max(group)
                span = end - start + 1

                if span > MAX_STACK_SPAN_LEVELS:
                    continue

                # V26 CLAVE:
                # Antes se hacía list(range(start, end + 1)) y eso rellenaba huecos.
                # Ahora SOLO se pintan los niveles reales detectados.
                zone_indices = group

                zone_max_winner_volume = 0
                zone_max_total_volume = 0

                for idx in zone_indices:
                    price = prices[idx]
                    bid, ask = get_level(matrix, price, minute)

                    zone_max_total_volume = max(zone_max_total_volume, bid + ask)

                    if side == "sell":
                        zone_max_winner_volume = max(zone_max_winner_volume, bid)
                    else:
                        zone_max_winner_volume = max(zone_max_winner_volume, ask)

                if not any(
                    passes_volume_filter(
                        side,
                        *get_level(matrix, prices[idx], minute),
                        volume_filter_mode
                    )
                    for idx in zone_indices
                ):
                    continue

                for idx in zone_indices:
                    price = prices[idx]
                    bid, ask = get_level(matrix, price, minute)
                    key = (price, minute)

                    if bid == 0 and ask == 0:
                        continue

                    base_info = raw_info.get((side, idx), {})

                    if key not in imbalances:
                        imbalances[key] = {
                            "buy": False,
                            "sell": False,
                            "bid": bid,
                            "ask": ask,
                            "upper_bid": base_info.get("upper_bid", 0),
                            "upper_ask": base_info.get("upper_ask", 0),
                            "lower_bid": base_info.get("lower_bid", 0),
                            "lower_ask": base_info.get("lower_ask", 0),
                            "level_volume": bid + ask,
                            "zone_max_total_volume": zone_max_total_volume,
                            "zone_max_winner_volume": zone_max_winner_volume,
                            "raw_group_size": len(group),
                            "span": span,
                            "is_zone_fill": False,
                            "comparison": base_info.get("comparison", ""),
                        }

                    if side == "buy":
                        imbalances[key]["buy"] = True
                    else:
                        imbalances[key]["sell"] = True

    return imbalances


def collect_imbalance_candidates(prices, minutes, matrix):
    candidates = []

    for minute in minutes:
        for i, price in enumerate(prices):
            bid, ask = get_level(matrix, price, minute)

            if bid == 0 and ask == 0:
                continue

            upper_price = prices[i - 1] if i > 0 else None
            lower_price = prices[i + 1] if i < len(prices) - 1 else None

            upper_bid, upper_ask = get_level(matrix, upper_price, minute) if upper_price is not None else (0, 0)
            lower_bid, lower_ask = get_level(matrix, lower_price, minute) if lower_price is not None else (0, 0)

            level_volume = bid + ask

            if upper_ask > 0:
                sell_ratio = bid / upper_ask
                if sell_ratio >= IMBALANCE_RATIO:
                    candidates.append({
                        "minute": minute,
                        "price": price,
                        "side": "SELL",
                        "value": f"{bid} x {ask}",
                        "bid": bid,
                        "ask": ask,
                        "level_volume": level_volume,
                        "compare_price": upper_price,
                        "compare_bid": upper_bid,
                        "compare_ask": upper_ask,
                        "compare_volume": upper_bid + upper_ask,
                        "ratio": round(sell_ratio, 4),
                        "passes_winner": passes_volume_filter("sell", bid, ask, "winner"),
                        "passes_level": passes_volume_filter("sell", bid, ask, "level"),
                    })

            if lower_bid > 0:
                buy_ratio = ask / lower_bid
                if buy_ratio >= IMBALANCE_RATIO:
                    candidates.append({
                        "minute": minute,
                        "price": price,
                        "side": "BUY",
                        "value": f"{bid} x {ask}",
                        "bid": bid,
                        "ask": ask,
                        "level_volume": level_volume,
                        "compare_price": lower_price,
                        "compare_bid": lower_bid,
                        "compare_ask": lower_ask,
                        "compare_volume": lower_bid + lower_ask,
                        "ratio": round(buy_ratio, 4),
                        "passes_winner": passes_volume_filter("buy", bid, ask, "winner"),
                        "passes_level": passes_volume_filter("buy", bid, ask, "level"),
                    })

    return candidates


# ============================================================
# BUILD FOOTPRINT
# ============================================================

def build_footprint(df, volume_filter_mode="winner"):
    df = df.copy()

    # V26: comprimir precios para acercar visualmente el ladder a ATAS
    df["_compressed_price"] = df["_price"].apply(compress_price)

    grouped = (
        df.groupby(["_compressed_price", "_minute"], as_index=False)
        .agg(bid=("_bid", "sum"), ask=("_ask", "sum"))
    )

    grouped.rename(columns={"_compressed_price": "_price"}, inplace=True)

    minutes = sorted(grouped["_minute"].unique())
    prices = sorted(grouped["_price"].unique(), reverse=True)

    matrix = {}
    for _, row in grouped.iterrows():
        price = float(row["_price"])
        minute = row["_minute"]
        matrix[(price, minute)] = {
            "bid": int(row["bid"]),
            "ask": int(row["ask"]),
        }

    summary = {}
    for minute in minutes:
        temp = grouped[grouped["_minute"] == minute]
        bid_sum = int(temp["bid"].sum())
        ask_sum = int(temp["ask"].sum())

        summary[minute] = {
            "delta": ask_sum - bid_sum,
            "volume": ask_sum + bid_sum,
        }

    candles = {}

    candle_df = (
        df.groupby("_minute", as_index=False)
        .agg(
            open=("_open", "first"),
            high=("_high", "max"),
            low=("_low", "min"),
            close=("_close", "last"),
        )
    )

    for _, row in candle_df.iterrows():
        if pd.isna(row["open"]) or pd.isna(row["high"]) or pd.isna(row["low"]) or pd.isna(row["close"]):
            continue

        # V26: también comprimimos OHLC para que el body/wick calce con la escala comprimida
        o = compress_price(float(row["open"]))
        h = compress_price(float(row["high"]))
        l = compress_price(float(row["low"]))
        c = compress_price(float(row["close"]))

        candles[row["_minute"]] = {
            "open": o,
            "high": h,
            "low": l,
            "close": c,
            "body_high": max(o, c),
            "body_low": min(o, c),
            "direction": "bullish" if c >= o else "bearish",
        }

    vwap = {}
    vwap_df = (
        df[(df["_minute"] >= VWAP_SESSION_BEGIN[:5]) & (df["_minute"] <= VWAP_SESSION_END[:5])]
        .groupby("_minute", as_index=False)
        .agg(
            close=("_close", "last"),
            volume=("_volume", "first"),
        )
        .sort_values("_minute")
    )

    cumulative_pv = 0.0
    cumulative_volume = 0.0
    previous_vwap = None

    for _, row in vwap_df.iterrows():
        if pd.isna(row["close"]) or pd.isna(row["volume"]) or float(row["volume"]) <= 0:
            continue

        close = float(row["close"])
        volume = float(row["volume"])
        cumulative_pv += close * volume
        cumulative_volume += volume

        if cumulative_volume <= 0:
            continue

        raw_vwap = cumulative_pv / cumulative_volume
        compressed_vwap = compress_price(raw_vwap)
        direction = "neutral"

        if previous_vwap is not None:
            if raw_vwap > previous_vwap:
                direction = "bullish"
            elif raw_vwap < previous_vwap:
                direction = "bearish"

        vwap[row["_minute"]] = {
            "raw": raw_vwap,
            "price": compressed_vwap,
            "direction": direction,
        }

        previous_vwap = raw_vwap

    imbalances = detect_diagonal_imbalances(prices, minutes, matrix, volume_filter_mode)
    candidates = collect_imbalance_candidates(prices, minutes, matrix)
    return prices, minutes, matrix, summary, candles, imbalances, candidates, vwap


# ============================================================
# EXCEL STYLE
# ============================================================

def make_body_border(price, candle):
    color = "00B050" if candle["direction"] == "bullish" else "C00000"
    thick = Side(style="medium", color=color)
    thin = Side(style="thin", color="B7B7B7")

    top = thick if abs(price - candle["body_high"]) < 1e-9 else thin
    bottom = thick if abs(price - candle["body_low"]) < 1e-9 else thin

    return Border(left=thick, right=thick, top=top, bottom=bottom)


def make_imbalance_border(imbalance):
    if imbalance["buy"] and imbalance["sell"]:
        side = Side(style="thick", color="7030A0")
    elif imbalance["buy"]:
        side = Side(style="thick", color="00AEEF")
    else:
        side = Side(style="thick", color="FF00FF")

    return Border(left=side, right=side, top=side, bottom=side)


def make_vwap_border(existing_border, direction):
    if direction == "bullish":
        color = VWAP_BULLISH_COLOR
    elif direction == "bearish":
        color = VWAP_BEARISH_COLOR
    else:
        color = VWAP_NEUTRAL_COLOR

    vwap_side = Side(style="thick", color=color)

    return Border(
        left=existing_border.left,
        right=existing_border.right,
        top=vwap_side,
        bottom=vwap_side,
    )


# ============================================================
# WRITE EXCEL
# ============================================================

def write_excel(output_path, title, prices, minutes, matrix, summary, candles, imbalances, candidates=None, vwap=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Footprint"

    white = "FFFFFF"
    black = "000000"
    header = "D9EAF7"

    bullish_body = "C6EFCE"
    bearish_body = "FFC7CE"
    wick_gray = "F2F2F2"

    buy_imbalance_fill = "00E5FF"
    sell_imbalance_fill = "FF00FF"
    both_imbalance_fill = "7030A0"

    thin_side = Side(style="thin", color="B7B7B7")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    vwap_price_by_minute = {}
    if DRAW_VWAP and vwap:
        for minute, item in vwap.items():
            if not prices:
                continue
            nearest_price = min(prices, key=lambda p: abs(float(p) - float(item["price"])))
            vwap_price_by_minute[minute] = {
                "price": nearest_price,
                "raw": item["raw"],
                "direction": item["direction"],
            }

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color=black, size=12)
    ws["A1"].fill = PatternFill("solid", fgColor=header)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(minutes) + 1))

    ws["A2"] = "Precio"
    ws["A2"].font = Font(bold=True, color=black)
    ws["A2"].fill = PatternFill("solid", fgColor=header)

    for col_idx, minute in enumerate(minutes, start=2):
        cell = ws.cell(row=2, column=col_idx, value=minute)
        cell.font = Font(bold=True, color=black)
        cell.fill = PatternFill("solid", fgColor=header)

    start_row = 3

    for row_idx, price in enumerate(prices, start=start_row):
        price_cell = ws.cell(row=row_idx, column=1, value=float(price))
        price_cell.font = Font(bold=True, color=black)
        price_cell.fill = PatternFill("solid", fgColor=white)
        price_cell.border = thin_border

        for col_idx, minute in enumerate(minutes, start=2):
            data = matrix.get((price, minute), {"bid": 0, "ask": 0})
            bid = int(data["bid"])
            ask = int(data["ask"])
            imbalance = imbalances.get((price, minute))

            value = "" if bid == 0 and ask == 0 else f"{bid} x {ask}"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = Font(color=black, size=9)
            cell.fill = PatternFill("solid", fgColor=white)

            candle = candles.get(minute)

            if candle:
                in_candle = candle["low"] <= price <= candle["high"]
                in_body = candle["body_low"] <= price <= candle["body_high"]

                if in_candle and not in_body:
                    cell.fill = PatternFill("solid", fgColor=wick_gray)
                elif in_body:
                    body_color = bullish_body if candle["direction"] == "bullish" else bearish_body
                    cell.fill = PatternFill("solid", fgColor=body_color)
                    cell.border = make_body_border(price, candle)

            if imbalance:
                if imbalance["buy"] and imbalance["sell"]:
                    cell.fill = PatternFill("solid", fgColor=both_imbalance_fill)
                    cell.font = Font(color="FFFFFF", bold=True, size=9)
                elif imbalance["buy"]:
                    cell.fill = PatternFill("solid", fgColor=buy_imbalance_fill)
                    cell.font = Font(color=black, bold=True, size=9)
                elif imbalance["sell"]:
                    cell.fill = PatternFill("solid", fgColor=sell_imbalance_fill)
                    cell.font = Font(color="FFFFFF", bold=True, size=9)

                cell.border = make_imbalance_border(imbalance)

                cell.comment = Comment(
                    f"ATAS diagonal imbalance V34\n\n"
                    f"Bid actual: {imbalance['bid']}\n"
                    f"Ask actual: {imbalance['ask']}\n"
                    f"Volumen nivel: {imbalance['level_volume']}\n"
                    f"Volumen max ganador zona: {imbalance['zone_max_winner_volume']}\n"
                    f"Volumen max total zona: {imbalance['zone_max_total_volume']}\n\n"
                    f"Comparación: {imbalance.get('comparison', '')}\n"
                    f"Raw group size: {imbalance.get('raw_group_size')}\n"
                    f"Span: {imbalance.get('span')}\n"
                    f"Zone fill intermedio: {imbalance.get('is_zone_fill')}\n\n"
                    f"Ratio: {IMBALANCE_RATIO}\n"
                    f"Volume filter: {IMBALANCE_MIN_VOLUME}\n"
                    f"Price group ticks: {PRICE_GROUP_TICKS}",
                    "05_footprint_generator_v34"
                )

            vwap_item = vwap_price_by_minute.get(minute)
            if vwap_item and abs(float(price) - float(vwap_item["price"])) < 1e-9:
                cell.border = make_vwap_border(cell.border, vwap_item["direction"])

                vwap_text = (
                    f"VWAP\n"
                    f"Raw: {vwap_item['raw']:.2f}\n"
                    f"Plotted price: {vwap_item['price']}\n"
                    f"Direction: {vwap_item['direction']}"
                )

                if cell.comment:
                    cell.comment = Comment(cell.comment.text + "\n\n" + vwap_text, cell.comment.author)
                else:
                    cell.comment = Comment(vwap_text, "05_footprint_generator_v34")

    vwap_row = start_row + len(prices) + 2

    ws.cell(row=vwap_row, column=1, value="VWAP")
    ws.cell(row=vwap_row, column=1).font = Font(bold=True, color=black)
    ws.cell(row=vwap_row, column=1).fill = PatternFill("solid", fgColor=header)
    ws.cell(row=vwap_row, column=1).border = thin_border

    for col_idx, minute in enumerate(minutes, start=2):
        cell = ws.cell(row=vwap_row, column=col_idx)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        vwap_item = vwap.get(minute) if vwap else None
        if vwap_item:
            cell.value = round(vwap_item["raw"], 2)

            if vwap_item["direction"] == "bullish":
                cell.fill = PatternFill("solid", fgColor=VWAP_BULLISH_COLOR)
                cell.font = Font(color="FFFFFF", bold=True, size=9)
            elif vwap_item["direction"] == "bearish":
                cell.fill = PatternFill("solid", fgColor=VWAP_BEARISH_COLOR)
                cell.font = Font(color="FFFFFF", bold=True, size=9)
            else:
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(color=black, bold=True, size=9)
        else:
            cell.fill = PatternFill("solid", fgColor="F2F2F2")

    summary_row = vwap_row + 2

    ws.cell(row=summary_row, column=1, value="DELTA / VOLUMEN POR MINUTO")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, color=black)

    header_row = summary_row + 1

    ws.cell(row=header_row, column=1, value="Métrica")
    ws.cell(row=header_row, column=1).font = Font(bold=True, color=black)
    ws.cell(row=header_row, column=1).fill = PatternFill("solid", fgColor=header)

    for col_idx, minute in enumerate(minutes, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=minute)
        cell.font = Font(bold=True, color=black)
        cell.fill = PatternFill("solid", fgColor=header)

    ws.cell(row=header_row + 1, column=1, value="Delta")
    ws.cell(row=header_row + 2, column=1, value="Volumen")

    for col_idx, minute in enumerate(minutes, start=2):
        delta = summary[minute]["delta"]

        delta_cell = ws.cell(row=header_row + 1, column=col_idx, value=delta)
        vol_cell = ws.cell(row=header_row + 2, column=col_idx, value=summary[minute]["volume"])

        if delta > 0:
            delta_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif delta < 0:
            delta_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        else:
            delta_cell.fill = PatternFill("solid", fgColor="F2F2F2")

        vol_cell.fill = PatternFill("solid", fgColor="D9D9D9")

    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = 13

    for col_idx in range(2, len(minutes) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    if candidates is not None:
        diag_ws = wb.create_sheet("ImbalanceCandidates")
        headers = [
            "minute",
            "price",
            "side",
            "value",
            "bid",
            "ask",
            "level_volume",
            "compare_price",
            "compare_bid",
            "compare_ask",
            "compare_volume",
            "ratio",
            "passes_winner",
            "passes_level",
            "painted",
        ]

        for col_idx, header_text in enumerate(headers, start=1):
            cell = diag_ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = Font(bold=True, color=black)
            cell.fill = PatternFill("solid", fgColor=header)
            cell.border = thin_border

        for row_idx, candidate in enumerate(candidates, start=2):
            painted = bool(imbalances.get((candidate["price"], candidate["minute"])))
            row_values = [
                candidate["minute"],
                candidate["price"],
                candidate["side"],
                candidate["value"],
                candidate["bid"],
                candidate["ask"],
                candidate["level_volume"],
                candidate["compare_price"],
                candidate["compare_bid"],
                candidate["compare_ask"],
                candidate["compare_volume"],
                candidate["ratio"],
                candidate["passes_winner"],
                candidate["passes_level"],
                painted,
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = diag_ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            side_fill = buy_imbalance_fill if candidate["side"] == "BUY" else sell_imbalance_fill
            diag_ws.cell(row=row_idx, column=3).fill = PatternFill("solid", fgColor=side_fill)
            if candidate["side"] == "SELL":
                diag_ws.cell(row=row_idx, column=3).font = Font(color="FFFFFF", bold=True)

            if painted:
                diag_ws.cell(row=row_idx, column=15).fill = PatternFill("solid", fgColor="C6EFCE")
            else:
                diag_ws.cell(row=row_idx, column=15).fill = PatternFill("solid", fgColor="FFC7CE")

        diag_ws.freeze_panes = "A2"
        widths = [10, 12, 10, 12, 10, 10, 14, 14, 12, 12, 16, 10, 14, 12, 10]
        for idx, width in enumerate(widths, start=1):
            diag_ws.column_dimensions[get_column_letter(idx)].width = width

    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 80)
    print("05 FOOTPRINT GENERATOR - V34 ATAS COMPARE FILTER + DIAGNOSTICS")
    print("=" * 80)

    files = [
        f for f in RAW_DIR.iterdir()
        if f.is_file()
        and f.suffix.lower() in [".csv", ".xlsx", ".xls"]
        and not f.name.startswith("~$")
        and (PROCESS_ONLY_DATE is None or PROCESS_ONLY_DATE in f.name)
    ]

    if PROCESS_ONLY_DATE:
        print(f"\nArchivos encontrados para {PROCESS_ONLY_DATE}: {len(files)}")
    else:
        print(f"\nArchivos encontrados para procesar: {len(files)}")

    for f in files:
        print(f"  - {f.name}")

    if not files:
        print("\nNo encontré archivos para procesar.")
        return

    for path in sorted(files):
        try:
            print("\n" + "-" * 80)
            print(f"Procesando: {path.name}")

            df = read_raw_file(path)

            if df.empty:
                print("  Saltado: no hay datos dentro del rango horario.")
                continue

            date_label = detect_date_from_filename(path)

            for volume_filter_mode in VOLUME_FILTER_MODES:
                prices, minutes, matrix, summary, candles, imbalances, candidates, vwap = build_footprint(
                    df,
                    volume_filter_mode=volume_filter_mode,
                )

                mode_label = volume_filter_mode.upper()
                title = (
                    f"MINI ATAS V34 — FILTER {mode_label} "
                    f"RATIO {IMBALANCE_RATIO} VOL {IMBALANCE_MIN_VOLUME} "
                    f"GROUP {PRICE_GROUP_TICKS}T — {date_label}"
                )

                output_path = OUTPUT_DIR / (
                    f"MINI_ATAS_FOOTPRINT_{date_label}_V34_FILTER_{mode_label}_DIAG.xlsx"
                )

                write_excel(output_path, title, prices, minutes, matrix, summary, candles, imbalances, candidates, vwap)

                print(f"  OK generado ({mode_label}): {output_path}")
                print(f"  Imbalances pintados ({mode_label}): {len(imbalances)}")

                # Debug rápido para la vela que estamos comparando
                target_minute = "09:31"
                target_imbalances = [
                    (price, info)
                    for (price, minute), info in imbalances.items()
                    if minute == target_minute
                ]

                print(f"  Imbalances {target_minute} ({mode_label}): {len(target_imbalances)}")
                for price, info in sorted(target_imbalances, key=lambda x: x[0], reverse=True):
                    side = "BOTH" if info["buy"] and info["sell"] else "BUY" if info["buy"] else "SELL"
                    print(f"    {target_minute} | {price} | {side} | {info['bid']} x {info['ask']} | {info.get('comparison', '')}")

        except Exception as e:
            print(f"  ERROR en {path.name}: {e}")


if __name__ == "__main__":
    main()
