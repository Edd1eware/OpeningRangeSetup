from pywinauto import Desktop
import csv
import os
import time
import pyperclip
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG
# =========================================================

# Fechas operables en horario DST de Nueva York 2026.
# Formato requerido por el panel Replay de ATAS: dd/mm/yyyy.
DATES_DST = [

    # MARZO 2025
    "10/03/2025",
    "11/03/2025",
    "12/03/2025",
    "13/03/2025",
    "14/03/2025",
    "17/03/2025",
    "18/03/2025",
    "19/03/2025",
    "20/03/2025",
    "21/03/2025",
    "24/03/2025",
    "25/03/2025",
    "26/03/2025",
    "27/03/2025",
    "28/03/2025",
    "31/03/2025",

    # ABRIL 2025
    "01/04/2025",
    "02/04/2025",
    "03/04/2025",
    "04/04/2025",
    "07/04/2025",
    "08/04/2025",
    "09/04/2025",
    "10/04/2025",
    "11/04/2025",
    "14/04/2025",
    "15/04/2025",
    "16/04/2025",
    "17/04/2025",
    "21/04/2025",
    "22/04/2025",
    "23/04/2025",
    "24/04/2025",
    "25/04/2025",
    "28/04/2025",
    "29/04/2025",
    "30/04/2025",

    # MAYO 2025
    "01/05/2025",
    "02/05/2025",
    "05/05/2025",
    "06/05/2025",
    "07/05/2025",
    "08/05/2025",
    "09/05/2025",
    "12/05/2025",
    "13/05/2025",
    "14/05/2025",
    "15/05/2025",
    "16/05/2025",
    "19/05/2025",
    "20/05/2025",
    "21/05/2025",
    "22/05/2025",
    "23/05/2025",
    "27/05/2025",
    "28/05/2025",
    "29/05/2025",
    "30/05/2025",

    # JUNIO 2025
    "02/06/2025",
    "03/06/2025",
    "04/06/2025",
    "05/06/2025",
    "06/06/2025",
    "09/06/2025",
    "10/06/2025",
    "11/06/2025",
    "12/06/2025",
    "13/06/2025",
    "16/06/2025",
    "17/06/2025",
    "18/06/2025",
    "20/06/2025",
    "23/06/2025",
    "24/06/2025",
    "25/06/2025",
    "26/06/2025",
    "27/06/2025",
    "30/06/2025",

    # JULIO 2025
    "01/07/2025",
    "02/07/2025",
    "03/07/2025",
    "07/07/2025",
    "08/07/2025",
    "09/07/2025",
    "10/07/2025",
    "11/07/2025",
    "14/07/2025",
    "15/07/2025",
    "16/07/2025",
    "17/07/2025",
    "18/07/2025",
    "21/07/2025",
    "22/07/2025",
    "23/07/2025",
    "24/07/2025",
    "25/07/2025",
    "28/07/2025",
    "29/07/2025",
    "30/07/2025",
    "31/07/2025",

    # AGOSTO 2025
    "01/08/2025",
    "04/08/2025",
    "05/08/2025",
    "06/08/2025",
    "07/08/2025",
    "08/08/2025",
    "11/08/2025",
    "12/08/2025",
    "13/08/2025",
    "14/08/2025",
    "15/08/2025",
    "18/08/2025",
    "19/08/2025",
    "20/08/2025",
    "21/08/2025",
    "22/08/2025",
    "25/08/2025",
    "26/08/2025",
    "27/08/2025",
    "28/08/2025",
    "29/08/2025",

    # SEPTIEMBRE 2025
    "02/09/2025",
    "03/09/2025",
    "04/09/2025",
    "05/09/2025",
    "08/09/2025",
    "09/09/2025",
    "10/09/2025",
    "11/09/2025",
    "12/09/2025",
    "15/09/2025",
    "16/09/2025",
    "17/09/2025",
    "18/09/2025",
    "19/09/2025",
    "22/09/2025",
    "23/09/2025",
    "24/09/2025",
    "25/09/2025",
    "26/09/2025",
    "29/09/2025",
    "30/09/2025",

    # OCTUBRE 2025
    "01/10/2025",
    "02/10/2025",
    "03/10/2025",
    "06/10/2025",
    "07/10/2025",
    "08/10/2025",
    "09/10/2025",
    "10/10/2025",
    "13/10/2025",
    "14/10/2025",
    "15/10/2025",
    "16/10/2025",
    "17/10/2025",
    "20/10/2025",
    "21/10/2025",
    "22/10/2025",
    "23/10/2025",
    "24/10/2025",
    "27/10/2025",
    "28/10/2025",
    "29/10/2025",
    "30/10/2025",
    "31/10/2025",
]
# Replay recomendado para esta prueba: X1.
# Ventana por dia: 09:30 a 09:50 NY. El exporter escribe TIME_OVER si no hay trade antes/de 09:40;
# si ya hay trade abierto, dejamos correr hasta 09:50 para que resuelva TP/SL/EXIT/BE.
REPLAY_END_TIME = "09:50"
POLL_SECONDS = 0.02
NO_TRADE_CUTOFF_SECONDS = 10 * 60 + 15
REPLAY_START_MAX_ATTEMPTS = 3
REPLAY_START_TIMEOUT_SECONDS = 2 * 60
REPLAY_START_POLL_SECONDS = 5

EXPORT_FOLDER = r"C:\Users\k_99_\Desktop\codding\data_footprint_generator"
RESULTS_FOLDER = os.path.join(EXPORT_FOLDER, "trade_results_score")
TARGET_FILE = os.path.join(EXPORT_FOLDER, "target_trade_result_date.txt")
REPLAY_STARTED_FILE = os.path.join(EXPORT_FOLDER, "replay_trade_result_started_at.txt")
TELEGRAM_CLEAR_REQUEST_FILE = os.path.join(RESULTS_FOLDER, "telegram_clear_requested.txt")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TESTING_OUTPUT_DIR = r"C:\Users\k_99_\Desktop\codding\corridas_testing_indicator"
SCORE_WORKBOOK_TEMPLATE = os.path.join(BASE_DIR, "Score_indicator_results_updated.xlsx")
SCORE_WORKBOOK_TEMPLATE_FALLBACK = os.path.join(BASE_DIR, "Score_indicator_results_updated_fallback.xlsx")
SCORE_WORKBOOK = os.path.join(TESTING_OUTPUT_DIR, "Score_indicator_results_updated.xlsx")
SCORE_WORKBOOK_FALLBACK = os.path.join(TESTING_OUTPUT_DIR, "Score_indicator_results_updated_fallback.xlsx")
RUN_STARTED_AT = time.time()


# =========================================================
# FUNCIONES
# =========================================================

def write_target_date(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    target = f"{yyyy}-{mm}-{dd}"

    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    os.makedirs(RESULTS_FOLDER, exist_ok=True)

    with open(TARGET_FILE, "w", encoding="utf-8") as f:
        f.write(target)

    print(f"Fecha objetivo escrita para ATAS: {target}")


def write_telegram_clear_request():
    os.makedirs(RESULTS_FOLDER, exist_ok=True)
    with open(TELEGRAM_CLEAR_REQUEST_FILE, "w", encoding="utf-8") as f:
        f.write(str(time.time()))
    print("Solicitud de limpieza de Telegram escrita (se borra conversacion anterior al inicio).")


def write_replay_started_marker():
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    with open(REPLAY_STARTED_FILE, "w", encoding="utf-8") as f:
        f.write(str(time.time()))
    print("Marcador de inicio de replay escrito.")


def get_replay():
    desktop = Desktop(backend="uia")
    candidates = desktop.windows(title_re=".*Replay.*", visible_only=True)

    if not candidates:
        raise RuntimeError("No encontre ninguna ventana Replay visible. Abre Replay en ATAS antes de correr el script.")

    print(f"Ventanas Replay encontradas: {len(candidates)}")

    for w in candidates:
        try:
            if w.is_visible() and w.is_enabled():
                print("Usando Replay:", w.window_text())
                w.set_focus()
                time.sleep(1)
                return w
        except Exception:
            pass

    replay = candidates[0]
    replay.set_focus()
    time.sleep(1)
    return replay


def paste_text(control, value):
    control.click_input()
    time.sleep(0.3)

    control.type_keys("^a")
    time.sleep(0.3)

    pyperclip.copy(value)
    time.sleep(0.3)

    control.type_keys("^v")
    time.sleep(0.8)


def get_controls():
    replay = get_replay()

    edits = replay.descendants(control_type="Edit")
    buttons = replay.descendants(control_type="Button")

    from_box = edits[0]
    current_box = edits[1]
    to_box = edits[2]

    start_button = None
    stop_button = None

    for b in buttons:
        txt = b.window_text()

        if txt == "Start":
            start_button = b

        if txt == "Stop":
            stop_button = b

    return replay, from_box, current_box, to_box, start_button, stop_button


def expected_result_path(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")

    return os.path.join(
        RESULTS_FOLDER,
        f"score_trade_result_{yyyy}-{mm}-{dd}_NY.csv"
    )


def print_result_file(path):
    if not os.path.exists(path):
        print("WARNING: no encontre archivo esperado:")
        print(path)
        return

    size_kb = round(os.path.getsize(path) / 1024, 2)
    print(f"OK EXPORTADO: {path}")
    print(f"Tamano: {size_kb} KB")

    with open(path, "r", encoding="utf-8") as f:
        print(f.read().strip())


def clear_previous_result(path):
    if os.path.exists(path):
        os.remove(path)
        print(f"Resultado anterior eliminado: {path}")


def clear_expected_results():
    for date in DATES_DST:
        clear_previous_result(expected_result_path(date))


def parse_result_ticks(value):
    if value in (None, "", "OPEN", "NO_TRADE"):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    normalized = str(value).strip().upper()

    if normalized in ("TP", "EXIT"):
        return 1.0

    if normalized == "SL":
        return -1.0

    if normalized == "BE":
        return 0.0

    try:
        return float(normalized.replace("+", ""))
    except ValueError:
        return None


def read_result_ticks(path):
    if not os.path.exists(path):
        return None

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return None

    if not row:
        return None

    value = row.get("result TP SL BE") or row.get("RESULT")
    return parse_result_ticks(value)


def result_is_terminal(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    if result_label in ("TP", "SL", "EXIT", "BE", "TIME_OVER", "NO_TRADE"):
        return True

    ticks = parse_result_ticks(row.get("result TP SL BE") or row.get("RESULT"))
    if ticks is None:
        return False

    return ticks != 0


def result_is_open_trade(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    return result_label == "OPEN"


def stop_replay(stop_button=None):
    print("Deteniendo replay...")

    candidates = [stop_button] if stop_button is not None else []

    for _ in range(5):
        try:
            replay, from_box, current_box, to_box, start_button, refreshed_stop_button = get_controls()
            replay.set_focus()
            if refreshed_stop_button is not None:
                candidates.insert(0, refreshed_stop_button)
        except Exception:
            pass

        for candidate in candidates:
            if candidate is None:
                continue

            try:
                if candidate.is_visible() and candidate.is_enabled():
                    candidate.click_input()
                    return
            except Exception:
                try:
                    candidate.click()
                    return
                except Exception:
                    pass

        time.sleep(0.05)

    print("No encontre boton Stop; probablemente el replay ya termino.")


def read_current_box_value(current_box):
    try:
        return current_box.window_text()
    except Exception:
        return None


def wait_for_replay_progress(current_box, timeout_seconds, poll_seconds):
    """Confirma que el replay realmente avanza (campo de tiempo actual cambia).
    Si el valor no se mueve dentro del timeout, asumimos que el Start no
    surtio efecto ("no se reproduce")."""
    initial_value = read_current_box_value(current_box)
    start = time.time()

    while time.time() - start < timeout_seconds:
        time.sleep(poll_seconds)
        current_value = read_current_box_value(current_box)

        if current_value is not None and current_value != initial_value:
            print(f"Replay avanzando (tiempo actual: {current_value}).")
            return True

    print(f"Replay no avanzo en {timeout_seconds}s (valor estancado en '{initial_value}').")
    return False


def start_replay_with_retries(date, result_path, max_attempts, timeout_seconds, poll_seconds):
    """Intenta iniciar el replay hasta max_attempts veces. Devuelve el
    stop_button vigente si el replay quedo confirmado reproduciendo, o
    None si se agotaron los intentos sin avance."""
    for attempt in range(1, max_attempts + 1):
        print(f"Intento de replay {attempt}/{max_attempts} para {date}...")

        replay, from_box, current_box, to_box, start_button, stop_button = get_controls()
        replay.set_focus()

        if start_button is None:
            raise RuntimeError("No se encontro boton Start")

        clear_previous_result(result_path)
        write_replay_started_marker()

        start_button.click_input()

        if wait_for_replay_progress(current_box, timeout_seconds, poll_seconds):
            return stop_button

        print(f"No se reproduce ({date}, intento {attempt}/{max_attempts}). Deteniendo y reintentando...")
        stop_replay(stop_button)
        time.sleep(2)

    print(f"Se agotaron {max_attempts} intentos sin reproducir para {date}. Se cambia de fecha.")
    return None


MAX_WAIT_SECONDS = 1200

def wait_until_result(path, min_modified_time=None, no_trade_cutoff_seconds=None, stop_button=None):
    print("Esperando resultado terminal en CSV; si el trade esta OPEN no se cambia de dia.")
    dot_count = 0
    start = time.time()
    last_print = 0

    while True:
        if result_is_terminal(path, min_modified_time):
            print("\rEsperando... listo.   ")
            print("Resultado terminal detectado en CSV; paso al siguiente dia.")
            stop_replay(stop_button)
            return True

        elapsed = time.time() - start
        if (
            no_trade_cutoff_seconds is not None and
            elapsed > no_trade_cutoff_seconds and
            not result_is_open_trade(path, min_modified_time)
        ):
            print(f"\rCorte 09:40 sin trade OPEN ({int(elapsed)}s). Deteniendo replay del dia.")
            stop_replay(stop_button)
            return False

        if elapsed > MAX_WAIT_SECONDS:
            print(f"\rTimeout ({MAX_WAIT_SECONDS}s) esperando resultado terminal. Saltando al siguiente dia.")
            stop_replay(stop_button)
            return False

        if time.time() - last_print >= 0.5:
            dot_count = dot_count % 3 + 1
            print(f"\rEsperando{'.' * dot_count}{' ' * (3 - dot_count)}", end="", flush=True)
            last_print = time.time()
        time.sleep(POLL_SECONDS)


def read_trade_result(path, date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    default_row = {
        "fecha": f"{yyyy}-{mm}-{dd}",
        "result TP SL BE": "NO_CSV",
    }

    if not os.path.exists(path):
        print(f"Sin CSV para {default_row['fecha']}; no se sobrescriben columnas de datos.")
        return default_row

    if os.path.getmtime(path) < RUN_STARTED_AT:
        print(f"CSV viejo ignorado para {default_row['fecha']}: {path}")
        return default_row

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        try:
            row = next(reader)
        except StopIteration:
            default_row["result TP SL BE"] = "EMPTY_CSV"
            return default_row

    if "fecha" not in row:
        legacy_result = row.get("RESULT", "NO_TRADE")
        legacy_ticks = None

        if legacy_result == "TP" and row.get("ENTRY") and row.get("TP"):
            legacy_ticks = abs((float(row["TP"]) - float(row["ENTRY"])) / 0.25)

        if legacy_result == "SL" and row.get("ENTRY") and row.get("SL"):
            legacy_ticks = -abs((float(row["ENTRY"]) - float(row["SL"])) / 0.25)

        row = {
            "fecha": default_row["fecha"],
            "SL_price": row.get("SL"),
            "Entry_price": row.get("ENTRY"),
            "TP_price": row.get("TP"),
            "result TP SL BE": legacy_ticks if legacy_ticks is not None else legacy_result,
        }

    row["fecha"] = row.get("fecha") or default_row["fecha"]
    row["result TP SL BE"] = row.get("result TP SL BE") or "OPEN"
    row["Exit_price"] = row.get("Exit_price") or calculate_exit_price(row)
    return row


def calculate_exit_price(row):
    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()

    if result_label == "TP":
        return row.get("TP_price") or row.get("TP") or ""

    if result_label == "SL":
        return row.get("SL_price") or row.get("SL") or ""

    if result_label != "EXIT":
        return ""

    return ""


def to_number(value):
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def get_or_create_headers(ws):
    default_headers = [
        "fecha",
        "or_low",
        "or_high",
        "range",
        "VWAP_entry",
        "Body",
        "Volume_entry",
        "Delta_entry",
        "BreakOut_SPEED",
        "BreakOut_TICKS_PER_SEC",
        "score total",
        "SL_price",
        "Entry_price",
        "TP_price",
        "SL_ticks",
        "TP_ticks",
        "Exit_price",
        "result TP SL BE",
        "Side",
        "Speed_Profile",
        "MAE_ticks",
        "MFE_ticks",
    ]

    headers = [
        ws.cell(row=3, column=col).value
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=3, column=col).value
    ]

    if not headers:
        headers = default_headers[:]

    for header in default_headers:
        if header not in headers:
            headers.append(header)

    for csv_header in get_csv_headers_for_dates():
        if csv_header not in headers:
            headers.append(csv_header)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    return headers


def get_csv_headers_for_dates():
    headers = []

    for date in DATES_DST:
        path = expected_result_path(date)

        if not os.path.exists(path):
            continue

        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue

                for field in reader.fieldnames:
                    if field and field not in headers:
                        headers.append(field)
        except OSError:
            continue

    return headers


def update_score_workbook():
    os.makedirs(os.path.dirname(SCORE_WORKBOOK), exist_ok=True)

    if os.path.exists(SCORE_WORKBOOK):
        wb = load_workbook(SCORE_WORKBOOK)
    elif os.path.exists(SCORE_WORKBOOK_TEMPLATE):
        wb = load_workbook(SCORE_WORKBOOK_TEMPLATE)
    elif os.path.exists(SCORE_WORKBOOK_FALLBACK):
        wb = load_workbook(SCORE_WORKBOOK_FALLBACK)
    elif os.path.exists(SCORE_WORKBOOK_TEMPLATE_FALLBACK):
        wb = load_workbook(SCORE_WORKBOOK_TEMPLATE_FALLBACK)
    else:
        wb = Workbook()

    ws = wb.active

    headers = get_or_create_headers(ws)
    first_row = 4
    last_row = first_row + len(DATES_DST) - 1

    for row in range(first_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    for row_offset, date in enumerate(DATES_DST, start=first_row):
        result = read_trade_result(expected_result_path(date), date)
        missing_result_file = result.get("result TP SL BE") in ("NO_CSV", "EMPTY_CSV")

        for col, header in enumerate(headers, start=1):
            if header not in result:
                if missing_result_file:
                    ws.cell(row=row_offset, column=col).value = None
                continue

            value = to_number(result.get(header))

            if value is None and header != "result TP SL BE":
                continue

            ws.cell(row=row_offset, column=col).value = value

    result_col = headers.index("result TP SL BE") + 1
    result_letter = get_column_letter(result_col)

    ws["A1"] = "win rate"
    ws["B1"] = "profit factor"
    ws["A2"] = f'=IF((COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0"))=0,"",COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")/(COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0")))'
    ws["B2"] = f'=IF(ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row}))=0,IF(SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})>0,"INF",""),SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})/ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row})))'
    ws["A2"].number_format = "0.00%"
    ws["B2"].number_format = "0.00"

    for row_idx in range(first_row, last_row + 1):
        ws.cell(row=row_idx, column=result_col).number_format = "+0.##;-0.##;0"

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    widths = {
        "A": 12,
        "B": 12,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 16,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    try:
        wb.save(SCORE_WORKBOOK)
        print(f"Excel actualizado: {SCORE_WORKBOOK}")
    except PermissionError:
        wb.save(SCORE_WORKBOOK_FALLBACK)
        print("WARNING: Excel original bloqueado/abierto; guarde copia actualizada en:")
        print(SCORE_WORKBOOK_FALLBACK)


# =========================================================
# LOOP PRINCIPAL
# =========================================================

print("\nINICIANDO REPLAY DST PARA SCORE TRADE RESULTS\n")
failed_dates = []

try:
    clear_expected_results()
    write_telegram_clear_request()

    for date in DATES_DST:
        print("\n" + "=" * 70)
        print(f"PROCESANDO {date}")
        print("=" * 70)

        result_path = expected_result_path(date)

        try:
            write_target_date(date)
            time.sleep(1)

            from_value = f"{date} 09:30 a. m."
            to_value = f"{date} {REPLAY_END_TIME} a. m."

            replay, from_box, current_box, to_box, start_button, stop_button = get_controls()

            paste_text(from_box, from_value)
            paste_text(to_box, to_value)

            print("Fechas configuradas:")
            print(f"FROM: {from_value}")
            print(f"TO:   {to_value}")

            time.sleep(2)

            started_at = time.time()
            stop_button = start_replay_with_retries(
                date, result_path,
                REPLAY_START_MAX_ATTEMPTS,
                REPLAY_START_TIMEOUT_SECONDS,
                REPLAY_START_POLL_SECONDS,
            )

            if stop_button is None:
                failed_dates.append((date, "no se reproduce tras 3 intentos"))
                print(f"Saltando {date}: no se reproduce tras {REPLAY_START_MAX_ATTEMPTS} intentos.")
                print("Pausa antes del siguiente dia...")
                time.sleep(10)
                continue

            print("Esperando hasta detectar TP/SL/EXIT/BE/TIME_OVER...")
            wait_until_result(result_path, started_at, NO_TRADE_CUTOFF_SECONDS, stop_button)

            time.sleep(5)

            print_result_file(result_path)
        except Exception as exc:
            failed_dates.append((date, str(exc)))
            print(f"ERROR procesando {date}: {exc}")
            try:
                stop_replay()
            except Exception as stop_exc:
                print(f"WARNING: no pude detener replay despues del error: {stop_exc}")

        print("Pausa antes del siguiente dia...")
        time.sleep(10)

finally:
    update_score_workbook()

if failed_dates:
    print("\nFECHAS CON ERROR DE SCRIPT/UI:")
    for failed_date, error in failed_dates:
        print(f"- {failed_date}: {error}")

print("\nTERMINO LA PRUEBA DST.\n")
