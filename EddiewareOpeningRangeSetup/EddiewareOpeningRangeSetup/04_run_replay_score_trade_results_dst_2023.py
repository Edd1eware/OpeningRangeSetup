from pywinauto import Desktop
import csv
import os
import time
from datetime import datetime, timedelta
import pyperclip
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG
# =========================================================

# Fechas operables en horario DST de Nueva York.
# Formato requerido por el panel Replay de ATAS: dd/mm/yyyy.
#
# Rango solicitado:
# - DST 2023 completo
# - DST 2024 completo
# - DST 2025 completo
# - DST 2026 hasta 02/06/2026
#
# NOTA: Solo incluye lunes a viernes.
def generate_weekdays_between(start_ddmmyyyy, end_ddmmyyyy):
    start = datetime.strptime(start_ddmmyyyy, "%d/%m/%Y")
    end = datetime.strptime(end_ddmmyyyy, "%d/%m/%Y")

    dates = []
    current = start

    while current <= end:
        if current.weekday() < 5:  # lunes a viernes
            dates.append(current.strftime("%d/%m/%Y"))

        current += timedelta(days=1)

    return dates


def normalize_replay_date(date_ddmmyyyy):
    return datetime.strptime(date_ddmmyyyy.strip(), "%d/%m/%Y").strftime("%d/%m/%Y")


DATES_DST = []

DATES_DST += generate_weekdays_between("13/03/2023", "03/11/2023")
DATES_DST += generate_weekdays_between("11/03/2024", "01/11/2024")
DATES_DST += generate_weekdays_between("10/03/2025", "31/10/2025")
DATES_DST += generate_weekdays_between("09/03/2026", "02/06/2026")

# Replay recomendado para esta prueba: X1.
# Ventana por dia: 09:30 a 09:50 NY.
# El exporter escribe TIME_OVER a 09:40 si no hay trade; si hay trade abierto,
# estos 10 minutos extra permiten esperar TP/SL.
REPLAY_END_TIME = "09:50"
POLL_SECONDS = 1

EXPORT_FOLDER = r"C:\Users\k_99_\Desktop\codding\data_footprint_generator"
RESULTS_FOLDER = os.path.join(EXPORT_FOLDER, "trade_results_score")
TARGET_FILE = os.path.join(EXPORT_FOLDER, "target_trade_result_date.txt")
REPLAY_STARTED_FILE = os.path.join(EXPORT_FOLDER, "replay_trade_result_started_at.txt")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCORE_WORKBOOK = os.path.join(BASE_DIR, "Score_indicator_results_updated.xlsx")
SCORE_WORKBOOK_FALLBACK = os.path.join(BASE_DIR, "Score_indicator_results_updated_fallback.xlsx")
RUN_STARTED_AT = time.time()


# =========================================================
# FUNCIONES
# =========================================================

def write_target_date(date_ddmmyyyy):
    date_ddmmyyyy = normalize_replay_date(date_ddmmyyyy)
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    target = f"{yyyy}-{mm}-{dd}"

    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    os.makedirs(RESULTS_FOLDER, exist_ok=True)

    with open(TARGET_FILE, "w", encoding="utf-8") as f:
        f.write(target)

    print(f"Fecha objetivo escrita para ATAS: {target}")


def write_replay_started_marker():
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    with open(REPLAY_STARTED_FILE, "w", encoding="utf-8") as f:
        f.write(str(time.time()))
    print("Marcador de inicio de replay escrito.")


def get_replay(verbose=True, focus=True):
    desktop = Desktop(backend="uia")
    candidates = desktop.windows(title_re=".*Replay.*", visible_only=True)

    if not candidates:
        raise RuntimeError("No encontre ninguna ventana Replay visible. Abre Replay en ATAS antes de correr el script.")

    if verbose:
        print(f"Ventanas Replay encontradas: {len(candidates)}")

    for w in candidates:
        try:
            if w.is_visible() and w.is_enabled():
                if verbose:
                    print("Usando Replay:", w.window_text())
                if focus:
                    w.set_focus()
                    time.sleep(1)
                return w
        except Exception:
            pass

    replay = candidates[0]
    if focus:
        replay.set_focus()
        time.sleep(1)
    return replay


def paste_text(control, value):
    control.click_input()
    time.sleep(0.3)

    control.type_keys("^a")
    time.sleep(0.3)

    pyperclip.copy(value)
    time.sleep(0.3)

    control.type_keys("^v")
    time.sleep(0.8)


def get_controls(verbose=True, focus=True):
    replay = get_replay(verbose=verbose, focus=focus)

    edits = replay.descendants(control_type="Edit")
    buttons = replay.descendants(control_type="Button")

    from_box = edits[0]
    to_box = edits[2]

    start_button = None
    stop_button = None

    for b in buttons:
        txt = b.window_text()

        if txt == "Start":
            start_button = b

        if txt == "Stop":
            stop_button = b

    return replay, from_box, to_box, start_button, stop_button


def expected_result_path(date_ddmmyyyy):
    date_ddmmyyyy = normalize_replay_date(date_ddmmyyyy)
    dd, mm, yyyy = date_ddmmyyyy.split("/")

    return os.path.join(
        RESULTS_FOLDER,
        f"score_trade_result_{yyyy}-{mm}-{dd}_NY.csv"
    )


def print_result_file(path):
    if not os.path.exists(path):
        print("WARNING: no encontre archivo esperado:")
        print(path)
        return

    size_kb = round(os.path.getsize(path) / 1024, 2)
    print(f"OK EXPORTADO: {path}")
    print(f"Tamano: {size_kb} KB")

    with open(path, "r", encoding="utf-8") as f:
        print(f.read().strip())


def clear_previous_result(path):
    if os.path.exists(path):
        os.remove(path)
        print(f"Resultado anterior eliminado: {path}")


def clear_expected_results():
    latest_path = None
    latest_modified = None

    if not os.path.isdir(RESULTS_FOLDER):
        return

    for filename in os.listdir(RESULTS_FOLDER):
        if not filename.startswith("score_trade_result_") or not filename.endswith("_NY.csv"):
            continue

        path = os.path.join(RESULTS_FOLDER, filename)
        try:
            modified = os.path.getmtime(path)
        except OSError:
            continue

        if latest_modified is None or modified > latest_modified:
            latest_path = path
            latest_modified = modified

    if latest_path is None:
        return

    clear_previous_result(latest_path)


def clear_current_result_only(path):
    clear_previous_result(path)


def result_already_exists(path):
    return os.path.exists(path) and os.path.getsize(path) > 0


def parse_result_ticks(value):
    if value in (None, "", "OPEN", "NO_TRADE"):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    normalized = str(value).strip().upper()

    if normalized in ("TP", "EXIT"):
        return 1.0

    if normalized == "SL":
        return -1.0

    if normalized == "BE":
        return 0.0

    try:
        return float(normalized.replace("+", ""))
    except ValueError:
        return None


def read_result_ticks(path):
    if not os.path.exists(path):
        return None

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return None

    if not row:
        return None

    value = row.get("result TP SL BE") or row.get("RESULT")
    return parse_result_ticks(value)


def result_is_terminal(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    if result_label in ("TP", "SL", "EXIT", "BE", "TIME_OVER", "NO_TRADE"):
        return True

    ticks = parse_result_ticks(row.get("result TP SL BE") or row.get("RESULT"))
    if ticks is None:
        return False

    return ticks != 0


def result_has_data_row(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    return bool(row)


def stop_replay():
    replay, from_box, to_box, start_button, stop_button = get_controls()

    if stop_button is not None:
        print("Deteniendo replay...")
        stop_button.click_input()
        return

    print("No encontre boton Stop; probablemente el replay ya termino.")


MAX_WAIT_SECONDS = 90
REPLAY_STILL_RUNNING_WARNING_SECONDS = 45 * 60

def wait_until_result(path, min_modified_time=None):
    print("Esperando resultado terminal en CSV; si el trade esta OPEN no se cambia de dia.")
    dot_count = 0
    start = time.time()

    while True:
        if result_is_terminal(path, min_modified_time):
            print("\rEsperando... listo.   ")
            print("Resultado terminal detectado en CSV; paso al siguiente dia.")
            return True

        elapsed = time.time() - start
        if elapsed > MAX_WAIT_SECONDS:
            print(f"\rTimeout ({MAX_WAIT_SECONDS}s) esperando resultado terminal. Saltando al siguiente dia.")
            return False

        dot_count = dot_count % 3 + 1
        print(f"\rEsperando{'.' * dot_count}{' ' * (3 - dot_count)}", end="", flush=True)
        time.sleep(POLL_SECONDS)


def wait_until_result_row(path, min_modified_time=None):
    print("Esperando primera fila en CSV; al capturarla se detiene Replay y se pasa al siguiente dia.")
    dot_count = 0
    start = time.time()

    while True:
        if result_has_data_row(path, min_modified_time):
            print("\rCSV capturado... listo.   ")
            return True

        elapsed = time.time() - start
        if elapsed > MAX_WAIT_SECONDS:
            print(f"\rTimeout ({MAX_WAIT_SECONDS}s) esperando primera fila de CSV. Saltando al siguiente dia.")
            return False

        dot_count = dot_count % 3 + 1
        print(f"\rEsperando CSV{'.' * dot_count}{' ' * (3 - dot_count)}", end="", flush=True)
        time.sleep(POLL_SECONDS)


def replay_looks_stopped(replay, start_button, stop_button):
    try:
        title = replay.window_text().lower()
        if "playback is stopped" in title or "reproduccion detenida" in title:
            return True
    except Exception:
        pass

    try:
        if stop_button is None or not stop_button.is_enabled():
            return True
    except Exception:
        return True

    try:
        if start_button is not None and start_button.is_enabled() and stop_button is not None and not stop_button.is_enabled():
            return True
    except Exception:
        pass

    return False


def wait_until_replay_stops_by_time():
    print(f"Esperando a que Replay termine por tiempo configurado ({REPLAY_END_TIME}); no se detiene por TP/SL/EXIT/BE/TIME_OVER.")
    dot_count = 0
    start = time.time()
    last_warning_at = start

    while True:
        try:
            replay, from_box, to_box, start_button, stop_button = get_controls(verbose=False, focus=False)
            if replay_looks_stopped(replay, start_button, stop_button):
                print("\rReplay detenido por tiempo configurado.   ")
                return True
        except Exception:
            print("\rReplay ya no esta activo o la ventana no esta disponible.   ")
            return True

        elapsed = time.time() - start
        if elapsed - (last_warning_at - start) > REPLAY_STILL_RUNNING_WARNING_SECONDS:
            print(f"\rReplay sigue corriendo despues de {int(elapsed)}s; sigo esperando el fin por tiempo configurado.   ")
            last_warning_at = time.time()

        dot_count = dot_count % 3 + 1
        print(f"\rReplay corriendo{'.' * dot_count}{' ' * (3 - dot_count)}", end="", flush=True)
        time.sleep(POLL_SECONDS)


def read_trade_result(path, date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    default_row = {
        "fecha": f"{yyyy}-{mm}-{dd}",
        "result TP SL BE": "NO_CSV",
    }

    if not os.path.exists(path):
        print(f"Sin CSV para {default_row['fecha']}; no se sobrescriben columnas de datos.")
        return default_row

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        try:
            row = next(reader)
        except StopIteration:
            default_row["result TP SL BE"] = "EMPTY_CSV"
            return default_row

    if "fecha" not in row:
        legacy_result = row.get("RESULT", "NO_TRADE")
        legacy_ticks = None

        if legacy_result == "TP" and row.get("ENTRY") and row.get("TP"):
            legacy_ticks = abs((float(row["TP"]) - float(row["ENTRY"])) / 0.25)

        if legacy_result == "SL" and row.get("ENTRY") and row.get("SL"):
            legacy_ticks = -abs((float(row["ENTRY"]) - float(row["SL"])) / 0.25)

        row = {
            "fecha": default_row["fecha"],
            "SL_price": row.get("SL"),
            "Entry_price": row.get("ENTRY"),
            "TP_price": row.get("TP"),
            "result TP SL BE": legacy_ticks if legacy_ticks is not None else legacy_result,
        }

    row["fecha"] = row.get("fecha") or default_row["fecha"]
    row["result TP SL BE"] = row.get("result TP SL BE") or "OPEN"
    row["Exit_price"] = row.get("Exit_price") or calculate_exit_price(row)
    return row


def calculate_exit_price(row):
    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()

    if result_label == "TP":
        return row.get("TP_price") or row.get("TP") or ""

    if result_label == "SL":
        return row.get("SL_price") or row.get("SL") or ""

    if result_label != "EXIT":
        return ""

    return ""


def to_number(value):
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def get_or_create_headers(ws):
    default_headers = [
        "fecha",
        "or_low",
        "or_high",
        "range",
        "VWAP_entry",
        "Body",
        "Volume_entry",
        "Delta_entry",
        "BreakOut_SPEED",
        "BreakOut_TICKS_PER_SEC",
        "score total",
        "SL_price",
        "Entry_price",
        "TP_price",
        "SL_ticks",
        "TP_ticks",
        "Exit_price",
        "result TP SL BE",
        "Side",
        "Speed_Profile",
        "MAE_ticks",
        "MFE_ticks",
    ]

    headers = [
        ws.cell(row=3, column=col).value
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=3, column=col).value
    ]

    if not headers:
        headers = default_headers[:]

    for header in default_headers:
        if header not in headers:
            headers.append(header)

    for csv_header in get_csv_headers_for_dates():
        if csv_header not in headers:
            headers.append(csv_header)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    return headers


def get_csv_headers_for_dates():
    headers = []

    for date in DATES_DST:
        path = expected_result_path(date)

        if not os.path.exists(path):
            continue

        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue

                for field in reader.fieldnames:
                    if field and field not in headers:
                        headers.append(field)
        except OSError:
            continue

    return headers


def update_score_workbook():
    if os.path.exists(SCORE_WORKBOOK):
        wb = load_workbook(SCORE_WORKBOOK)
    elif os.path.exists(SCORE_WORKBOOK_FALLBACK):
        wb = load_workbook(SCORE_WORKBOOK_FALLBACK)
    else:
        os.makedirs(os.path.dirname(SCORE_WORKBOOK), exist_ok=True)
        wb = Workbook()

    ws = wb.active

    headers = get_or_create_headers(ws)
    first_row = 4
    last_row = first_row + len(DATES_DST) - 1

    for row in range(first_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    for row_offset, date in enumerate(DATES_DST, start=first_row):
        result = read_trade_result(expected_result_path(date), date)
        missing_result_file = result.get("result TP SL BE") in ("NO_CSV", "EMPTY_CSV")

        for col, header in enumerate(headers, start=1):
            if header not in result:
                if missing_result_file:
                    ws.cell(row=row_offset, column=col).value = None
                continue

            value = to_number(result.get(header))

            if value is None and header != "result TP SL BE":
                continue

            ws.cell(row=row_offset, column=col).value = value

    result_col = headers.index("result TP SL BE") + 1
    result_letter = get_column_letter(result_col)

    ws["A1"] = "win rate"
    ws["B1"] = "profit factor"
    ws["A2"] = f'=IF((COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0"))=0,"",COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")/(COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0")))'
    ws["B2"] = f'=IF(ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row}))=0,IF(SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})>0,"INF",""),SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})/ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row})))'
    ws["A2"].number_format = "0.00%"
    ws["B2"].number_format = "0.00"

    for row_idx in range(first_row, last_row + 1):
        ws.cell(row=row_idx, column=result_col).number_format = "+0.##;-0.##;0"

    widths = {
        "A": 12,
        "B": 12,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 16,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    try:
        wb.save(SCORE_WORKBOOK)
        print(f"Excel actualizado: {SCORE_WORKBOOK}")
    except PermissionError:
        wb.save(SCORE_WORKBOOK_FALLBACK)
        print("WARNING: Excel original bloqueado/abierto; guarde copia actualizada en:")
        print(SCORE_WORKBOOK_FALLBACK)


# =========================================================
# LOOP PRINCIPAL
# =========================================================

print("\nINICIANDO REPLAY DST PARA SCORE TRADE RESULTS\n")

try:
    clear_expected_results()

    for raw_date in DATES_DST:
        date = normalize_replay_date(raw_date)

        print("\n" + "=" * 70)
        print(f"PROCESANDO {date}")
        print("=" * 70)

        result_path = expected_result_path(date)

        if result_already_exists(result_path):
            print(f"CSV existente para {date}; se salta replay y se conserva resultado.")
            continue

        write_target_date(date)
        time.sleep(1)

        from_value = f"{date} 09:30 a. m."
        to_value = f"{date} {REPLAY_END_TIME} a. m."

        replay, from_box, to_box, start_button, stop_button = get_controls()

        paste_text(from_box, from_value)
        paste_text(to_box, to_value)

        print("Fechas configuradas:")
        print(f"FROM: {from_value}")
        print(f"TO:   {to_value}")

        time.sleep(2)

        replay, from_box, to_box, start_button, stop_button = get_controls()

        if start_button is None:
            raise RuntimeError("No se encontro boton Start")

        clear_current_result_only(result_path)
        write_replay_started_marker()

        print("Iniciando replay...")
        started_at = time.time()
        start_button.click_input()

        print("Replay se detendra en cuanto el exporter escriba la primera fila del CSV.")
        wait_until_result_row(result_path, started_at)
        stop_replay()

        time.sleep(5)

        print_result_file(result_path)

        print("Pausa antes del siguiente dia...")
        time.sleep(10)

finally:
    update_score_workbook()

print("\nTERMINO LA PRUEBA DST.\n")
