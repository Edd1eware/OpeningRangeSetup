from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SHEET_ORDER = ("Summary", "Daily_Profile", "LVN_Profile", "LVN_Events", "No_Retest", "Debug")


def _summary_row(section: str, segment: str, metric: str, value: object, n: int | None = None) -> dict[str, object]:
    return {"section": section, "segment": segment, "metric": metric, "value": value, "n": n}


def build_summary(events: pd.DataFrame, daily: pd.DataFrame, lvn_profiles: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    session_count = int(daily["date"].nunique()) if not daily.empty and "date" in daily else 0
    lvn_days = int(lvn_profiles["date"].nunique()) if not lvn_profiles.empty and "date" in lvn_profiles else 0
    event_count = len(events)
    rows.extend([
        _summary_row("Overall", "ALL", "sessions", session_count, session_count),
        _summary_row("Overall", "ALL", "sessions_with_first_minute_lvn", lvn_days, session_count),
        _summary_row("Overall", "ALL", "lvn_occurrence_day_rate", lvn_days / session_count if session_count else math.nan, session_count),
        _summary_row("Overall", "ALL", "lvn_nodes", len(lvn_profiles), len(lvn_profiles)),
        _summary_row("Overall", "ALL", "retest_events", event_count, event_count),
        _summary_row("Overall", "ALL", "retests_per_lvn", event_count / len(lvn_profiles) if len(lvn_profiles) else math.nan, len(lvn_profiles)),
    ])
    if not lvn_profiles.empty:
        lvn_profiles = lvn_profiles.copy()
        lvn_profiles["month"] = pd.to_datetime(lvn_profiles["date"]).dt.to_period("M").astype(str)
        for shape, cohort in lvn_profiles.groupby("context_profile_shape", dropna=False):
            rows.extend([
                _summary_row("Shape occurrence", str(shape), "days_with_lvn", cohort["date"].nunique(), len(cohort)),
                _summary_row("Shape occurrence", str(shape), "lvn_nodes", len(cohort), len(cohort)),
                _summary_row("Shape occurrence", str(shape), "nodes_with_retest_rate", float(cohort["had_retest"].astype(bool).mean()), len(cohort)),
                _summary_row("Shape occurrence", str(shape), "mean_retests_per_node", float(cohort["retest_count"].mean()), len(cohort)),
            ])
        for month, cohort in lvn_profiles.groupby("month"):
            rows.extend([
                _summary_row("Monthly occurrence", month, "days_with_lvn", cohort["date"].nunique(), len(cohort)),
                _summary_row("Monthly occurrence", month, "lvn_nodes", len(cohort), len(cohort)),
                _summary_row("Monthly occurrence", month, "nodes_with_retest", int(cohort["had_retest"].astype(bool).sum()), len(cohort)),
            ])
    if not events.empty:
        event_dates = pd.to_datetime(events["date"])
        events = events.copy()
        events["month"] = event_dates.dt.to_period("M").astype(str)
        monthly_counts = events.groupby("month").size()
        rows.extend([
            _summary_row("Overall", "ALL", "active_months", int(monthly_counts.size), int(monthly_counts.size)),
            _summary_row("Overall", "ALL", "mean_events_per_month", float(monthly_counts.mean()), int(monthly_counts.size)),
            _summary_row("Overall", "ALL", "median_events_per_month", float(monthly_counts.median()), int(monthly_counts.size)),
            _summary_row("Overall", "ALL", "mean_mfe_ticks", float(events["mfe_ticks"].mean()), event_count),
            _summary_row("Overall", "ALL", "mean_mae_ticks", float(events["mae_ticks"].mean()), event_count),
            _summary_row("Overall", "ALL", "median_mfe_mae_ratio", float(events["mfe_mae_ratio"].median()), int(events["mfe_mae_ratio"].notna().sum())),
        ])

        # Shape cohort = medición pura (decisión usuario 2026-07-09): frecuencia del grupo y
        # su máxima extensión (MFE). Sin WR/PF aquí — la estrategia se diseña después.
        total_months = int(events["month"].nunique()) or 1
        for shape, cohort in events.groupby("context_profile_shape", dropna=False):
            mfe = pd.to_numeric(cohort["mfe_ticks"], errors="coerce").dropna()
            time_to_mfe = pd.to_numeric(cohort["time_to_mfe_seconds"], errors="coerce").dropna()
            rows.extend([
                _summary_row("Shape cohort", str(shape), "events", len(cohort), len(cohort)),
                _summary_row("Shape cohort", str(shape), "unique_days", cohort["date"].nunique(), len(cohort)),
                _summary_row("Shape cohort", str(shape), "events_per_month", len(cohort) / total_months, len(cohort)),
                _summary_row("Shape cohort", str(shape), "mfe_ticks_median", float(mfe.median()) if len(mfe) else math.nan, len(mfe)),
                _summary_row("Shape cohort", str(shape), "mfe_ticks_mean", float(mfe.mean()) if len(mfe) else math.nan, len(mfe)),
                _summary_row("Shape cohort", str(shape), "mfe_ticks_p90", float(mfe.quantile(0.9)) if len(mfe) else math.nan, len(mfe)),
                _summary_row("Shape cohort", str(shape), "mfe_ticks_max", float(mfe.max()) if len(mfe) else math.nan, len(mfe)),
                _summary_row("Shape cohort", str(shape), "mae_ticks_median", float(pd.to_numeric(cohort["mae_ticks"], errors="coerce").median()), len(cohort)),
                _summary_row("Shape cohort", str(shape), "time_to_mfe_seconds_median", float(time_to_mfe.median()) if len(time_to_mfe) else math.nan, len(time_to_mfe)),
            ])

        if "lvn_interaction" in events.columns:
            for interaction, cohort in events.groupby("lvn_interaction", dropna=False):
                rows.append(_summary_row("Interaction cohort", str(interaction), "events", len(cohort), len(cohort)))
                rows.append(_summary_row("Interaction cohort", str(interaction), "unique_days", cohort["date"].nunique(), len(cohort)))
                rows.append(_summary_row("Interaction cohort", str(interaction), "mean_mfe_ticks", float(cohort["mfe_ticks"].mean()), len(cohort)))
                rows.append(_summary_row("Interaction cohort", str(interaction), "mean_mae_ticks", float(cohort["mae_ticks"].mean()), len(cohort)))
                for target in (20, 40, 60, 80):
                    result = cohort[f"tp_sl_{target}_{target}_result"].astype(str)
                    wins = int((result == "TP").sum())
                    losses = int((result == "SL").sum())
                    rows.extend([
                        _summary_row("Interaction cohort", str(interaction), f"wr_{target}_{target}_all", wins / len(cohort) if len(cohort) else math.nan, len(cohort)),
                        _summary_row("Interaction cohort", str(interaction), f"wr_{target}_{target}_resolved", wins / (wins + losses) if wins + losses else math.nan, wins + losses),
                        _summary_row("Interaction cohort", str(interaction), f"mean_realized_r_{target}_{target}", float(cohort[f"realized_r_{target}_{target}"].mean()), int(cohort[f"realized_r_{target}_{target}"].notna().sum())),
                    ])
            if "trade_logic" in events.columns and "context_profile_shape" in events.columns:
                for (interaction, shape), cohort in events.groupby(["lvn_interaction", "context_profile_shape"], dropna=False):
                    segment = f"{interaction}|{shape}"
                    rows.append(_summary_row("Interaction x shape", segment, "events", len(cohort), len(cohort)))
                    for target in (20, 40, 60, 80):
                        result = cohort[f"tp_sl_{target}_{target}_result"].astype(str)
                        wins = int((result == "TP").sum())
                        losses = int((result == "SL").sum())
                        rows.append(_summary_row(
                            "Interaction x shape",
                            segment,
                            f"wr_{target}_{target}_resolved",
                            wins / (wins + losses) if wins + losses else math.nan,
                            wins + losses,
                        ))

        for month, cohort in events.groupby("month"):
            rows.append(_summary_row("Monthly", month, "events", len(cohort), len(cohort)))
            rows.append(_summary_row("Monthly", month, "unique_days", cohort["date"].nunique(), len(cohort)))
            for target in (20, 40, 60, 80):
                result = cohort[f"tp_sl_{target}_{target}_result"].astype(str)
                resolved = result.isin(["TP", "SL"])
                rows.append(_summary_row(
                    "Monthly",
                    month,
                    f"wr_{target}_{target}_resolved",
                    float((result[resolved] == "TP").mean()) if resolved.any() else math.nan,
                    int(resolved.sum()),
                ))
    return pd.DataFrame(rows)


def _excel_safe(frame: pd.DataFrame) -> pd.DataFrame:
    safe = frame.copy()
    for column in safe.columns:
        if isinstance(safe[column].dtype, pd.DatetimeTZDtype):
            safe[column] = safe[column].astype(str)
        elif safe[column].dtype == "object":
            safe[column] = safe[column].map(
                lambda value: "|".join(map(str, value)) if isinstance(value, (list, tuple, set)) else value
            )
    return safe.replace([np.inf, -np.inf], np.nan)


def _style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    summary_fill = PatternFill("solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_row >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 30
        for column_index in range(1, worksheet.max_column + 1):
            letter = get_column_letter(column_index)
            header = str(worksheet.cell(1, column_index).value or "")
            sample_lengths = [len(str(worksheet.cell(row, column_index).value or "")) for row in range(1, min(worksheet.max_row, 200) + 1)]
            width = min(34, max(10, max(sample_lengths, default=len(header)) + 2))
            if header in {"event_id", "lvn_id", "lvn_ids"}:
                width = min(55, max(40, max(sample_lengths, default=40) + 2))
            elif header == "source_file":
                width = 60
            elif "reason" in header:
                width = min(55, max(30, max(sample_lengths, default=30) + 2))
            elif "time" in header:
                width = min(30, max(width, 22))
            worksheet.column_dimensions[letter].width = width
        if worksheet.title == "Summary" and worksheet.max_row > 1:
            for row in range(2, worksheet.max_row + 1):
                if worksheet.cell(row, 1).value == "Overall":
                    for cell in worksheet[row]:
                        cell.fill = summary_fill
    workbook.save(path)


def export_results(output_path: str | Path, sheets: dict[str, pd.DataFrame]) -> tuple[Path, Path]:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    # Regla 2026-07-08: NUNCA sobrescribir/borrar un Excel de una corrida anterior.
    # Si el destino existe, la corrida nueva escribe a un archivo con timestamp;
    # borrar reportes viejos es SIEMPRE decisión manual del usuario.
    if path.exists():
        from datetime import datetime as _dt

        stamp = _dt.now().strftime("%Y%m%d_%H%M%S")
        path = path.with_name(f"{path.stem}_run{stamp}{path.suffix}")
    csv_dir = path.parent / f"{path.stem}_csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    ordered = {name: _excel_safe(sheets.get(name, pd.DataFrame())) for name in SHEET_ORDER}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in ordered.items():
            if frame.empty:
                frame = pd.DataFrame({"status": ["NO_ROWS"]})
            frame.to_excel(writer, sheet_name=name, index=False)
            frame.to_csv(csv_dir / f"{name}.csv", index=False, encoding="utf-8-sig")
    _style_workbook(path)
    return path, csv_dir
