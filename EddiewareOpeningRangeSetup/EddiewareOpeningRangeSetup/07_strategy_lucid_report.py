"""Reporte Lucid del Execution Manager: PASASTE/QUEMASTE + meses/dias + grafica + Excel + Telegram.

Lee los trades REALES que la estrategia registro (strategy_tester_trades.csv).
Si no existe aun, usa el backtest (trailing sim) como PREVIEW.

Lucid 100k: profit target $6,000 | MLL trailing $3,000 (EOD) | NQ $5/tick.

Uso:  python 07_strategy_lucid_report.py            (Contracts del log o 5)
      python 07_strategy_lucid_report.py --contracts 6
"""

import argparse
import csv
import re
from datetime import date
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

import replay_sync_runner_common_after_sync as rs
import telegram_run_summary_after_sync as telegram

RES = rs.RESULTS_FOLDER
TESTER = RES / "visual_tests" / "strategy_tester_results"
TRADES_CSV = TESTER / "strategy_tester_trades.csv"
LADDER_ROOT = RES / "visual_tests" / "sync_ladder_runs" / "sync_v11_ladder_001_resume"
EDGE_NAUTILUS = RES / "visual_tests" / "orb_nq_databento_edge_nautilus"
EDGE_NAUTILUS.mkdir(parents=True, exist_ok=True)
PNG = str(EDGE_NAUTILUS / "lucid_resultado.png")
XLSX = str(EDGE_NAUTILUS / "EW_Strategy_Lucid_Resultado.xlsx")

TICK_USD = 5.0          # NQ
TARGET = 6000.0
MLL = 3000.0
OPER = rs.OPERATIVA_COMPARISON_FIELDS


def num(v):
    try:
        return float(str(v).strip().replace("+", ""))
    except Exception:
        return None


def load_real_trades():
    """(date_iso, pnl_usd) de los fills reales registrados por la estrategia."""
    if not TRADES_CSV.exists():
        return []
    out = []
    for r in csv.DictReader(open(TRADES_CSV, encoding="utf-8-sig")):
        d = (r.get("fecha") or "").strip()
        p = num(r.get("pnl_usd"))
        if d and p is not None:
            out.append((d, p))
    return sorted(out)


def backtest_preview(contracts):
    """Fallback: trailing sim sobre las 39 fechas A+ Speed x contratos x $5."""
    def row_of(p):
        try:
            return next(csv.DictReader(open(p, encoding="utf-8-sig")), None)
        except Exception:
            return None

    def rows_of(p):
        try:
            return list(csv.DictReader(open(p, encoding="utf-8-sig")))
        except Exception:
            return []

    def b(v):
        return str(v).strip().upper() == "TRUE"

    NO_TRADE = {"", "TIME_OVER", "NO_TRADE", "HOLYDAY NO DATA", "OPEN"}
    trades = {}
    for x10 in LADDER_ROOT.glob("*/X10_R1/score_trade_result_*_NY.csv"):
        d = re.search(r"(\d{4}-\d{2}-\d{2})", x10.name).group(1)
        r10 = row_of(x10)
        if not r10:
            continue
        lab = (r10.get("Result_Label") or "").strip().upper()
        if lab in NO_TRADE or not (r10.get("Entry_price") or "").strip():
            continue
        if not b(r10.get("APlus_Speed")):
            continue
        if d not in trades:
            trades[d] = (r10, x10.parent / f"dynamic_timeline_{d}_NY.csv")

    def exit_ticks(tl, real, SL=50, act=20, trail=10):
        active = False
        for s in rows_of(tl):
            mfe = num(s.get("MFE_ticks_To_Now")) or 0
            mae = num(s.get("MAE_ticks_To_Now")) or 0
            dd = num(s.get("Drawdown_From_MFE_Ticks")) or 0
            if mae >= SL and not active:
                return -SL
            if mfe >= act:
                active = True
            if active and dd >= trail:
                return max(0, mfe - trail)
        return max(-SL, real)

    out = []
    for d, (r, tl) in sorted(trades.items()):
        ticks = exit_ticks(tl, num(r.get("result TP SL BE")) or 0)
        out.append((d, ticks * TICK_USD * contracts))
    return out


def months_days(d0, d1):
    months = (d1.year - d0.year) * 12 + (d1.month - d0.month)
    days = d1.day - d0.day
    if days < 0:
        months -= 1
        days += 30
    return max(0, months), max(0, days)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--contracts", type=int, default=None)
    args = parser.parse_args()

    trades = load_real_trades()
    source = "REAL (fills de la estrategia)"
    if not trades:
        contracts = args.contracts or 5
        trades = backtest_preview(contracts)
        source = f"PREVIEW backtest ({contracts} contratos)"
    if not trades:
        print("No hay trades.")
        return 1

    # Curva de equity + resultado Lucid (trailing MLL EOD).
    eq = 0.0
    peak = 0.0
    cum = []
    floor = []
    result = "EN CURSO"
    pass_idx = blow_idx = None
    for i, (d, pnl) in enumerate(trades):
        eq += pnl
        peak = max(peak, eq)
        f = min(peak - MLL, 0.0)
        cum.append(eq)
        floor.append(f)
        if blow_idx is None and eq <= f:
            blow_idx = i
            result = "QUEMASTE"
            break
        if pass_idx is None and eq >= TARGET:
            pass_idx = i
            result = "PASASTE"
            break

    first = date.fromisoformat(trades[0][0])
    n = len(cum)
    x = range(1, n + 1)
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.plot(x, cum, marker="o", ms=3, lw=1.8, color="#2ca02c", label="Equity (USD)")
    ax.axhline(TARGET, color="#1f77b4", lw=1.5, ls="--", label="Profit Target +$6,000")
    ax.plot(x, floor, color="#d62728", lw=1.3, label="Trailing MLL piso $3,000")
    ax.axhline(0, color="grey", lw=0.8)

    detail = ""
    if result == "PASASTE":
        pd = date.fromisoformat(trades[pass_idx][0])
        mm, dd = months_days(first, pd)
        ax.axvline(pass_idx + 1, color="#1f77b4", ls=":", lw=0.8)
        ax.annotate(f"PASO ({mm}m {dd}d)", (pass_idx + 1, TARGET),
                    xytext=(6, 8), textcoords="offset points", color="#1f77b4")
        detail = f"en {mm} meses {dd} dias (trade {pass_idx+1}, {pd})"
    elif result == "QUEMASTE":
        bd = date.fromisoformat(trades[blow_idx][0])
        ax.axvline(blow_idx + 1, color="#d62728", ls=":", lw=0.8)
        ax.annotate("QUEMO", (blow_idx + 1, floor[blow_idx]),
                    xytext=(6, -12), textcoords="offset points", color="#d62728")
        detail = f"en trade {blow_idx+1} ({bd})"
    else:
        detail = f"equity ${cum[-1]:,.0f} de ${TARGET:,.0f} ({n} trades)"

    title_word = {"PASASTE": "PASASTE LA CUENTA", "QUEMASTE": "QUEMASTE LA CUENTA",
                  "EN CURSO": "EN CURSO"}[result]
    ax.set_title(f"EW Execution Manager - Resultado Lucid 100k ({source})\n"
                 f"{title_word} {detail}")
    ax.set_xlabel("# trade"); ax.set_ylabel("Equity acumulada (USD)")
    ax.legend(loc="upper left"); ax.grid(alpha=0.3)
    TESTER.mkdir(parents=True, exist_ok=True)
    plt.tight_layout(); plt.savefig(PNG, dpi=110)
    print("PNG:", PNG)

    # Excel con tabla + grafica
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
        wb = Workbook(); ws = wb.active; ws.title = "Trades"
        ws.append(["fecha", "PnL_USD", "equity_acum"])
        for c in ws[1]:
            c.font = Font(bold=True)
        run = 0.0
        for d, pnl in trades:
            run += pnl
            ws.append([d, round(pnl, 2), round(run, 2)])
        ws2 = wb.create_sheet("Resultado_Lucid")
        ws2["A1"] = f"{title_word} {detail}"; ws2["A1"].font = Font(bold=True, size=12)
        ws2.add_image(XLImage(PNG), "A3")
        wb.save(XLSX); print("XLSX:", XLSX)
    except Exception as exc:
        print("Excel fallo:", exc)

    # Telegram: resultado + grafica. Es un resultado PARCIAL/estimacion; el replay
    # NO se detiene por esto (este script es read-only, separado del runner).
    head = {"PASASTE": "PASASTE LA CUENTA! 🎉", "QUEMASTE": "QUEMASTE LA CUENTA",
            "EN CURSO": "Resultado parcial"}[result]
    cap = (f"EW Execution Manager | LUCID 100k - {head}\n"
           f"{title_word} {detail}.\n"
           f"(Resultado PARCIAL/estimacion - el replay sigue corriendo)\n"
           f"Fuente: {source}. NQ $5/tick. Target $6,000 / MLL $3,000.")
    print("telegram foto:", telegram.send_photo(RES, PNG, cap))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
