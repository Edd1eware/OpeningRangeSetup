from __future__ import annotations

import math
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from lvn_retest_engine.config import ResearchConfig
from lvn_retest_engine.main import run_research


def add_profile(rows: list[dict[str, object]], day: str, clock: str, prices: list[float], volumes: list[float]) -> None:
    for offset, (price, volume) in enumerate(zip(prices, volumes, strict=True)):
        rows.append({
            "timestamp": f"{day} {clock}.{offset:03d}",
            "price": price,
            "volume": volume,
            "side": "BUY" if offset % 2 == 0 else "SELL",
            "symbol": "NQ",
        })


def synthetic_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    context_prices = [99.50, 99.75, 100.00, 100.25, 100.50, 100.75, 101.00, 101.25, 101.50]
    context_volumes = [20, 45, 80, 120, 150, 120, 80, 45, 20]
    minute_prices = [100.00, 100.25, 100.50, 100.75, 101.00]
    minute_volumes = [90, 80, 10, 90, 110]
    for day in ("2026-06-01", "2026-06-02"):
        add_profile(rows, day, "08:30:00", context_prices, context_volumes)
        add_profile(rows, day, "09:30:00", minute_prices, minute_volumes)

    # Day 1: approach from above, touch 100.50, then travel exactly +20 ticks.
    for second, price in ((1, 101.50), (2, 101.00), (3, 100.75), (4, 100.50), (5, 101.00), (6, 105.50)):
        rows.append({"timestamp": f"2026-06-01 09:31:{second:02d}", "price": price, "volume": 5, "side": "BUY", "symbol": "NQ"})
    # This post-09:31 valley-looking pattern must never alter the frozen LVN profile.
    add_profile(rows, "2026-06-01", "09:32:00", [102.25, 102.50, 102.75], [100, 1, 100])

    # Day 2: LVN exists, but price remains above it until 09:40.
    for minute, price in ((31, 102.00), (33, 102.50), (35, 103.00), (39, 102.25)):
        rows.append({"timestamp": f"2026-06-02 09:{minute:02d}:00", "price": price, "volume": 5, "side": "SELL", "symbol": "NQ"})
    return rows


class LvnResearchEngineTests(unittest.TestCase):
    def test_end_to_end_causal_event_and_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "synthetic_ticks.csv"
            output = root / "results.xlsx"
            pd.DataFrame(synthetic_rows()).to_csv(source, index=False)
            config = ResearchConfig(
                lvn_neighbor_levels=1,
                hvn_neighbor_levels=1,
                lvn_max_percent_of_neighbors=0.50,
                max_lvn_volume_percent_of_poc=0.50,
            )
            result = run_research([str(source)], output, config)
            self.assertEqual(result["sessions"], 2)
            self.assertEqual(result["lvns"], 2)
            self.assertEqual(result["events"], 1)
            self.assertEqual(result["no_retest_days"], 1)

            events = pd.read_csv(root / "results_csv" / "LVN_Events.csv")
            lvns = pd.read_csv(root / "results_csv" / "LVN_Profile.csv")
            no_retest = pd.read_csv(root / "results_csv" / "No_Retest.csv")
            self.assertEqual(len(events), 1)
            self.assertAlmostEqual(float(events.iloc[0]["lvn_price"]), 100.50)
            self.assertEqual(events.iloc[0]["approach"], "FROM_ABOVE")
            self.assertEqual(events.iloc[0]["expected_reaction_side"], "LONG")
            self.assertTrue(bool(events.iloc[0]["hit_plus_20"]))
            self.assertEqual(events.iloc[0]["tp_sl_20_20_result"], "TP")
            self.assertEqual(set(lvns["price"].round(2)), {100.50})
            self.assertNotIn(102.50, set(lvns["price"].round(2)))
            self.assertEqual(no_retest.iloc[0]["date"], "2026-06-02")

            for prefix in ("context_", "minute_"):
                probability_columns = [
                    f"{prefix}prob_D", f"{prefix}prob_P", f"{prefix}prob_b",
                    f"{prefix}prob_double", f"{prefix}prob_trend_up",
                    f"{prefix}prob_trend_down", f"{prefix}prob_unknown",
                ]
                total = float(lvns.iloc[0][probability_columns].sum())
                self.assertTrue(math.isclose(total, 1.0, rel_tol=1e-9, abs_tol=1e-9))

            workbook = load_workbook(output, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["Summary", "Daily_Profile", "LVN_Profile", "LVN_Events", "No_Retest", "Debug"],
            )
            workbook.close()

    def test_footprint_bar_uses_next_bar_and_marks_same_bar_order_ambiguous(self) -> None:
        rows: list[dict[str, object]] = []
        day = "2026-06-03"
        context_prices = [99.50, 99.75, 100.00, 100.25, 100.50, 100.75, 101.00]
        context_volumes = [20, 50, 100, 140, 100, 50, 20]
        minute_prices = [100.00, 100.25, 100.50, 100.75, 101.00]
        minute_volumes = [90, 80, 10, 90, 110]
        for timestamp, prices, volumes, ohlc in (
            (f"{day} 08:30:00", context_prices, context_volumes, (100.0, 101.0, 99.5, 100.5)),
            (f"{day} 09:30:00", minute_prices, minute_volumes, (100.0, 101.0, 100.0, 101.0)),
            (f"{day} 09:31:00", [100.50], [20], (101.0, 101.0, 100.5, 100.75)),
            (f"{day} 09:32:00", [100.50], [20], (100.5, 105.5, 95.5, 100.5)),
        ):
            for price, volume in zip(prices, volumes, strict=True):
                rows.append({
                    "timestamp": timestamp,
                    "price": price,
                    "bid_volume": volume / 2,
                    "ask_volume": volume / 2,
                    "open": ohlc[0], "high": ohlc[1], "low": ohlc[2], "close": ohlc[3],
                    "symbol": "NQ",
                })
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source, output = root / "footprint.csv", root / "footprint.xlsx"
            pd.DataFrame(rows).to_csv(source, index=False)
            result = run_research(
                [str(source)],
                output,
                ResearchConfig(lvn_neighbor_levels=1, hvn_neighbor_levels=1),
            )
            self.assertEqual(result["events"], 1)
            event = pd.read_csv(root / "footprint_csv" / "LVN_Events.csv").iloc[0]
            self.assertEqual(event["timing_precision"], "BAR_UNORDERED")
            self.assertEqual(event["outcome_start_policy"], "NEXT_BAR_CAUSAL_CONSERVATIVE")
            self.assertEqual(event["tp_sl_20_20_result"], "AMBIGUOUS")


if __name__ == "__main__":
    unittest.main()
