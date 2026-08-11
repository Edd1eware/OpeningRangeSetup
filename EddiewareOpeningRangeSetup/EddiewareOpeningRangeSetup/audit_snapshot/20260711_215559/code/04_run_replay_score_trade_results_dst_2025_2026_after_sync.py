import argparse
from pywinauto import Desktop
import csv
from datetime import date, datetime, timedelta
import os
import shutil
import time
from pathlib import Path
import pyperclip
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import replay_sync_runner_common_after_sync as replay_sync
from telegram_run_summary_after_sync import clear_telegram_before_run, send_run_summary, send_text

# =========================================================
# CONFIG
# =========================================================

# Temporadas DST completas de Nueva York de 2022 a 2026.
# 2022 arranca el 04/04/2022 (inicio de datos disponibles), no en el cambio DST.
# Se excluyen fines de semana y cierres completos del mercado de EE. UU.
# Los cierres tempranos se incluyen porque el replay solo usa 09:30-09:50 NY.
# Formato requerido por el panel Replay de ATAS: dd/mm/yyyy.

DST_SEASONS = (
    (date(2022, 4, 4), date(2022, 11, 4)),
    (date(2023, 3, 13), date(2023, 11, 3)),
    (date(2024, 3, 11), date(2024, 11, 1)),
    (date(2025, 3, 10), date(2025, 10, 31)),
    (date(2026, 3, 9), date(2026, 10, 30)),
)

MARKET_CLOSED_DATES = {
    # 2022
    date(2022, 1, 17),  # Martin Luther King Jr. Day
    date(2022, 2, 21),  # Washington's Birthday
    date(2022, 4, 15),  # Good Friday
    date(2022, 5, 30),  # Memorial Day
    date(2022, 6, 20),  # Juneteenth observed
    date(2022, 7, 4),   # Independence Day
    date(2022, 9, 5),   # Labor Day
    date(2022, 11, 24), # Thanksgiving Day
    date(2022, 12, 26), # Christmas Day observed

    # 2023
    date(2023, 1, 2),   # New Year's Day observed
    date(2023, 1, 16),  # Martin Luther King Jr. Day
    date(2023, 2, 20),  # Washington's Birthday
    date(2023, 4, 7),   # Good Friday
    date(2023, 5, 29),  # Memorial Day
    date(2023, 6, 19),  # Juneteenth
    date(2023, 7, 4),   # Independence Day
    date(2023, 9, 4),   # Labor Day
    date(2023, 11, 23), # Thanksgiving Day
    date(2023, 12, 25), # Christmas Day

    # 2024
    date(2024, 1, 1),   # New Year's Day
    date(2024, 1, 15),  # Martin Luther King Jr. Day
    date(2024, 2, 19),  # Washington's Birthday
    date(2024, 3, 29),  # Good Friday
    date(2024, 5, 27),  # Memorial Day
    date(2024, 6, 19),  # Juneteenth
    date(2024, 7, 4),   # Independence Day
    date(2024, 9, 2),   # Labor Day
    date(2024, 11, 28), # Thanksgiving Day
    date(2024, 12, 25), # Christmas Day

    # 2025
    date(2025, 1, 1),   # New Year's Day
    date(2025, 1, 9),   # National Day of Mourning
    date(2025, 1, 20),  # Martin Luther King Jr. Day
    date(2025, 2, 17),  # Washington's Birthday
    date(2025, 4, 18),  # Good Friday
    date(2025, 5, 26),  # Memorial Day
    date(2025, 6, 19),  # Juneteenth
    date(2025, 7, 4),   # Independence Day
    date(2025, 9, 1),   # Labor Day
    date(2025, 11, 27), # Thanksgiving Day
    date(2025, 12, 25), # Christmas Day

    # 2026
    date(2026, 1, 1),   # New Year's Day
    date(2026, 1, 19),  # Martin Luther King Jr. Day
    date(2026, 2, 16),  # Washington's Birthday
    date(2026, 4, 3),   # Good Friday
    date(2026, 5, 25),  # Memorial Day
    date(2026, 6, 19),  # Juneteenth
    date(2026, 7, 3),   # Independence Day observed
    date(2026, 9, 7),   # Labor Day
    date(2026, 11, 26), # Thanksgiving Day
    date(2026, 12, 25), # Christmas Day
}


def build_trading_dates(start_date, end_date):
    trading_dates = []
    current = start_date

    while current <= end_date:
        if current.weekday() < 5 and current not in MARKET_CLOSED_DATES:
            trading_dates.append(current.strftime("%d/%m/%Y"))
        current += timedelta(days=1)

    return trading_dates


def replay_date_is_allowed(date_ddmmyyyy):
    today_ny = datetime.now(ZoneInfo("America/New_York")).date()
    last_replay_date = today_ny - timedelta(days=1)
    replay_date = datetime.strptime(date_ddmmyyyy, "%d/%m/%Y").date()
    return replay_date <= last_replay_date


ALL_DATES_DST = [
    replay_date
    for season_start, season_end in DST_SEASONS
    for replay_date in build_trading_dates(season_start, season_end)
]

TODAY_NY = datetime.now(ZoneInfo("America/New_York")).date()
LAST_REPLAY_DATE = TODAY_NY - timedelta(days=1)
DATES_DST = [
    replay_date
    for replay_date in ALL_DATES_DST
    if datetime.strptime(replay_date, "%d/%m/%Y").date() <= LAST_REPLAY_DATE
]

# Fase 1: primeras N fechas en X1 y luego X10 para validar sincronia X1==X10.
# Fase 2: si la sincronia pasa, corrida completa desde 04/04/2022 SOLO en X10.
SYNC_CHECK_DATE_COUNT = 10

# Replay recomendado para esta prueba: X10.
# Ventana por dia: 09:30 a 09:50 NY. El exporter escribe TIME_OVER si no hay trade antes/de 09:40;
# si ya hay trade abierto, dejamos correr hasta 09:50 para que resuelva TP/SL/EXIT/BE.
REPLAY_END_TIME = "09:50"
POLL_SECONDS = 0.02
NO_TRADE_CUTOFF_SECONDS = 10 * 60 + 15
HOLIDAY_RETRY_COUNT = 3
HOLIDAY_NORMAL_WAIT_SECONDS = 2 * 60
HOLIDAY_FINAL_WAIT_SECONDS = 3 * 60
HOLIDAY_NO_DATA_LABEL = "HOLYDAY NO DATA"
REQUIRED_TIMELINE_VERSION = replay_sync.EXPECTED_TIMELINE_VERSION

EXPORT_FOLDER = r"C:\Users\k_99_\Desktop\codding\data_footprint_generator"
RESULTS_FOLDER = os.path.join(EXPORT_FOLDER, "trade_results_score")
TIMELINE_FOLDER = os.path.join(RESULTS_FOLDER, "dynamic_management_timeline")
TARGET_FILE = os.path.join(EXPORT_FOLDER, "target_trade_result_date.txt")
REPLAY_STARTED_FILE = os.path.join(EXPORT_FOLDER, "replay_trade_result_started_at.txt")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCORE_WORKBOOK_TEMPLATE = os.path.join(BASE_DIR, "Score_indicator_results_updated.xlsx")
SCORE_WORKBOOK_TEMPLATE_FALLBACK = os.path.join(BASE_DIR, "Score_indicator_results_updated_fallback.xlsx")
SCORE_WORKBOOK = os.path.join(BASE_DIR, "Score_indicator_results_updated_2022_2026.xlsx")
SCORE_WORKBOOK_FALLBACK = os.path.join(BASE_DIR, "Score_indicator_results_updated_2022_2026_fallback.xlsx")
RUN_STARTED_AT = time.time()
RESUME_EXISTING_RESULTS = True
STALE_RESULT_BACKUP_DIR = os.path.join(RESULTS_FOLDER, "_replay_result_backups")
STALE_TIMELINE_BACKUP_DIR = os.path.join(TIMELINE_FOLDER, "_replay_timeline_backups")
EXCLUDED_EXCEL_HEADERS = {"Contracts"}


# =========================================================
# FUNCIONES
# =========================================================

def write_target_date(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    target = f"{yyyy}-{mm}-{dd}"

    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    os.makedirs(RESULTS_FOLDER, exist_ok=True)

    with open(TARGET_FILE, "w", encoding="utf-8") as f:
        f.write(target)

    print(f"Fecha objetivo escrita para ATAS: {target}")


def write_replay_started_marker():
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    with open(REPLAY_STARTED_FILE, "w", encoding="utf-8") as f:
        f.write(str(time.time()))
    print("Marcador de inicio de replay escrito.")


def get_replay():
    desktop = Desktop(backend="uia")
    candidates = desktop.windows(title_re=".*Replay.*", visible_only=True)

    if not candidates:
        raise RuntimeError("No encontre ninguna ventana Replay visible. Abre Replay en ATAS antes de correr el script.")

    print(f"Ventanas Replay encontradas: {len(candidates)}")

    for w in candidates:
        try:
            if w.is_visible() and w.is_enabled():
                print("Usando Replay:", w.window_text())
                w.set_focus()
                time.sleep(1)
                return w
        except Exception:
            pass

    replay = candidates[0]
    replay.set_focus()
    time.sleep(1)
    return replay


def paste_text(control, value):
    control.click_input()
    time.sleep(0.3)

    control.type_keys("^a")
    time.sleep(0.3)

    pyperclip.copy(value)
    time.sleep(0.3)

    control.type_keys("^v")
    time.sleep(0.8)


def get_controls():
    replay = get_replay()

    edits = replay.descendants(control_type="Edit")
    buttons = replay.descendants(control_type="Button")

    from_box = edits[0]
    to_box = edits[2]

    start_button = None
    stop_button = None

    for b in buttons:
        txt = b.window_text()

        if txt == "Start":
            start_button = b

        if txt == "Stop":
            stop_button = b

    return replay, from_box, to_box, start_button, stop_button


def expected_result_path(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")

    return os.path.join(
        RESULTS_FOLDER,
        f"score_trade_result_{yyyy}-{mm}-{dd}_NY.csv"
    )


def expected_timeline_path(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")

    return os.path.join(
        TIMELINE_FOLDER,
        f"dynamic_timeline_{yyyy}-{mm}-{dd}_NY.csv"
    )


def print_result_file(path):
    if not os.path.exists(path):
        print("WARNING: no encontre archivo esperado:")
        print(path)
        return

    size_kb = round(os.path.getsize(path) / 1024, 2)
    print(f"OK EXPORTADO: {path}")
    print(f"Tamano: {size_kb} KB")

    with open(path, "r", encoding="utf-8") as f:
        print(f.read().strip())


def expected_feature_scan_path(date_ddmmyyyy):
    # Sidecar written by the "Feature Scanner" indicator (research sweep).
    # Same folder + one-date gate as the score CSV, joinable by `fecha`.
    dd, mm, yyyy = date_ddmmyyyy.split("/")

    return os.path.join(
        RESULTS_FOLDER,
        f"features_scan_{yyyy}-{mm}-{dd}_NY.csv"
    )


def print_feature_scans():
    # Print the NEW features produced by the Feature Scanner for every replayed
    # date. Header once, then one data line per date (each sidecar holds 1 row).
    print("\n=== FEATURES NUEVAS (Feature Scanner sidecar) ===")
    header_printed = False
    found = 0

    for date in DATES_DST:
        path = expected_feature_scan_path(date)

        if not os.path.exists(path):
            continue

        with open(path, "r", encoding="utf-8-sig") as f:
            lines = f.read().strip().splitlines()

        if not lines:
            continue

        if not header_printed:
            print(f"Columnas ({len(lines[0].split(','))}): {lines[0]}")
            header_printed = True

        for data_line in lines[1:]:
            print(data_line)

        found += 1

    if found == 0:
        print(
            "No se encontraron archivos features_scan_*.csv. "
            "Confirma que el indicador 'Feature Scanner' este agregado al chart "
            "durante el replay (gate por target_trade_result_date.txt)."
        )
    else:
        print(f"Total fechas con features: {found}")


def print_excel_sizes_by_year():
    # Total size of the score workbooks in BASE_DIR, grouped by the year token in
    # the filename (e.g. "2025_2026", "2024", "1mes"). Console only.
    import glob
    import re

    print("\n=== PESO DE EXCELS (workbooks de score) ===")
    files = sorted(glob.glob(os.path.join(BASE_DIR, "Score_indicator_results_updated*.xlsx")))

    if not files:
        print(f"No hay workbooks Score_indicator_results_updated*.xlsx en {BASE_DIR}")
        return

    by_year = {}
    for path in files:
        name = os.path.basename(path)
        match = re.search(r"(\d{4}(?:_\d{4})?|1mes)", name)
        year = match.group(1) if match else "sin_ano"
        size = os.path.getsize(path)
        by_year[year] = by_year.get(year, 0) + size
        print(f"  {name:<58s} {size / 1024:8.1f} KB")

    print("  " + "-" * 78)
    total = 0
    for year in sorted(by_year):
        size = by_year[year]
        total += size
        print(f"  Ano {year:<12s} total: {size / 1024:9.1f} KB  ({size / 1048576:.2f} MB)")
    print(f"  {'TOTAL':<17s}      {total / 1024:9.1f} KB  ({total / 1048576:.2f} MB)")


def backup_previous_result(path, reason):
    if not os.path.exists(path):
        return

    os.makedirs(STALE_RESULT_BACKUP_DIR, exist_ok=True)
    base_name = os.path.basename(path)
    name, ext = os.path.splitext(base_name)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(STALE_RESULT_BACKUP_DIR, f"{name}_{reason}_{timestamp}{ext}")
    os.replace(path, backup_path)
    print(f"Resultado previo no terminal movido a backup: {backup_path}")


def backup_previous_timeline(path, reason):
    if not os.path.exists(path):
        return

    os.makedirs(STALE_TIMELINE_BACKUP_DIR, exist_ok=True)
    base_name = os.path.basename(path)
    name, ext = os.path.splitext(base_name)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(
        STALE_TIMELINE_BACKUP_DIR,
        f"{name}_{reason}_{timestamp}{ext}"
    )
    os.replace(path, backup_path)
    print(f"Timeline previo movido a backup: {backup_path}")


def clear_previous_result(path, force=False):
    if not os.path.exists(path):
        return False

    if RESUME_EXISTING_RESULTS and result_is_terminal(path) and not force:
        print(f"Resultado terminal existente conservado: {path}")
        return False

    backup_previous_result(path, "stale")
    return True


def clear_previous_timeline(path):
    if not os.path.exists(path):
        return False

    backup_previous_timeline(path, "stale")
    return True


def clear_expected_results():
    for date in DATES_DST:
        result_path = expected_result_path(date)
        timeline_path = expected_timeline_path(date)

        if date_has_complete_artifacts(result_path, timeline_path):
            continue

        # Preserve terminal historical results until their date is actually
        # replayed. This keeps a stopped run resumable without moving every
        # future result to backup at startup.
        if result_is_terminal(result_path):
            continue

        clear_previous_result(result_path)
        clear_previous_timeline(timeline_path)


def parse_result_ticks(value):
    if value in (None, "", "OPEN", "NO_TRADE"):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    normalized = str(value).strip().upper()

    if normalized in ("TP", "EXIT"):
        return 1.0

    if normalized == "SL":
        return -1.0

    if normalized == "BE":
        return 0.0

    try:
        return float(normalized.replace("+", ""))
    except ValueError:
        return None


def read_result_ticks(path):
    if not os.path.exists(path):
        return None

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return None

    if not row:
        return None

    value = row.get("result TP SL BE") or row.get("RESULT")
    return parse_result_ticks(value)


def result_is_terminal(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    if result_label in ("TP", "SL", "EXIT", "BE", "TIME_OVER", "NO_TRADE", HOLIDAY_NO_DATA_LABEL):
        return True

    ticks = parse_result_ticks(row.get("result TP SL BE") or row.get("RESULT"))
    if ticks is None:
        return False

    return ticks != 0


def result_requires_timeline(path):
    if not os.path.exists(path):
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    return result_label in ("TP", "SL", "EXIT", "BE")


def timeline_is_terminal(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            first_row = None
            last_row = None
            for row in reader:
                if first_row is None:
                    first_row = row
                last_row = row
    except (OSError, PermissionError):
        return False

    if not first_row or not last_row:
        return False

    if str(first_row.get("Timeline_VERSION") or "").strip() != REQUIRED_TIMELINE_VERSION:
        return False

    result_label = str(last_row.get("Result") or "").strip().upper()
    event_name = str(last_row.get("Event") or "").strip().upper()
    return (
        result_label in ("TP", "SL", "EXIT", "BE") and
        event_name.startswith("EXIT_")
    )


def date_has_complete_artifacts(result_path, timeline_path):
    if not result_is_terminal(result_path):
        return False

    if not result_requires_timeline(result_path):
        return True

    return timeline_is_terminal(timeline_path)


def result_is_open_trade(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    return result_label == "OPEN"


def write_holiday_no_data_result(path, date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    os.makedirs(os.path.dirname(path), exist_ok=True)

    headers = [
        "Exporter_VERSION",
        "fecha",
        "Result_Label",
        "result TP SL BE",
        "Signal_Source",
    ]
    row = [
        "python-replay-holiday-no-data",
        f"{yyyy}-{mm}-{dd}",
        HOLIDAY_NO_DATA_LABEL,
        HOLIDAY_NO_DATA_LABEL,
        "NO_DATA",
    ]

    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerow(row)

    print(f"CSV marcado como {HOLIDAY_NO_DATA_LABEL}: {path}")


def stop_replay(stop_button=None):
    print("Deteniendo replay...")

    candidates = [stop_button] if stop_button is not None else []

    for _ in range(5):
        try:
            replay, from_box, to_box, start_button, refreshed_stop_button = get_controls()
            replay.set_focus()
            if refreshed_stop_button is not None:
                candidates.insert(0, refreshed_stop_button)
        except Exception:
            pass

        for candidate in candidates:
            if candidate is None:
                continue

            try:
                if candidate.is_visible() and candidate.is_enabled():
                    candidate.click_input()
                    return
            except Exception:
                try:
                    candidate.click()
                    return
                except Exception:
                    pass

        time.sleep(0.05)

    print("No encontre boton Stop; probablemente el replay ya termino.")


MAX_WAIT_SECONDS = 1200

def format_countdown(seconds):
    seconds = max(0, int(seconds))
    minutes, seconds = divmod(seconds, 60)
    return f"{minutes:02d}:{seconds:02d}"


def wait_until_result(path, min_modified_time=None, no_trade_cutoff_seconds=None, stop_button=None):
    print("Esperando resultado terminal en CSV; si el trade esta OPEN no se cambia de dia.")
    start = time.time()
    last_print_second = -1
    last_status_line = ""

    while True:
        if result_is_terminal(path, min_modified_time):
            print("\r" + " " * max(len(last_status_line), 80), end="\r", flush=True)
            print("Esperando... listo.")
            print("Resultado terminal detectado en CSV; paso al siguiente dia.")
            stop_replay(stop_button)
            return True

        elapsed = time.time() - start
        has_open_trade = result_is_open_trade(path, min_modified_time)
        countdown_limit = no_trade_cutoff_seconds if no_trade_cutoff_seconds is not None else MAX_WAIT_SECONDS
        remaining = countdown_limit - elapsed

        if (
            no_trade_cutoff_seconds is not None and
            elapsed > no_trade_cutoff_seconds and
            not has_open_trade
        ):
            print("\r" + " " * max(len(last_status_line), 80), end="\r", flush=True)
            print(f"Corte 09:40 sin trade OPEN ({int(elapsed)}s). Deteniendo replay del dia.")
            stop_replay(stop_button)
            return False

        if elapsed > MAX_WAIT_SECONDS:
            print("\r" + " " * max(len(last_status_line), 80), end="\r", flush=True)
            print(f"Timeout ({MAX_WAIT_SECONDS}s) esperando resultado terminal. Saltando al siguiente dia.")
            stop_replay(stop_button)
            return False

        elapsed_second = int(elapsed)
        if elapsed_second != last_print_second:
            open_status = "OPEN" if has_open_trade else "sin OPEN"
            last_status_line = (
                f"Esperando resultado... transcurrido {format_countdown(elapsed)} | "
                f"restan {format_countdown(remaining)} | {open_status}"
            )
            print("\r" + last_status_line + " " * max(0, 100 - len(last_status_line)), end="", flush=True)
            last_print_second = elapsed_second

        time.sleep(POLL_SECONDS)


def read_trade_result(path, date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    default_row = {
        "fecha": f"{yyyy}-{mm}-{dd}",
        "result TP SL BE": "NO_CSV",
    }

    if not os.path.exists(path):
        print(f"Sin CSV para {default_row['fecha']}; no se sobrescriben columnas de datos.")
        return default_row

    if os.path.getmtime(path) < RUN_STARTED_AT:
        if not result_is_terminal(path):
            print(f"CSV viejo no terminal ignorado para {default_row['fecha']}: {path}")
            return default_row
        print(f"CSV terminal preservado usado para {default_row['fecha']}: {path}")

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        try:
            row = next(reader)
        except StopIteration:
            default_row["result TP SL BE"] = "EMPTY_CSV"
            return default_row

    if "fecha" not in row:
        legacy_result = row.get("RESULT", "NO_TRADE")
        legacy_ticks = None

        if legacy_result == "TP" and row.get("ENTRY") and row.get("TP"):
            legacy_ticks = abs((float(row["TP"]) - float(row["ENTRY"])) / 0.25)

        if legacy_result == "SL" and row.get("ENTRY") and row.get("SL"):
            legacy_ticks = -abs((float(row["ENTRY"]) - float(row["SL"])) / 0.25)

        row = {
            "fecha": default_row["fecha"],
            "SL_price": row.get("SL"),
            "Entry_price": row.get("ENTRY"),
            "TP_price": row.get("TP"),
            "result TP SL BE": legacy_ticks if legacy_ticks is not None else legacy_result,
        }

    row["fecha"] = row.get("fecha") or default_row["fecha"]
    row["result TP SL BE"] = row.get("result TP SL BE") or "OPEN"
    row["Exit_price"] = row.get("Exit_price") or calculate_exit_price(row)
    return row


def calculate_exit_price(row):
    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()

    if result_label == "TP":
        return row.get("TP_price") or row.get("TP") or ""

    if result_label == "SL":
        return row.get("SL_price") or row.get("SL") or ""

    if result_label != "EXIT":
        return ""

    return ""


def to_number(value):
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def get_or_create_headers(ws):
    for col in range(ws.max_column, 0, -1):
        if ws.cell(row=3, column=col).value in EXCLUDED_EXCEL_HEADERS:
            ws.delete_cols(col)

    default_headers = [
        "fecha",
        "or_low",
        "or_high",
        "range",
        "VWAP_entry",
        "Body",
        "Volume_entry",
        "Delta_entry",
        "BreakOut_SPEED",
        "BreakOut_TICKS_PER_SEC",
        "score total",
        "SL_price",
        "Entry_price",
        "TP_price",
        "SL_ticks",
        "TP_ticks",
        "Exit_price",
        "result TP SL BE",
        "Side",
        "Speed_Profile",
        "MAE_ticks",
        "MFE_ticks",
    ]

    headers = [
        ws.cell(row=3, column=col).value
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=3, column=col).value and ws.cell(row=3, column=col).value not in EXCLUDED_EXCEL_HEADERS
    ]

    if not headers:
        headers = default_headers[:]

    for header in default_headers:
        if header not in headers:
            headers.append(header)

    for csv_header in get_csv_headers_for_dates():
        if csv_header not in headers:
            headers.append(csv_header)

    for feature_header in get_feature_scan_headers_for_dates():
        if feature_header not in headers:
            headers.append(feature_header)

    for header in ("ExitTime_NY", "Trade_Duration"):
        if header in headers:
            headers.remove(header)

    entry_time_index = headers.index("EntryTime_NY") + 1
    headers[entry_time_index:entry_time_index] = ["ExitTime_NY", "Trade_Duration"]

    trade_path_headers = [
        "Largest_MAE_pullback_ticks",
        "Largest_MFE_pullup_ticks",
        "Number_of_Pullbacks_during_Trade",
        "Number_of_PullUps_during_Trade",
        "Max_Speed_MAE_during_trade",
        "Max_Speed_MFE_during_trade",
    ]
    for header in trade_path_headers:
        if header in headers:
            headers.remove(header)

    mfe_index = headers.index("MFE_ticks") + 1
    headers[mfe_index:mfe_index] = trade_path_headers

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    return headers


def get_csv_headers_for_dates():
    headers = []

    for date in DATES_DST:
        path = expected_result_path(date)

        if not os.path.exists(path):
            continue

        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue

                for field in reader.fieldnames:
                    if field and field not in EXCLUDED_EXCEL_HEADERS and field not in headers:
                        headers.append(field)
        except OSError:
            continue

    return headers


def get_feature_scan_headers_for_dates():
    # Column names from the Feature Scanner sidecars, appended after the score
    # columns. Shared keys (fecha, or_low, or_high...) are filtered by the caller.
    headers = []

    for date in DATES_DST:
        path = expected_feature_scan_path(date)

        if not os.path.exists(path):
            continue

        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue

                for field in reader.fieldnames:
                    if field and field not in EXCLUDED_EXCEL_HEADERS and field not in headers:
                        headers.append(field)
        except OSError:
            continue

    return headers


def read_feature_scan(path):
    # First (only) data row of a Feature Scanner sidecar as a dict, or {}.
    if not os.path.exists(path):
        return {}

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return {}

    return row or {}


def update_score_workbook():
    os.makedirs(os.path.dirname(SCORE_WORKBOOK), exist_ok=True)

    if os.path.exists(SCORE_WORKBOOK):
        os.remove(SCORE_WORKBOOK)
        print(f"Excel anterior eliminado: {SCORE_WORKBOOK}")

    if os.path.exists(SCORE_WORKBOOK_TEMPLATE):
        wb = load_workbook(SCORE_WORKBOOK_TEMPLATE)
    elif os.path.exists(SCORE_WORKBOOK_FALLBACK):
        wb = load_workbook(SCORE_WORKBOOK_FALLBACK)
    elif os.path.exists(SCORE_WORKBOOK_TEMPLATE_FALLBACK):
        wb = load_workbook(SCORE_WORKBOOK_TEMPLATE_FALLBACK)
    else:
        wb = Workbook()

    ws = wb.active

    headers = get_or_create_headers(ws)
    first_row = 4
    last_row = first_row + len(DATES_DST) - 1

    for row in range(first_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    for row_offset, date in enumerate(DATES_DST, start=first_row):
        result = read_trade_result(expected_result_path(date), date)

        # Merge the new Feature Scanner columns. setdefault => score CSV wins on
        # shared keys (fecha, or_low, or_high...); feature-only columns are added.
        feature_row = read_feature_scan(expected_feature_scan_path(date))
        for key, value in feature_row.items():
            result.setdefault(key, value)

        missing_result_file = result.get("result TP SL BE") in ("NO_CSV", "EMPTY_CSV")

        for col, header in enumerate(headers, start=1):
            if header not in result:
                if missing_result_file:
                    ws.cell(row=row_offset, column=col).value = None
                continue

            value = to_number(result.get(header))

            if value is None and header != "result TP SL BE":
                continue

            ws.cell(row=row_offset, column=col).value = value

    result_col = headers.index("result TP SL BE") + 1
    result_letter = get_column_letter(result_col)

    ws["A1"] = "win rate"
    ws["B1"] = "profit factor"
    ws["A2"] = f'=IF((COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0"))=0,"",COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")/(COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0")))'
    ws["B2"] = f'=IF(ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row}))=0,IF(SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})>0,"INF",""),SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})/ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row})))'
    ws["A2"].number_format = "0.00%"
    ws["B2"].number_format = "0.00"

    for row_idx in range(first_row, last_row + 1):
        ws.cell(row=row_idx, column=result_col).number_format = "+0.##;-0.##;0"

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header in ("ExitTime_NY", "Trade_Duration"):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    widths = {
        "A": 12,
        "B": 12,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 16,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header in ("ExitTime_NY", "Trade_Duration"):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    try:
        wb.save(SCORE_WORKBOOK)
        print(f"Excel actualizado: {SCORE_WORKBOOK}")
    except PermissionError:
        wb.save(SCORE_WORKBOOK_FALLBACK)
        print("WARNING: Excel original bloqueado/abierto; guarde copia actualizada en:")
        print(SCORE_WORKBOOK_FALLBACK)


# =========================================================
# LOOP PRINCIPAL V11 X1/X10
# =========================================================


def reset_replay_run_state(output_folder):
    """Archiva resultados previos y reinicia el estado contable de Telegram."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(output_folder)

    if output_path.exists():
        archive_path = output_path.parent / f"_archive_{output_path.name}_reset_{timestamp}"
        suffix = 1
        while archive_path.exists():
            archive_path = output_path.parent / (
                f"_archive_{output_path.name}_reset_{timestamp}_{suffix}"
            )
            suffix += 1
        shutil.move(str(output_path), str(archive_path))
        print(f"Resultados previos archivados: {archive_path}")

    output_path.mkdir(parents=True, exist_ok=True)

    # No tocar telegram_message_ids.txt aqui: clear_telegram_before_run lo usa
    # para borrar mensajes viejos antes de vaciarlo. El balance si debe volver a
    # arrancar desde TelegramStartingBalance ($150,000 en el exporter).
    for file_name in ("telegram_balance.json", "telegram_challenge_passed.flag"):
        state_path = Path(RESULTS_FOLDER) / file_name
        if not state_path.exists():
            continue
        backup_path = state_path.with_name(
            f"{state_path.stem}_backup_{timestamp}{state_path.suffix}"
        )
        shutil.move(str(state_path), str(backup_path))
        print(f"Estado Telegram reiniciado; backup: {backup_path}")


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Corre las temporadas DST 2022-2026 (desde 04/04/2022) con flujo v11: "
            "Fase 1 valida sincronia X1/X10 en las primeras "
            f"{SYNC_CHECK_DATE_COUNT} fechas; Fase 2 corre todo SOLO en X10."
        )
    )
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Compatibilidad con test_fechas_conflictivas.py; este runner ya usa modo quick por defecto.",
    )
    parser.add_argument(
        "--prepare-only",
        action="store_true",
        help="Muestra plan y versiones sin iniciar Replay.",
    )
    parser.add_argument(
        "--compare-only",
        action="store_true",
        help="Regenera el reporte con corridas guardadas.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Ignora corridas guardadas y vuelve a correr todo.",
    )
    parser.add_argument(
        "--reset-state",
        action="store_true",
        help="Archiva resultados de esta corrida y reinicia balance/stats Telegram.",
    )
    parser.add_argument(
        "--step",
        action="store_true",
        help="Pide ENTER antes de cada fecha.",
    )
    parser.add_argument(
        "--x1-only",
        action="store_true",
        help="Solo corre la fase X1 canónica.",
    )
    parser.add_argument(
        "--x10-only",
        action="store_true",
        help="Solo corre la fase X10; requiere snapshots v11 previos de X1.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        metavar="N",
        help="Recorta a las ultimas N fechas (prueba rapida). 0 = todas. Una semana = 5.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    global DATES_DST
    if args.limit and args.limit > 0:
        DATES_DST = DATES_DST[-args.limit:]
        print(
            f"LIMITADO a las ultimas {len(DATES_DST)} fechas para prueba rapida: "
            f"{DATES_DST[0]} -> {DATES_DST[-1]}"
        )

    print_excel_sizes_by_year()

    date_iso_list = [replay_sync.date_iso_from_replay(date) for date in DATES_DST]
    output_folder = os.path.join(
        RESULTS_FOLDER,
        "visual_tests",
        "04_run_replay_score_trade_results_dst_2025_2026_runs",
    )

    # --x1-only / --x10-only conservan el modo manual de UNA fase sobre todas
    # las fechas. Sin flags corre el flujo de dos fases (sync 10 -> full X10).
    single_phase = args.x1_only or args.x10_only

    sync_dates = date_iso_list[:SYNC_CHECK_DATE_COUNT]
    run_plan_sync = replay_sync.build_run_plan(quick=True)
    run_plan_x10 = replay_sync.build_run_plan(quick=True, x10_only=True)

    print(
        f"\nINICIANDO REPLAY DE TEMPORADAS DST 2022-2026 V11 "
        f"({len(DATES_DST)} sesiones desde {DATES_DST[0]})\n"
        f"Fecha NY actual: {TODAY_NY:%d/%m/%Y} | "
        f"Ultima fecha permitida: {LAST_REPLAY_DATE:%d/%m/%Y}\n"
        f"Version esperada: {replay_sync.EXPECTED_EXPORTER_VERSION}\n"
        f"Resultados de validacion: {output_folder}\n"
    )
    if single_phase:
        run_plan_manual = replay_sync.build_run_plan(
            quick=True,
            x1_only=args.x1_only,
            x10_only=args.x10_only,
        )
        print("Plan (modo manual, una fase):")
        for run_name, speed_label, _ in run_plan_manual:
            print(f"  - {run_name}: Replay {speed_label}")
    else:
        print("Plan:")
        print(
            f"  - FASE 1 (sincronia): primeras {len(sync_dates)} fechas en "
            "X1_R1 y X10_R1; aborta si divergen."
        )
        print(
            f"  - FASE 2 (historia completa): {len(date_iso_list)} fechas "
            "desde 04/04/2022 SOLO en X10 (las ya corridas se saltan)."
        )

    if args.prepare_only:
        print("\nPREPARE-ONLY correcto. No se inició Replay.")
        return 0

    if args.reset_state and args.compare_only:
        print("WARNING: --reset-state se ignora en --compare-only.")
    elif args.reset_state:
        print("Reiniciando resultados/Telegram para corrida desde cero...")
        reset_replay_run_state(output_folder)

    if not args.compare_only:
        print("Ejecutando limpieza unica de Telegram antes de la primera fecha...")
        clear_telegram_before_run(RESULTS_FOLDER)

    if single_phase:
        progress_meta = {
            "stage_index": 1,
            "stage_total": 1,
            "stage_label": "DST 2022-2026 v11 (modo manual)",
            "stage_period": f"{DATES_DST[0]} -> {DATES_DST[-1]}",
            "session_roots": [output_folder],
            "run_label": "DST 2022-2026",
            "stats_root": output_folder,
        }
        passed, failures = replay_sync.run_replay_period(
            date_iso_list,
            output_folder=Path(output_folder),
            run_plan=run_plan_manual,
            report_prefix="dst_2022_2026_v11",
            force=args.force,
            step=args.step,
            compare_only=args.compare_only,
            replay_to_time=REPLAY_END_TIME,
            progress_meta=progress_meta,
        )
    else:
        # ── FASE 1: sincronia X1==X10 en las primeras fechas ──────────────
        progress_meta_sync = {
            "stage_index": 1,
            "stage_total": 2,
            "stage_label": f"Sincronia X1/X10 ({len(sync_dates)} fechas)",
            "stage_period": f"{DATES_DST[0]} -> {DATES_DST[len(sync_dates) - 1]}",
            "global_target": len(sync_dates),
            "session_roots": [output_folder],
            "run_label": "DST 2022-2026",
            "stats_root": output_folder,
        }
        passed_sync, failures_sync = replay_sync.run_replay_period(
            sync_dates,
            output_folder=Path(output_folder),
            run_plan=run_plan_sync,
            report_prefix="dst_2022_2026_sync10_v11",
            force=args.force,
            step=args.step,
            compare_only=args.compare_only,
            replay_to_time=REPLAY_END_TIME,
            progress_meta=progress_meta_sync,
        )

        phase1_synced = passed_sync and not failures_sync
        sync_report_path = (
            Path(output_folder) / "dst_2022_2026_sync10_v11_resumen.txt"
        )
        sync_message = [
            "EW Opening Range | FASE 1 X1/X10 "
            + ("SINCRONIZADO" if phase1_synced else "NO SINCRONIZADO"),
            f"Fechas verificadas: {len(sync_dates)}",
            f"Periodo: {DATES_DST[0]} -> {DATES_DST[len(sync_dates) - 1]}",
            "Comparacion: campos operativos (entrada/SL/TP/salida/resultado/score).",
        ]
        if phase1_synced:
            sync_message.append("X1_R1 y X10_R1 coinciden. Se habilita FASE 2 X10 completa.")
        else:
            sync_message.append("Replay detenido: revisar diferencias antes de FASE 2.")
            if sync_report_path.exists():
                sync_message.append(f"Reporte: {sync_report_path}")
            if failures_sync:
                sync_message.append("Errores:")
                for date_iso, run_name, reason in failures_sync[:5]:
                    sync_message.append(f"- {date_iso} {run_name}: {reason}")
        send_text(RESULTS_FOLDER, "\n".join(sync_message))

        if not passed_sync or failures_sync:
            failed_dates = [
                (date_iso, f"{run_name}: {reason}")
                for date_iso, run_name, reason in failures_sync
            ]
            print("\nFASE 1 FALLO: sincronia X1/X10 divergente o fechas con error.")
            for failed_date, error in failed_dates:
                print(f"- {failed_date}: {error}")
            send_run_summary(
                RESULTS_FOLDER,
                DATES_DST[: len(sync_dates)],
                failed_dates,
                "DST 2022-2026 FASE 1 (sincronia X1/X10) FALLO — corrida detenida",
            )
            print("\nCORRIDA DETENIDA: NO se inicia la fase X10 completa.\n")
            return 1

        print(
            f"\nFASE 1 OK: {len(sync_dates)} fechas sincronizadas X1==X10. "
            "Iniciando FASE 2 (X10 completo desde 04/04/2022)...\n"
        )

        # ── FASE 2: historia completa SOLO X10 ────────────────────────────
        # Sin global_target/session_roots: la metrica de sincronia X1==X10 no
        # aplica en X10-only; el avance real va en la linea de etapa (done/total).
        progress_meta_full = {
            "stage_index": 2,
            "stage_total": 2,
            "stage_label": "Historia completa X10 desde 04/04/2022",
            "stage_period": f"{DATES_DST[0]} -> {DATES_DST[-1]}",
            "run_label": "DST 2022-2026",
            "stats_root": output_folder,
        }
        passed, failures = replay_sync.run_replay_period(
            date_iso_list,
            output_folder=Path(output_folder),
            run_plan=run_plan_x10,
            report_prefix="dst_2022_2026_x10_full_v11",
            force=args.force,
            step=args.step,
            compare_only=args.compare_only,
            replay_to_time=REPLAY_END_TIME,
            progress_meta=progress_meta_full,
        )

    update_score_workbook()

    print_feature_scans()

    failed_dates = [
        (date_iso, f"{run_name}: {reason}")
        for date_iso, run_name, reason in failures
    ]
    if failed_dates:
        print("\nFECHAS CON ERROR DE SCRIPT/UI:")
        for failed_date, error in failed_dates:
            print(f"- {failed_date}: {error}")

    send_run_summary(
        RESULTS_FOLDER,
        DATES_DST,
        failed_dates,
        "Temporadas DST 2022-2026 v11 (sync 10 X1/X10 + full X10)",
    )

    print("\nTERMINO LA PRUEBA DE TEMPORADAS DST COMPLETAS 2022-2026 V11.\n")
    return 0 if passed and not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
