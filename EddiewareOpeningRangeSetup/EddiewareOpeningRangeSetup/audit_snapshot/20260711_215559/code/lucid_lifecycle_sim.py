# -*- coding: utf-8 -*-
"""Full Lucid (LucidPro) lifecycle Monte Carlo: eval -> pass -> funded -> payouts.

Objective is NOT zero losing years; it is to pass an account and pull >=3
withdrawals, retrying on bust. So we simulate the whole cash cycle on the real
UP + anti-absorption ORB trade distribution (block bootstrap preserves streaks).

LucidPro rules (verified 2026):
  100k: target $6,000, trailing max-loss $3,000 (EOD), 6 mini, payout>=$750/3d, cap $2,500 then $3,000
  150k: target $9,000, trailing max-loss $4,500 (EOD), 10 mini, payout>=$1,000/3d, cap $3,000 then $3,500
Trailing drawdown is END-OF-DAY and locks at the start balance once the buffer
(=max-loss) is built. We treat 1 trade per trading day. 90/10 split applied to
withdrawals. min withdrawal $500.

Outputs: per-account expected withdrawals & pass rate by size, plus an Excel of a
representative run that passed and withdrew >=3 times.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT_DIR = Path(r"C:\Users\k_99_\Desktop\codding\OpeningRangeSetup\EddiewareOpeningRangeSetup"
               r"\_x") and Path(r"C:\Users\k_99_\Documents\Indicador ATAS\outputs\orb_bigmove_1s")
MD_DIR = Path(r"C:\Users\k_99_\Desktop\contexto_OR_catboost")
TICK_USD = 5.0
COMMISSION = 2.0
SPLIT = 0.90
MIN_WD = 500.0
MIN_DAYS = 3
RNG = np.random.default_rng(2026)
BLOCK = 8

ACC = {
    "100k": dict(target=6000, maxloss=3000, max_mini=6, pay_min=750, cap1=2500, cap2=3000),
    "150k": dict(target=9000, maxloss=4500, max_mini=10, pay_min=1000, cap1=3000, cap2=3500),
}

# real trade stream: UP + anti-absorption, trail 40/20/40
pnl = pd.read_csv(OUT_DIR / "orb_trailing_pnl.csv")
feat = pd.read_csv(OUT_DIR / "orb_features_labels_1s.csv")[["date", "direction_up", "vpt_30"]]
t = pnl.merge(feat, on="date", how="left")
vhi = t["vpt_30"].quantile(2/3)
t = t[(t.direction_up == 1.0) & (t.vpt_30 <= vhi)].sort_values("date")
SERIES = t["trail_40_20_40"].to_numpy() - COMMISSION      # net ticks / trade, 1 contract


def draw(n):
    out = []
    while len(out) < n:
        i = RNG.integers(0, len(SERIES))
        out.extend(SERIES[i:i + BLOCK])
    return np.asarray(out[:n])


def eval_phase(size, a, stream, k):
    """Return (passed, k_next). size in mini contracts."""
    bal = high = 0.0
    while k < len(stream):
        bal += stream[k] * size * TICK_USD
        k += 1
        high = max(high, bal)
        floor = high - a["maxloss"]
        if bal <= floor:
            return False, k
        if bal >= a["target"]:
            return True, k
    return False, k


def funded_phase(size, a, stream, k, record=None):
    """Return (n_withdrawals, total_withdrawn_net, k_next). Records rows if given."""
    bal = high = 0.0
    days = 0
    since_pay = 0
    n_wd = 0
    total = 0.0
    cyc = 0
    while k < len(stream):
        days += 1
        since_pay += 1
        bal += stream[k] * size * TICK_USD
        high = max(high, bal)
        # trailing floor, locked at start (0) once buffer built
        floor = min(high - a["maxloss"], 0.0)
        event = ""
        if bal <= floor:
            if record is not None:
                record.append((days, stream[k] * size * TICK_USD, bal, high, floor, "FUNDED", "BUST"))
            return n_wd, total, k + 1
        # withdrawal: eligible profit and cadence
        if bal >= a["pay_min"] and since_pay >= MIN_DAYS:
            cap = a["cap1"] if cyc == 0 else a["cap2"]
            amt = min(bal, cap)              # pull profit down toward start buffer
            if amt >= MIN_WD:
                net = amt * SPLIT
                bal -= amt
                total += net
                n_wd += 1
                cyc += 1
                since_pay = 0
                event = f"WITHDRAW ${amt:.0f} (net ${net:.0f})"
        if record is not None:
            record.append((days, stream[k] * size * TICK_USD, bal, high, floor, "FUNDED", event))
        k += 1
    return n_wd, total, k


def simulate_lifecycle(size, a, max_trades=600, record=None):
    stream = draw(max_trades)
    passed, k = eval_phase(size, a, stream, 0)
    if record is not None and passed:
        # rebuild eval trace for the record
        bal = high = 0.0
        for j in range(k):
            bal += stream[j] * size * TICK_USD
            high = max(high, bal)
            floor = high - a["maxloss"]
            ev = "PASSED" if bal >= a["target"] else ""
            record.append((j + 1, stream[j] * size * TICK_USD, bal, high, floor, "EVAL", ev))
            if ev:
                break
    if not passed:
        return False, 0, 0.0
    n_wd, total, _ = funded_phase(size, a, stream, k, record)
    return True, n_wd, total


def sweep():
    print(f"Trade stream: n={len(SERIES)}, WR={ (SERIES>0).mean()*100:.1f}%, "
          f"EV/trade=${SERIES.mean()*TICK_USD:.1f} (1 mini)\n")
    N = 8000
    best = {}
    for name, a in ACC.items():
        print(f"=== Lucid {name} (target ${a['target']}, DD ${a['maxloss']}, max {a['max_mini']} mini) ===")
        print(f"{'size':>4} {'pass%':>6} {'E[wd]':>6} {'P(>=3wd)':>8} {'E[$net]':>9}")
        bwd = (0, -1)
        for size in range(1, a["max_mini"] + 1):
            res = [simulate_lifecycle(size, a) for _ in range(N)]
            passed = np.array([r[0] for r in res])
            wds = np.array([r[1] for r in res])
            nets = np.array([r[2] for r in res])
            e_wd = wds.mean()
            p3 = (wds >= 3).mean() * 100
            print(f"{size:>4} {passed.mean()*100:>5.1f}% {e_wd:>6.2f} {p3:>7.1f}% {nets.mean():>9.0f}")
            if e_wd > bwd[1]:
                bwd = (size, e_wd)
        best[name] = bwd[0]
        print(f"  -> size max E[withdrawals] = {bwd[0]} mini\n")
    return best


def make_excel(name, size):
    a = ACC[name]
    # find a run that passes and withdraws >=3 times
    for _ in range(20000):
        rec = []
        passed, n_wd, total = simulate_lifecycle(size, a, record=rec)
        if passed and n_wd >= 3:
            break
    wb = Workbook()
    ws = wb.active
    ws.title = "Passed_Run"
    hdr = ["trade/day", "pnl_$", "balance_$", "eod_high_$", "dd_floor_$", "phase", "event"]
    fillh = PatternFill("solid", fgColor="1F4E5F")
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h); cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fillh
    for r, row in enumerate(rec, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, round(v, 1) if isinstance(v, float) else v)
            cell.alignment = Alignment(horizontal="center")
            if row[6] == "PASSED":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
            elif "WITHDRAW" in str(row[6]):
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            elif row[6] == "BUST":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
    for col, w in zip("ABCDEFG", [10, 10, 12, 12, 12, 8, 26]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    s = wb.create_sheet("Summary")
    rows = [
        ["Lucid account", name], ["Setup", "UP breakout + anti-absorption + trail 40/20/40"],
        ["Contracts (mini)", size], ["Profit target $", a["target"]],
        ["Trailing DD $ (EOD)", a["maxloss"]], ["Withdrawals in this run", n_wd],
        ["Net withdrawn $ (after 90/10)", round(total)],
        ["Note", "Backtest label + block-bootstrap; EOD trailing DD; forward test before capital"],
    ]
    for r, (k, v) in enumerate(rows, 1):
        s.cell(r, 1, k).font = Font(bold=True); s.cell(r, 2, v)
    s.column_dimensions["A"].width = 30; s.column_dimensions["B"].width = 48

    dest = MD_DIR / f"Lucid_{name}_passed_run_simulation.xlsx"
    wb.save(dest)
    print(f"Excel escrito: {dest}  (withdrawals={n_wd}, net ${total:.0f})")
    return dest


if __name__ == "__main__":
    best = sweep()
    for name in ("100k", "150k"):
        make_excel(name, best[name])
