import csv
import os
import re
from copy import copy
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG
# =========================================================

EXPORT_FOLDER = r"C:\Users\k_99_\Desktop\codding\data_footprint_generator"
RESULTS_FOLDER = os.path.join(EXPORT_FOLDER, "trade_results_score")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCORE_WORKBOOK_NAME = "Score_indicator_results_from_root_folder"
SCORE_WORKBOOK = os.path.join(BASE_DIR, f"{SCORE_WORKBOOK_NAME}.xlsx")
FORMAT_TEMPLATE_WORKBOOK = os.path.join(BASE_DIR, "Score_indicator_results_updated_fallback.xlsx")
RESULT_FILE_RE = re.compile(r"^score_trade_result_(\d{4})-(\d{2})-(\d{2})_NY\.csv$", re.IGNORECASE)


# =========================================================
# FUNCIONES
# =========================================================

def discover_result_files():
    files = []

    if not os.path.isdir(RESULTS_FOLDER):
        raise FileNotFoundError(f"No existe la carpeta de resultados: {RESULTS_FOLDER}")

    for filename in os.listdir(RESULTS_FOLDER):
        match = RESULT_FILE_RE.match(filename)

        if not match:
            continue

        yyyy, mm, dd = match.groups()
        files.append({
            "date_ddmmyyyy": f"{dd}/{mm}/{yyyy}",
            "date_iso": f"{yyyy}-{mm}-{dd}",
            "path": os.path.join(RESULTS_FOLDER, filename),
        })

    return sorted(files, key=lambda item: item["date_iso"])


def get_new_workbook_copy_path():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(BASE_DIR, f"{SCORE_WORKBOOK_NAME}_{timestamp}.xlsx")


def load_score_workbook_template():
    if os.path.exists(FORMAT_TEMPLATE_WORKBOOK):
        print(f"Usando formato base: {FORMAT_TEMPLATE_WORKBOOK}")
        return load_workbook(FORMAT_TEMPLATE_WORKBOOK)

    if os.path.exists(SCORE_WORKBOOK):
        print(f"Usando workbook existente: {SCORE_WORKBOOK}")
        return load_workbook(SCORE_WORKBOOK)

    os.makedirs(os.path.dirname(SCORE_WORKBOOK), exist_ok=True)
    return Workbook()


def copy_row_format(ws, source_row, target_row):
    if source_row == target_row:
        return

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def parse_result_ticks(value):
    if value in (None, "", "OPEN", "NO_TRADE"):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    normalized = str(value).strip().upper()

    if normalized in ("TP", "EXIT"):
        return 1.0

    if normalized == "SL":
        return -1.0

    if normalized == "BE":
        return 0.0

    try:
        return float(normalized.replace("+", ""))
    except ValueError:
        return None


def read_result_ticks(path):
    if not os.path.exists(path):
        return None

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return None

    if not row:
        return None

    value = row.get("result TP SL BE") or row.get("RESULT")
    return parse_result_ticks(value)


def read_trade_result(path, date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    default_row = {
        "fecha": f"{yyyy}-{mm}-{dd}",
        "result TP SL BE": "NO_CSV",
    }

    if not os.path.exists(path):
        print(f"Sin CSV para {default_row['fecha']}; no se sobrescriben columnas de datos.")
        return default_row

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        try:
            row = next(reader)
        except StopIteration:
            default_row["result TP SL BE"] = "EMPTY_CSV"
            return default_row

    if "fecha" not in row:
        legacy_result = row.get("RESULT", "NO_TRADE")
        legacy_ticks = None

        if legacy_result == "TP" and row.get("ENTRY") and row.get("TP"):
            legacy_ticks = abs((float(row["TP"]) - float(row["ENTRY"])) / 0.25)

        if legacy_result == "SL" and row.get("ENTRY") and row.get("SL"):
            legacy_ticks = -abs((float(row["ENTRY"]) - float(row["SL"])) / 0.25)

        row = {
            "fecha": default_row["fecha"],
            "SL_price": row.get("SL"),
            "Entry_price": row.get("ENTRY"),
            "TP_price": row.get("TP"),
            "result TP SL BE": legacy_ticks if legacy_ticks is not None else legacy_result,
        }

    row["fecha"] = row.get("fecha") or default_row["fecha"]
    row["result TP SL BE"] = row.get("result TP SL BE") or "OPEN"
    row["Exit_price"] = row.get("Exit_price") or calculate_exit_price(row)
    return row


def calculate_exit_price(row):
    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()

    if result_label == "TP":
        return row.get("TP_price") or row.get("TP") or ""

    if result_label == "SL":
        return row.get("SL_price") or row.get("SL") or ""

    if result_label != "EXIT":
        return ""

    return ""


def to_number(value):
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def get_csv_headers_for_result_files(result_files):
    headers = []

    for result_file in result_files:
        path = result_file["path"]

        if not os.path.exists(path):
            continue

        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue

                for field in reader.fieldnames:
                    if field and field not in headers:
                        headers.append(field)
        except OSError:
            continue

    return headers


def get_or_create_headers(ws, result_files):
    default_headers = [
        "fecha",
        "or_low",
        "or_high",
        "range",
        "VWAP_entry",
        "Body",
        "Volume_entry",
        "Delta_entry",
        "BreakOut_SPEED",
        "BreakOut_TICKS_PER_SEC",
        "Entry_Score",
        "Min_IN_TRADE_SCORE",
        "Max_IN_TRADE_SCORE",
        "SL_price",
        "Entry_price",
        "TP_price",
        "SL_ticks",
        "TP_ticks",
        "Exit_price",
        "result TP SL BE",
        "Side",
        "Speed_Profile",
        "MAE_ticks",
        "MFE_ticks",
    ]

    headers = [
        ws.cell(row=3, column=col).value
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=3, column=col).value
    ]

    if not headers:
        headers = default_headers[:]

    for header in default_headers:
        if header not in headers:
            headers.append(header)

    for csv_header in get_csv_headers_for_result_files(result_files):
        if csv_header not in headers:
            headers.append(csv_header)

    in_trade_score_headers = ["Min_IN_TRADE_SCORE", "Max_IN_TRADE_SCORE"]
    for header in in_trade_score_headers:
        if header in headers:
            headers.remove(header)

    entry_score_index = headers.index("Entry_Score") + 1
    headers[entry_score_index:entry_score_index] = in_trade_score_headers

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    return headers


def update_score_workbook():
    result_files = discover_result_files()

    if not result_files:
        print(f"No encontre CSV de resultados en: {RESULTS_FOLDER}")
        return

    wb = load_score_workbook_template()

    ws = wb.active

    headers = get_or_create_headers(ws, result_files)
    first_row = 4
    last_row = first_row + len(result_files) - 1

    for row in range(first_row, max(ws.max_row, last_row) + 1):
        copy_row_format(ws, first_row, row)

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    for row_offset, result_file in enumerate(result_files, start=first_row):
        result = read_trade_result(result_file["path"], result_file["date_ddmmyyyy"])
        missing_result_file = result.get("result TP SL BE") in ("NO_CSV", "EMPTY_CSV")

        for col, header in enumerate(headers, start=1):
            if header not in result:
                if missing_result_file:
                    ws.cell(row=row_offset, column=col).value = None
                continue

            value = to_number(result.get(header))

            if value is None and header != "result TP SL BE":
                continue

            ws.cell(row=row_offset, column=col).value = value

    result_col = headers.index("result TP SL BE") + 1
    result_letter = get_column_letter(result_col)

    ws["A1"] = "win rate"
    ws["B1"] = "profit factor"
    ws["A2"] = f'=IF((COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0"))=0,"",COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")/(COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0")))'
    ws["B2"] = f'=IF(ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row}))=0,IF(SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})>0,"INF",""),SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})/ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row})))'
    ws["A2"].number_format = "0.00%"
    ws["B2"].number_format = "0.00"

    for row_idx in range(first_row, last_row + 1):
        ws.cell(row=row_idx, column=result_col).number_format = "+0.##;-0.##;0"

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    widths = {
        "A": 12,
        "B": 12,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 16,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    try:
        wb.save(SCORE_WORKBOOK)
        print(f"Excel actualizado: {SCORE_WORKBOOK}")
    except PermissionError:
        new_workbook_path = get_new_workbook_copy_path()
        wb.save(new_workbook_path)
        print("WARNING: Excel principal bloqueado/abierto; cree un archivo nuevo en:")
        print(new_workbook_path)

    print(f"Resultados leidos sin borrar archivos: {len(result_files)}")
    print(f"Carpeta fuente: {RESULTS_FOLDER}")


# =========================================================
# LOOP PRINCIPAL
# =========================================================

if __name__ == "__main__":
    print("\nLEYENDO SCORE TRADE RESULTS DESDE CARPETA RAIZ\n")
    update_score_workbook()
    print("\nTERMINO LA LECTURA DE RESULTADOS.\n")
