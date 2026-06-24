import argparse
from pywinauto import Desktop
import csv
from datetime import date, datetime, timedelta
import os
import time
from pathlib import Path
import pyperclip
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import replay_sync_runner_common_after_sync as replay_sync
from telegram_run_summary_after_sync import clear_telegram_before_run, send_run_summary

# =========================================================
# CONFIG
# =========================================================

# Temporadas DST completas de Nueva York para 2025 y 2026.
# Se excluyen fines de semana y cierres completos del mercado de EE. UU.
# Los cierres tempranos se incluyen porque el replay solo usa 09:30-09:50 NY.
# Formato requerido por el panel Replay de ATAS: dd/mm/yyyy.

DST_SEASONS = (
    (date(2025, 3, 10), date(2025, 10, 31)),
    (date(2026, 3, 9), date(2026, 10, 30)),
)

MARKET_CLOSED_DATES = {
    # 2025
    date(2025, 1, 1),   # New Year's Day
    date(2025, 1, 9),   # National Day of Mourning
    date(2025, 1, 20),  # Martin Luther King Jr. Day
    date(2025, 2, 17),  # Washington's Birthday
    date(2025, 4, 18),  # Good Friday
    date(2025, 5, 26),  # Memorial Day
    date(2025, 6, 19),  # Juneteenth
    date(2025, 7, 4),   # Independence Day
    date(2025, 9, 1),   # Labor Day
    date(2025, 11, 27), # Thanksgiving Day
    date(2025, 12, 25), # Christmas Day

    # 2026
    date(2026, 1, 1),   # New Year's Day
    date(2026, 1, 19),  # Martin Luther King Jr. Day
    date(2026, 2, 16),  # Washington's Birthday
    date(2026, 4, 3),   # Good Friday
    date(2026, 5, 25),  # Memorial Day
    date(2026, 6, 19),  # Juneteenth
    date(2026, 7, 3),   # Independence Day observed
    date(2026, 9, 7),   # Labor Day
    date(2026, 11, 26), # Thanksgiving Day
    date(2026, 12, 25), # Christmas Day
}


def build_trading_dates(start_date, end_date):
    trading_dates = []
    current = start_date

    while current <= end_date:
        if current.weekday() < 5 and current not in MARKET_CLOSED_DATES:
            trading_dates.append(current.strftime("%d/%m/%Y"))
        current += timedelta(days=1)

    return trading_dates


def replay_date_is_allowed(date_ddmmyyyy):
    today_ny = datetime.now(ZoneInfo("America/New_York")).date()
    last_replay_date = today_ny - timedelta(days=1)
    replay_date = datetime.strptime(date_ddmmyyyy, "%d/%m/%Y").date()
    return replay_date <= last_replay_date


ALL_DATES_DST = [
    replay_date
    for season_start, season_end in DST_SEASONS
    for replay_date in build_trading_dates(season_start, season_end)
]

TODAY_NY = datetime.now(ZoneInfo("America/New_York")).date()
LAST_REPLAY_DATE = TODAY_NY - timedelta(days=1)
DATES_DST = [
    replay_date
    for replay_date in ALL_DATES_DST
    if datetime.strptime(replay_date, "%d/%m/%Y").date() <= LAST_REPLAY_DATE
]

# Replay recomendado para esta prueba: X10.
# Ventana por dia: 09:30 a 09:50 NY. El exporter escribe TIME_OVER si no hay trade antes/de 09:40;
# si ya hay trade abierto, dejamos correr hasta 09:50 para que resuelva TP/SL/EXIT/BE.
REPLAY_END_TIME = "09:50"
POLL_SECONDS = 0.02
NO_TRADE_CUTOFF_SECONDS = 10 * 60 + 15
HOLIDAY_RETRY_COUNT = 3
HOLIDAY_NORMAL_WAIT_SECONDS = 2 * 60
HOLIDAY_FINAL_WAIT_SECONDS = 3 * 60
HOLIDAY_NO_DATA_LABEL = "HOLYDAY NO DATA"
REQUIRED_TIMELINE_VERSION = replay_sync.EXPECTED_TIMELINE_VERSION

EXPORT_FOLDER = r"C:\Users\k_99_\Desktop\codding\data_footprint_generator"
RESULTS_FOLDER = os.path.join(EXPORT_FOLDER, "trade_results_score")
TIMELINE_FOLDER = os.path.join(RESULTS_FOLDER, "dynamic_management_timeline")
TARGET_FILE = os.path.join(EXPORT_FOLDER, "target_trade_result_date.txt")
REPLAY_STARTED_FILE = os.path.join(EXPORT_FOLDER, "replay_trade_result_started_at.txt")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCORE_WORKBOOK_TEMPLATE = os.path.join(BASE_DIR, "Score_indicator_results_updated.xlsx")
SCORE_WORKBOOK_TEMPLATE_FALLBACK = os.path.join(BASE_DIR, "Score_indicator_results_updated_fallback.xlsx")
SCORE_WORKBOOK = os.path.join(BASE_DIR, "Score_indicator_results_updated_2025_2026.xlsx")
SCORE_WORKBOOK_FALLBACK = os.path.join(BASE_DIR, "Score_indicator_results_updated_2025_2026_fallback.xlsx")
RUN_STARTED_AT = time.time()
RESUME_EXISTING_RESULTS = True
STALE_RESULT_BACKUP_DIR = os.path.join(RESULTS_FOLDER, "_replay_result_backups")
STALE_TIMELINE_BACKUP_DIR = os.path.join(TIMELINE_FOLDER, "_replay_timeline_backups")
EXCLUDED_EXCEL_HEADERS = {"Contracts"}


# =========================================================
# FUNCIONES
# =========================================================

def write_target_date(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    target = f"{yyyy}-{mm}-{dd}"

    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    os.makedirs(RESULTS_FOLDER, exist_ok=True)

    with open(TARGET_FILE, "w", encoding="utf-8") as f:
        f.write(target)

    print(f"Fecha objetivo escrita para ATAS: {target}")


def write_replay_started_marker():
    os.makedirs(EXPORT_FOLDER, exist_ok=True)
    with open(REPLAY_STARTED_FILE, "w", encoding="utf-8") as f:
        f.write(str(time.time()))
    print("Marcador de inicio de replay escrito.")


def get_replay():
    desktop = Desktop(backend="uia")
    candidates = desktop.windows(title_re=".*Replay.*", visible_only=True)

    if not candidates:
        raise RuntimeError("No encontre ninguna ventana Replay visible. Abre Replay en ATAS antes de correr el script.")

    print(f"Ventanas Replay encontradas: {len(candidates)}")

    for w in candidates:
        try:
            if w.is_visible() and w.is_enabled():
                print("Usando Replay:", w.window_text())
                w.set_focus()
                time.sleep(1)
                return w
        except Exception:
            pass

    replay = candidates[0]
    replay.set_focus()
    time.sleep(1)
    return replay


def paste_text(control, value):
    control.click_input()
    time.sleep(0.3)

    control.type_keys("^a")
    time.sleep(0.3)

    pyperclip.copy(value)
    time.sleep(0.3)

    control.type_keys("^v")
    time.sleep(0.8)


def get_controls():
    replay = get_replay()

    edits = replay.descendants(control_type="Edit")
    buttons = replay.descendants(control_type="Button")

    from_box = edits[0]
    to_box = edits[2]

    start_button = None
    stop_button = None

    for b in buttons:
        txt = b.window_text()

        if txt == "Start":
            start_button = b

        if txt == "Stop":
            stop_button = b

    return replay, from_box, to_box, start_button, stop_button


def expected_result_path(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")

    return os.path.join(
        RESULTS_FOLDER,
        f"score_trade_result_{yyyy}-{mm}-{dd}_NY.csv"
    )


def expected_timeline_path(date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")

    return os.path.join(
        TIMELINE_FOLDER,
        f"dynamic_timeline_{yyyy}-{mm}-{dd}_NY.csv"
    )


def print_result_file(path):
    if not os.path.exists(path):
        print("WARNING: no encontre archivo esperado:")
        print(path)
        return

    size_kb = round(os.path.getsize(path) / 1024, 2)
    print(f"OK EXPORTADO: {path}")
    print(f"Tamano: {size_kb} KB")

    with open(path, "r", encoding="utf-8") as f:
        print(f.read().strip())


def backup_previous_result(path, reason):
    if not os.path.exists(path):
        return

    os.makedirs(STALE_RESULT_BACKUP_DIR, exist_ok=True)
    base_name = os.path.basename(path)
    name, ext = os.path.splitext(base_name)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(STALE_RESULT_BACKUP_DIR, f"{name}_{reason}_{timestamp}{ext}")
    os.replace(path, backup_path)
    print(f"Resultado previo no terminal movido a backup: {backup_path}")


def backup_previous_timeline(path, reason):
    if not os.path.exists(path):
        return

    os.makedirs(STALE_TIMELINE_BACKUP_DIR, exist_ok=True)
    base_name = os.path.basename(path)
    name, ext = os.path.splitext(base_name)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(
        STALE_TIMELINE_BACKUP_DIR,
        f"{name}_{reason}_{timestamp}{ext}"
    )
    os.replace(path, backup_path)
    print(f"Timeline previo movido a backup: {backup_path}")


def clear_previous_result(path, force=False):
    if not os.path.exists(path):
        return False

    if RESUME_EXISTING_RESULTS and result_is_terminal(path) and not force:
        print(f"Resultado terminal existente conservado: {path}")
        return False

    backup_previous_result(path, "stale")
    return True


def clear_previous_timeline(path):
    if not os.path.exists(path):
        return False

    backup_previous_timeline(path, "stale")
    return True


def clear_expected_results():
    for date in DATES_DST:
        result_path = expected_result_path(date)
        timeline_path = expected_timeline_path(date)

        if date_has_complete_artifacts(result_path, timeline_path):
            continue

        # Preserve terminal historical results until their date is actually
        # replayed. This keeps a stopped run resumable without moving every
        # future result to backup at startup.
        if result_is_terminal(result_path):
            continue

        clear_previous_result(result_path)
        clear_previous_timeline(timeline_path)


def parse_result_ticks(value):
    if value in (None, "", "OPEN", "NO_TRADE"):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    normalized = str(value).strip().upper()

    if normalized in ("TP", "EXIT"):
        return 1.0

    if normalized == "SL":
        return -1.0

    if normalized == "BE":
        return 0.0

    try:
        return float(normalized.replace("+", ""))
    except ValueError:
        return None


def read_result_ticks(path):
    if not os.path.exists(path):
        return None

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return None

    if not row:
        return None

    value = row.get("result TP SL BE") or row.get("RESULT")
    return parse_result_ticks(value)


def result_is_terminal(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    if result_label in ("TP", "SL", "EXIT", "BE", "TIME_OVER", "NO_TRADE", HOLIDAY_NO_DATA_LABEL):
        return True

    ticks = parse_result_ticks(row.get("result TP SL BE") or row.get("RESULT"))
    if ticks is None:
        return False

    return ticks != 0


def result_requires_timeline(path):
    if not os.path.exists(path):
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    return result_label in ("TP", "SL", "EXIT", "BE")


def timeline_is_terminal(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            first_row = None
            last_row = None
            for row in reader:
                if first_row is None:
                    first_row = row
                last_row = row
    except (OSError, PermissionError):
        return False

    if not first_row or not last_row:
        return False

    if str(first_row.get("Timeline_VERSION") or "").strip() != REQUIRED_TIMELINE_VERSION:
        return False

    result_label = str(last_row.get("Result") or "").strip().upper()
    event_name = str(last_row.get("Event") or "").strip().upper()
    return (
        result_label in ("TP", "SL", "EXIT", "BE") and
        event_name.startswith("EXIT_")
    )


def date_has_complete_artifacts(result_path, timeline_path):
    if not result_is_terminal(result_path):
        return False

    if not result_requires_timeline(result_path):
        return True

    return timeline_is_terminal(timeline_path)


def result_is_open_trade(path, min_modified_time=None):
    if not os.path.exists(path):
        return False

    if min_modified_time is not None and os.path.getmtime(path) < min_modified_time:
        return False

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            row = next(reader, None)
    except (OSError, PermissionError):
        return False

    if not row:
        return False

    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()
    return result_label == "OPEN"


def write_holiday_no_data_result(path, date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    os.makedirs(os.path.dirname(path), exist_ok=True)

    headers = [
        "Exporter_VERSION",
        "fecha",
        "Result_Label",
        "result TP SL BE",
        "Signal_Source",
    ]
    row = [
        "python-replay-holiday-no-data",
        f"{yyyy}-{mm}-{dd}",
        HOLIDAY_NO_DATA_LABEL,
        HOLIDAY_NO_DATA_LABEL,
        "NO_DATA",
    ]

    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerow(row)

    print(f"CSV marcado como {HOLIDAY_NO_DATA_LABEL}: {path}")


def stop_replay(stop_button=None):
    print("Deteniendo replay...")

    candidates = [stop_button] if stop_button is not None else []

    for _ in range(5):
        try:
            replay, from_box, to_box, start_button, refreshed_stop_button = get_controls()
            replay.set_focus()
            if refreshed_stop_button is not None:
                candidates.insert(0, refreshed_stop_button)
        except Exception:
            pass

        for candidate in candidates:
            if candidate is None:
                continue

            try:
                if candidate.is_visible() and candidate.is_enabled():
                    candidate.click_input()
                    return
            except Exception:
                try:
                    candidate.click()
                    return
                except Exception:
                    pass

        time.sleep(0.05)

    print("No encontre boton Stop; probablemente el replay ya termino.")


MAX_WAIT_SECONDS = 1200

def format_countdown(seconds):
    seconds = max(0, int(seconds))
    minutes, seconds = divmod(seconds, 60)
    return f"{minutes:02d}:{seconds:02d}"


def wait_until_result(path, min_modified_time=None, no_trade_cutoff_seconds=None, stop_button=None):
    print("Esperando resultado terminal en CSV; si el trade esta OPEN no se cambia de dia.")
    start = time.time()
    last_print_second = -1
    last_status_line = ""

    while True:
        if result_is_terminal(path, min_modified_time):
            print("\r" + " " * max(len(last_status_line), 80), end="\r", flush=True)
            print("Esperando... listo.")
            print("Resultado terminal detectado en CSV; paso al siguiente dia.")
            stop_replay(stop_button)
            return True

        elapsed = time.time() - start
        has_open_trade = result_is_open_trade(path, min_modified_time)
        countdown_limit = no_trade_cutoff_seconds if no_trade_cutoff_seconds is not None else MAX_WAIT_SECONDS
        remaining = countdown_limit - elapsed

        if (
            no_trade_cutoff_seconds is not None and
            elapsed > no_trade_cutoff_seconds and
            not has_open_trade
        ):
            print("\r" + " " * max(len(last_status_line), 80), end="\r", flush=True)
            print(f"Corte 09:40 sin trade OPEN ({int(elapsed)}s). Deteniendo replay del dia.")
            stop_replay(stop_button)
            return False

        if elapsed > MAX_WAIT_SECONDS:
            print("\r" + " " * max(len(last_status_line), 80), end="\r", flush=True)
            print(f"Timeout ({MAX_WAIT_SECONDS}s) esperando resultado terminal. Saltando al siguiente dia.")
            stop_replay(stop_button)
            return False

        elapsed_second = int(elapsed)
        if elapsed_second != last_print_second:
            open_status = "OPEN" if has_open_trade else "sin OPEN"
            last_status_line = (
                f"Esperando resultado... transcurrido {format_countdown(elapsed)} | "
                f"restan {format_countdown(remaining)} | {open_status}"
            )
            print("\r" + last_status_line + " " * max(0, 100 - len(last_status_line)), end="", flush=True)
            last_print_second = elapsed_second

        time.sleep(POLL_SECONDS)


def read_trade_result(path, date_ddmmyyyy):
    dd, mm, yyyy = date_ddmmyyyy.split("/")
    default_row = {
        "fecha": f"{yyyy}-{mm}-{dd}",
        "result TP SL BE": "NO_CSV",
    }

    if not os.path.exists(path):
        print(f"Sin CSV para {default_row['fecha']}; no se sobrescriben columnas de datos.")
        return default_row

    if os.path.getmtime(path) < RUN_STARTED_AT:
        if not result_is_terminal(path):
            print(f"CSV viejo no terminal ignorado para {default_row['fecha']}: {path}")
            return default_row
        print(f"CSV terminal preservado usado para {default_row['fecha']}: {path}")

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        try:
            row = next(reader)
        except StopIteration:
            default_row["result TP SL BE"] = "EMPTY_CSV"
            return default_row

    if "fecha" not in row:
        legacy_result = row.get("RESULT", "NO_TRADE")
        legacy_ticks = None

        if legacy_result == "TP" and row.get("ENTRY") and row.get("TP"):
            legacy_ticks = abs((float(row["TP"]) - float(row["ENTRY"])) / 0.25)

        if legacy_result == "SL" and row.get("ENTRY") and row.get("SL"):
            legacy_ticks = -abs((float(row["ENTRY"]) - float(row["SL"])) / 0.25)

        row = {
            "fecha": default_row["fecha"],
            "SL_price": row.get("SL"),
            "Entry_price": row.get("ENTRY"),
            "TP_price": row.get("TP"),
            "result TP SL BE": legacy_ticks if legacy_ticks is not None else legacy_result,
        }

    row["fecha"] = row.get("fecha") or default_row["fecha"]
    row["result TP SL BE"] = row.get("result TP SL BE") or "OPEN"
    row["Exit_price"] = row.get("Exit_price") or calculate_exit_price(row)
    return row


def calculate_exit_price(row):
    result_label = str(row.get("Result_Label") or row.get("RESULT") or "").strip().upper()

    if result_label == "TP":
        return row.get("TP_price") or row.get("TP") or ""

    if result_label == "SL":
        return row.get("SL_price") or row.get("SL") or ""

    if result_label != "EXIT":
        return ""

    return ""


def to_number(value):
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def get_or_create_headers(ws):
    for col in range(ws.max_column, 0, -1):
        if ws.cell(row=3, column=col).value in EXCLUDED_EXCEL_HEADERS:
            ws.delete_cols(col)

    default_headers = [
        "fecha",
        "or_low",
        "or_high",
        "range",
        "VWAP_entry",
        "Body",
        "Volume_entry",
        "Delta_entry",
        "BreakOut_SPEED",
        "BreakOut_TICKS_PER_SEC",
        "score total",
        "SL_price",
        "Entry_price",
        "TP_price",
        "SL_ticks",
        "TP_ticks",
        "Exit_price",
        "result TP SL BE",
        "Side",
        "Speed_Profile",
        "MAE_ticks",
        "MFE_ticks",
    ]

    headers = [
        ws.cell(row=3, column=col).value
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=3, column=col).value and ws.cell(row=3, column=col).value not in EXCLUDED_EXCEL_HEADERS
    ]

    if not headers:
        headers = default_headers[:]

    for header in default_headers:
        if header not in headers:
            headers.append(header)

    for csv_header in get_csv_headers_for_dates():
        if csv_header not in headers:
            headers.append(csv_header)

    for header in ("ExitTime_NY", "Trade_Duration"):
        if header in headers:
            headers.remove(header)

    entry_time_index = headers.index("EntryTime_NY") + 1
    headers[entry_time_index:entry_time_index] = ["ExitTime_NY", "Trade_Duration"]

    trade_path_headers = [
        "Largest_MAE_pullback_ticks",
        "Largest_MFE_pullup_ticks",
        "Number_of_Pullbacks_during_Trade",
        "Number_of_PullUps_during_Trade",
        "Max_Speed_MAE_during_trade",
        "Max_Speed_MFE_during_trade",
    ]
    for header in trade_path_headers:
        if header in headers:
            headers.remove(header)

    mfe_index = headers.index("MFE_ticks") + 1
    headers[mfe_index:mfe_index] = trade_path_headers

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    return headers


def get_csv_headers_for_dates():
    headers = []

    for date in DATES_DST:
        path = expected_result_path(date)

        if not os.path.exists(path):
            continue

        try:
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    continue

                for field in reader.fieldnames:
                    if field and field not in EXCLUDED_EXCEL_HEADERS and field not in headers:
                        headers.append(field)
        except OSError:
            continue

    return headers


def update_score_workbook():
    os.makedirs(os.path.dirname(SCORE_WORKBOOK), exist_ok=True)

    if os.path.exists(SCORE_WORKBOOK):
        os.remove(SCORE_WORKBOOK)
        print(f"Excel anterior eliminado: {SCORE_WORKBOOK}")

    if os.path.exists(SCORE_WORKBOOK_TEMPLATE):
        wb = load_workbook(SCORE_WORKBOOK_TEMPLATE)
    elif os.path.exists(SCORE_WORKBOOK_FALLBACK):
        wb = load_workbook(SCORE_WORKBOOK_FALLBACK)
    elif os.path.exists(SCORE_WORKBOOK_TEMPLATE_FALLBACK):
        wb = load_workbook(SCORE_WORKBOOK_TEMPLATE_FALLBACK)
    else:
        wb = Workbook()

    ws = wb.active

    headers = get_or_create_headers(ws)
    first_row = 4
    last_row = first_row + len(DATES_DST) - 1

    for row in range(first_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None

    for row_offset, date in enumerate(DATES_DST, start=first_row):
        result = read_trade_result(expected_result_path(date), date)
        missing_result_file = result.get("result TP SL BE") in ("NO_CSV", "EMPTY_CSV")

        for col, header in enumerate(headers, start=1):
            if header not in result:
                if missing_result_file:
                    ws.cell(row=row_offset, column=col).value = None
                continue

            value = to_number(result.get(header))

            if value is None and header != "result TP SL BE":
                continue

            ws.cell(row=row_offset, column=col).value = value

    result_col = headers.index("result TP SL BE") + 1
    result_letter = get_column_letter(result_col)

    ws["A1"] = "win rate"
    ws["B1"] = "profit factor"
    ws["A2"] = f'=IF((COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0"))=0,"",COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")/(COUNTIF({result_letter}{first_row}:{result_letter}{last_row},">0")+COUNTIF({result_letter}{first_row}:{result_letter}{last_row},"<0")))'
    ws["B2"] = f'=IF(ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row}))=0,IF(SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})>0,"INF",""),SUMIF({result_letter}{first_row}:{result_letter}{last_row},">0",{result_letter}{first_row}:{result_letter}{last_row})/ABS(SUMIF({result_letter}{first_row}:{result_letter}{last_row},"<0",{result_letter}{first_row}:{result_letter}{last_row})))'
    ws["A2"].number_format = "0.00%"
    ws["B2"].number_format = "0.00"

    for row_idx in range(first_row, last_row + 1):
        ws.cell(row=row_idx, column=result_col).number_format = "+0.##;-0.##;0"

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header in ("ExitTime_NY", "Trade_Duration"):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    widths = {
        "A": 12,
        "B": 12,
        "C": 12,
        "D": 10,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 16,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col_idx, header in enumerate(headers, start=1):
        if header == "EntryTime_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header in ("ExitTime_NY", "Trade_Duration"):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

        if header == "EntrySecond_NY":
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
            for row_idx in range(first_row, last_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0"

    try:
        wb.save(SCORE_WORKBOOK)
        print(f"Excel actualizado: {SCORE_WORKBOOK}")
    except PermissionError:
        wb.save(SCORE_WORKBOOK_FALLBACK)
        print("WARNING: Excel original bloqueado/abierto; guarde copia actualizada en:")
        print(SCORE_WORKBOOK_FALLBACK)


# =========================================================
# LOOP PRINCIPAL V10 X1/X10
# =========================================================


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Corre las temporadas DST 2025-2026 usando el flujo v10 canónico: "
            "X1 genera la fila oficial y X10 la sincroniza."
        )
    )
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Compatibilidad con test_fechas_conflictivas.py; este runner ya usa modo quick por defecto.",
    )
    parser.add_argument(
        "--prepare-only",
        action="store_true",
        help="Muestra plan y versiones sin iniciar Replay.",
    )
    parser.add_argument(
        "--compare-only",
        action="store_true",
        help="Regenera el reporte con corridas guardadas.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Ignora corridas guardadas y vuelve a correr todo.",
    )
    parser.add_argument(
        "--step",
        action="store_true",
        help="Pide ENTER antes de cada fecha.",
    )
    parser.add_argument(
        "--x1-only",
        action="store_true",
        help="Solo corre la fase X1 canónica.",
    )
    parser.add_argument(
        "--x10-only",
        action="store_true",
        help="Solo corre la fase X10; requiere snapshots v10 previos de X1.",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    quick = True
    run_plan = replay_sync.build_run_plan(
        quick=quick,
        x1_only=args.x1_only,
        x10_only=args.x10_only,
    )
    date_iso_list = [replay_sync.date_iso_from_replay(date) for date in DATES_DST]
    output_folder = os.path.join(
        RESULTS_FOLDER,
        "visual_tests",
        "04_run_replay_score_trade_results_dst_2025_2026_runs",
    )

    print(
        f"\nINICIANDO REPLAY DE TEMPORADAS DST 2025-2026 V10 X1/X10 "
        f"({len(DATES_DST)} sesiones)\n"
        f"Fecha NY actual: {TODAY_NY:%d/%m/%Y} | "
        f"Ultima fecha permitida: {LAST_REPLAY_DATE:%d/%m/%Y}\n"
        f"Version esperada: {replay_sync.EXPECTED_EXPORTER_VERSION}\n"
        f"Resultados de validacion: {output_folder}\n"
    )
    print("Plan:")
    for run_name, speed_label, _ in run_plan:
        print(f"  - {run_name}: Replay {speed_label}")

    if args.prepare_only:
        print("\nPREPARE-ONLY correcto. No se inició Replay.")
        return 0

    if not args.compare_only:
        print("Ejecutando limpieza unica de Telegram antes de la primera fecha...")
        clear_telegram_before_run(RESULTS_FOLDER)

    passed, failures = replay_sync.run_replay_period(
        date_iso_list,
        output_folder=Path(output_folder),
        run_plan=run_plan,
        report_prefix="dst_2025_2026_v10",
        force=args.force,
        step=args.step,
        compare_only=args.compare_only,
        replay_to_time=REPLAY_END_TIME,
    )

    update_score_workbook()

    failed_dates = [
        (date_iso, f"{run_name}: {reason}")
        for date_iso, run_name, reason in failures
    ]
    if failed_dates:
        print("\nFECHAS CON ERROR DE SCRIPT/UI:")
        for failed_date, error in failed_dates:
            print(f"- {failed_date}: {error}")

    send_run_summary(
        RESULTS_FOLDER,
        DATES_DST,
        failed_dates,
        "Temporadas DST completas 2025-2026 v10 X1/X10",
    )

    print("\nTERMINO LA PRUEBA DE TEMPORADAS DST COMPLETAS 2025-2026 V10.\n")
    return 0 if passed and not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
