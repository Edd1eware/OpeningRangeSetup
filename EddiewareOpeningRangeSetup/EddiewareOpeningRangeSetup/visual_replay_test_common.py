import csv
import os
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


EXPORT_FOLDER = Path(r"C:\Users\k_99_\Desktop\codding\data_footprint_generator")
RESULTS_FOLDER = EXPORT_FOLDER / "trade_results_score"
TIMELINE_FOLDER = RESULTS_FOLDER / "dynamic_management_timeline"
TARGET_FILE = EXPORT_FOLDER / "target_trade_result_date.txt"
VISUAL_FOLDER = RESULTS_FOLDER / "visual_tests"
BACKUP_FOLDER = VISUAL_FOLDER / "_artifact_backups"


def read_single_csv(path):
    with open(path, encoding="utf-8-sig", newline="") as handle:
        return next(csv.DictReader(handle), None)


def read_csv_rows(path):
    with open(path, encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def to_float(value):
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def discover_trade_rows():
    rows = []
    for path in sorted(RESULTS_FOLDER.glob("score_trade_result_*_NY.csv")):
        row = read_single_csv(path)
        if row and row.get("Result_Label", "").strip() in {"TP", "SL"}:
            rows.append(row)
    return rows


def timeline_path(date_iso):
    return TIMELINE_FOLDER / f"dynamic_timeline_{date_iso}_NY.csv"


def create_visual_workbook(test_name, rows, result_options, description):
    VISUAL_FOLDER.mkdir(parents=True, exist_ok=True)
    output_path = VISUAL_FOLDER / f"{test_name}.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Validacion"
    instructions = workbook.create_sheet("Instrucciones")

    headers = list(rows[0].keys()) if rows else ["Fecha"]
    headers.extend(["Resultado_visual_que_vi", "Comentarios"])
    sheet.append(headers)

    for row in rows:
        sheet.append([row.get(header, "") for header in headers[:-2]] + ["", ""])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    good_fill = PatternFill("solid", fgColor="E2F0D9")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    result_column = len(headers) - 1
    for row_number in range(2, max(2, sheet.max_row + 1)):
        sheet.cell(row_number, result_column).fill = input_fill
        sheet.cell(row_number, result_column + 1).fill = input_fill

    if rows and result_options:
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(result_options) + '"',
            allow_blank=True,
        )
        sheet.add_data_validation(validation)
        validation.add(
            f"{sheet.cell(2, result_column).coordinate}:"
            f"{sheet.cell(sheet.max_row, result_column).coordinate}"
        )

    if rows:
        sheet.conditional_formatting.add(
            f"{sheet.cell(2, result_column).coordinate}:"
            f"{sheet.cell(sheet.max_row, result_column).coordinate}",
            FormulaRule(
                formula=[f'LEN({sheet.cell(2, result_column).coordinate})>0'],
                fill=good_fill,
            ),
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 42

    for index, header in enumerate(headers, 1):
        width = max(13, min(38, len(header) + 3))
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width

    instructions["A1"] = test_name
    instructions["A1"].font = Font(size=16, bold=True, color="1F4E78")
    instructions["A3"] = description
    instructions["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions["A5"] = (
        "Usa Replay X1 para inspección visual. En cada fila escribe lo observado "
        "en la celda amarilla Resultado_visual_que_vi. El script conserva y restaura "
        "los CSV/timelines originales después de cada fecha."
    )
    instructions["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions.column_dimensions["A"].width = 110
    instructions.row_dimensions[3].height = 80
    instructions.row_dimensions[5].height = 80

    workbook.save(output_path)
    return output_path


def get_replay_controls():
    from pywinauto import Desktop

    desktop = Desktop(backend="uia")
    candidates = desktop.windows(title_re=".*Replay.*", visible_only=True)
    if not candidates:
        raise RuntimeError("Abre una ventana Replay visible en ATAS.")

    replay = candidates[0]
    replay.set_focus()
    edits = replay.descendants(control_type="Edit")
    buttons = replay.descendants(control_type="Button")

    if len(edits) < 2:
        raise RuntimeError("No encontré los campos FROM/TO del Replay.")

    start_button = next(
        (button for button in buttons if "start" in button.window_text().lower()),
        None,
    )
    stop_button = next(
        (button for button in buttons if "stop" in button.window_text().lower()),
        None,
    )
    if start_button is None:
        raise RuntimeError("No encontré el botón Start del Replay.")

    return replay, edits[0], edits[1], start_button, stop_button


def paste_text(control, value):
    import pyperclip

    control.set_focus()
    control.type_keys("^a")
    pyperclip.copy(value)
    control.type_keys("^v")


def stop_replay(stop_button=None):
    try:
        if stop_button is None:
            _, _, _, _, stop_button = get_replay_controls()
        if stop_button is not None:
            stop_button.click_input()
            time.sleep(1)
    except Exception as exc:
        print(f"WARNING: no pude detener Replay automáticamente: {exc}")


def artifact_paths(date_iso):
    return (
        RESULTS_FOLDER / f"score_trade_result_{date_iso}_NY.csv",
        TIMELINE_FOLDER / f"dynamic_timeline_{date_iso}_NY.csv",
    )


def backup_artifacts(date_iso):
    BACKUP_FOLDER.mkdir(parents=True, exist_ok=True)
    saved = []
    for source in artifact_paths(date_iso):
        backup = BACKUP_FOLDER / f"{date_iso}_{source.name}"
        existed = source.exists()
        if existed:
            shutil.copy2(source, backup)
        saved.append((source, backup, existed))
    return saved


def restore_artifacts(saved):
    for destination, backup, existed in saved:
        if existed and backup.exists():
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(backup, destination)
            backup.unlink()
        elif not existed and destination.exists():
            destination.unlink()


def write_target_date(date_iso):
    EXPORT_FOLDER.mkdir(parents=True, exist_ok=True)
    TARGET_FILE.write_text(date_iso, encoding="utf-8")


def run_visual_dates(test_name, rows, result_options, description):
    if not rows:
        print("No se encontraron fechas candidatas.")
        return

    workbook_path = create_visual_workbook(
        test_name,
        rows,
        result_options,
        description,
    )
    print(f"Excel de validación: {workbook_path}")

    if "--prepare-only" in sys.argv:
        return

    os.startfile(workbook_path)
    print("Configura Replay en X1 antes de comenzar.")

    for index, row in enumerate(rows, 1):
        date_iso = row["Fecha"]
        date_value = datetime.strptime(date_iso, "%Y-%m-%d")
        date_replay = date_value.strftime("%d/%m/%Y")
        saved = backup_artifacts(date_iso)

        try:
            input(
                f"\n[{index}/{len(rows)}] {date_iso}. "
                "Presiona ENTER para configurar e iniciar Replay..."
            )
            write_target_date(date_iso)
            replay, from_box, to_box, start_button, stop_button = get_replay_controls()
            paste_text(from_box, f"{date_replay} 09:30 a. m.")
            paste_text(to_box, f"{date_replay} 09:41 a. m.")
            time.sleep(1)
            replay.set_focus()
            start_button.click_input()
            input(
                "Observa ATAS, llena Resultado_visual_que_vi en el Excel "
                "y presiona ENTER para detener y continuar..."
            )
            stop_replay(stop_button)
        finally:
            restore_artifacts(saved)

    print(f"\nPrueba terminada. Guarda tus observaciones en:\n{workbook_path}")
